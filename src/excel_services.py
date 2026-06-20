"""
Excel background services for the Sub-Process Traceability application.
Handles background queuing of Excel file writes and formatting logic.
"""
import os
import queue
import threading
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from config import EXCEL_FILE
from database import get_db_connection, logger

# ── Excel Background Worker ───────────────────────────────────────────────────

_excel_queue = queue.Queue()
excel_error_state = False

def _excel_worker():
    import time
    while True:
        func, args, kwargs = _excel_queue.get()
        while True:
            try:
                func(*args, **kwargs)
                global excel_error_state
                excel_error_state = False
                break
            except PermissionError:
                excel_error_state = True
                logger.warning("Excel file is open. Retrying background save in 5 seconds...")
                time.sleep(5)
            except Exception as e:
                logger.error("Excel worker error: %s", e)
                break
        _excel_queue.task_done()

_excel_thread = threading.Thread(target=_excel_worker, daemon=True)
_excel_thread.start()

def queue_excel_task(func, *args, **kwargs):
    """Adds a task to the background Excel writer queue."""
    _excel_queue.put((func, args, kwargs))

# ── Excel File Generation ─────────────────────────────────────────────────────

def rebuild_excel_from_db():
  if os.path.exists(EXCEL_FILE):
    try:
      os.remove(EXCEL_FILE)
    except PermissionError:
      raise
    except Exception:
      pass
      
  wb = Workbook()
  ws = wb.active
  ws.title = "Sub-process fill by TL"
  
  headers = [
    "SB ID", "FULL PN Semi fini", "PART NAME (SF)", 
    "RM PN", "RM Name", "Batch No. 1", "Batch No. 2", "Batch No. 3", "Quantity",
    "Work in Sub-Process by Shift", "Op ID", "Station",
    "Sub-Process Work Date/Time", "Production Line Entry Date/Time",
    "Work in PROD line by Shift", "Remarks", "Registered by ID"
  ]
  ws.append(headers)
  ws.freeze_panes = "A2"
  
  header_font = Font(name='Calibri', size=10, bold=True, color="000000")
  header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
  header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
  thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

  for col_idx in range(1, 18):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    
  ws.row_dimensions[1].height = 34
  widths = [20, 22, 25, 22, 25, 12, 12, 12, 10, 15, 10, 10, 20, 20, 15, 25, 15]
  from openpyxl.utils import get_column_letter
  for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

  conn = get_db_connection()
  conn.row_factory = sqlite3.Row
  c = conn.cursor()
  c.execute("SELECT * FROM records ORDER BY id ASC")
  rows = c.fetchall()
  conn.close()
  
  batch_tracker = {}
  
  for row in rows:
    change_counts = []
    rm_pns = [row["rm1_pn"], row["rm2_pn"], row["rm3_pn"], row["rm4_pn"]]
    batch4_val = row["batch4"] if "batch4" in row.keys() else ""
    batches = [row["batch1"], row["batch2"], row["batch3"], batch4_val]
    
    for i in range(4):
      rm_pn = rm_pns[i]
      batch = str(batches[i]).strip() if batches[i] else ""
      if rm_pn and batch:
        if rm_pn not in batch_tracker:
          batch_tracker[rm_pn] = {"current": batch, "count": 0}
        elif batch_tracker[rm_pn]["current"] != batch:
          batch_tracker[rm_pn]["count"] += 1
          batch_tracker[rm_pn]["current"] = batch
        change_counts.append(batch_tracker[rm_pn]["count"])
      else:
        change_counts.append(0)
        
    data = [
      row["sub_batch_id"], row["pn_sf"], row["part_sf"],
      row["rm1_pn"], row["rm1_name"], row["rm2_pn"], row["rm2_name"],
      row["rm3_pn"], row["rm3_name"], row["rm4_pn"], row["rm4_name"],
      row["batch1"], row["batch2"], row["batch3"], batch4_val,
      row["quantity"], row["shift_sp"], row["op_id"], row["station"], row["dt_sp"], row["dt_line"],
      row["shift_line"], row["remarks"], row["registered_by"]
    ]
    _write_record_to_ws(ws, data, change_counts)
    wb.save(EXCEL_FILE)

def save_to_excel(data):
  if not os.path.exists(EXCEL_FILE):
    rebuild_excel_from_db()
    return

  try:
    wb = load_workbook(EXCEL_FILE)
    if "Sub-process fill by TL" in wb.sheetnames:
      ws = wb["Sub-process fill by TL"]
    else:
      rebuild_excel_from_db()
      return
      
    change_counts = []
    conn = get_db_connection()
    c = conn.cursor()
    for i in range(4):
      rm_pn = data[3 + i*2]
      if not rm_pn:
        change_counts.append(0)
        continue
      
      try:
        c.execute("""
          SELECT 
            CASE 
              WHEN rm1_pn = ? THEN batch1
              WHEN rm2_pn = ? THEN batch2
              WHEN rm3_pn = ? THEN batch3
              WHEN rm4_pn = ? THEN batch4
            END
          FROM records 
          WHERE ? IN (rm1_pn, rm2_pn, rm3_pn, rm4_pn)
          ORDER BY id ASC
        """, (rm_pn, rm_pn, rm_pn, rm_pn, rm_pn))
        
        hist_rows = c.fetchall()
        count = 0
        current = None
        for r in hist_rows:
          b = str(r[0]).strip() if r[0] else ""
          if not b: continue
          if current is None:
            current = b
          elif b != current:
            count += 1
            current = b
        change_counts.append(count)
      except Exception as e:
        logger.error("Error calculating batch change: %s", e)
        change_counts.append(0)
        
    conn.close()
    
    _write_record_to_ws(ws, data, change_counts)
    wb.save(EXCEL_FILE)
  except PermissionError:
    raise
  except Exception as e:
    logger.error("Error loading Excel file: %s", e)
    return

def _write_record_to_ws(ws, data, change_counts):
  sb_id = data[0] if data[0] else "-"
  sf_pn = data[1] if data[1] else "-"
  part_sf = data[2] if data[2] else "-"
  rms = []
  for i in range(4):
    rm_pn = data[3 + i*2]
    rm_name = data[4 + i*2]
    batch_no = str(data[11 + i]).strip() if data[11 + i] else ""
    c_count = change_counts[i] if i < len(change_counts) else 0
    b1, b2, b3 = "-", "-", "-"
    
    if batch_no:
      if c_count == 0:
        b1 = batch_no
      elif c_count == 1:
        b2 = batch_no
      else:
        b3 = batch_no
        
    if rm_pn:
      rms.append((rm_pn if rm_pn else "-", rm_name if rm_name else "-", b1, b2, b3))
      
  if not rms:
    rms = [("-", "-", "-", "-", "-")]
    
  num_rms = len(rms)
  rest_data = ["-" if (x == "" or x is None) else x for x in data[15:]]
  
  start_row = ws.max_row + 1
  end_row = start_row + num_rms - 1
  
  for i in range(num_rms):
    row_data = []
    if i == 0:
      row_data.extend([sb_id, sf_pn, part_sf])
      row_data.extend([rms[i][0], rms[i][1], rms[i][2], rms[i][3], rms[i][4]])
      row_data.extend(rest_data)
    else:
      row_data.extend([None, None, None])
      row_data.extend([rms[i][0], rms[i][1], rms[i][2], rms[i][3], rms[i][4]])
      row_data.extend([None] * len(rest_data))
    ws.append(row_data)
    
  if num_rms > 1:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
    ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
    for col in range(9, 18):
      ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
  
  sf_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
  rm_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
  rest_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
  row_font = Font(name='Calibri', size=10, color="000000")
  align = Alignment(horizontal="center", vertical="center", wrap_text=True)
  thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
  
  for r in range(start_row, end_row + 1):
    for c in range(1, 18):
      cell = ws.cell(row=r, column=c)
      cell.font = row_font
      cell.alignment = align
      cell.border = thin_border
      if c in (1, 2, 3):
        cell.fill = sf_fill
      elif c in (4, 5):
        cell.fill = rm_fill
      else:
        cell.fill = rest_fill
