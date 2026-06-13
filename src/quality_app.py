import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
import datetime
import os
import sys
import threading
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import matplotlib
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import hashlib
import binascii

def resource_path(relative_path):
  try:
    base_path = sys._MEIPASS
  except Exception:
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
  return os.path.join(base_path, relative_path)

def persistent_path(relative_path):
  if hasattr(sys, '_MEIPASS'):
    base_path = os.path.dirname(sys.executable)
  else:
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
  return os.path.join(base_path, relative_path)

DATA_DIR = persistent_path("data")
os.makedirs(DATA_DIR, exist_ok=True)
DB_FILE = os.path.join(DATA_DIR, "traceability.db")

def get_db_connection():
  conn = sqlite3.connect(DB_FILE, check_same_thread=False, timeout=10)
  conn.execute('PRAGMA journal_mode=WAL')
  return conn

def hash_password(password):
  salt = b"subproc_trace_salt_2026"
  hash_obj = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 100000)
  return binascii.hexlify(hash_obj).decode("utf-8")
EXCEL_FILE = os.path.join(DATA_DIR, "quality_defects.xlsx")

APP_VERSION = "1.0.0"

# ─────────────────────────────────────────────────────────────────────────────
# Theme (100% identical to main.py)
# ─────────────────────────────────────────────────────────────────────────────
BG_COLOR   = "#0F1419"
SURFACE_COLOR = "#1B232C"
ACCENT_COLOR = "#00B4D8"
SUCCESS_COLOR = "#2DC653"
WARNING_COLOR = "#F59E0B"
ERROR_COLOR  = "#DC2626"
BORDER_COLOR = "#334155"
TEXT_COLOR  = "#F8FAFC"
TEXT_MUTED  = "#94A3B8"
STATUS_IDLE  = "#64748B"

HEADER_BG = "#FFFFFF"
HEADER_FG = "#005A8C"
QUALITY_HL = "#F59E0B"  # amber highlight for quality-specific elements

HMI_FONT_XL = ("Segoe UI", 20, "bold")
HMI_FONT_L = ("Segoe UI", 16, "bold")
HMI_FONT_M = ("Segoe UI", 12, "bold")
HMI_FONT_S = ("Segoe UI", 10)
HMI_FONT_XS = ("Segoe UI", 9)

DEFECT_TYPES = [
  "Cosmetic",
  "Assembly Error",
  "Functional Failure",
  "Missing Part",
  "Dimension Out of Spec",
  "Other",
]

# ─────────────────────────────────────────────────────────────────────────────
# DB init
# ─────────────────────────────────────────────────────────────────────────────
def init_quality_db():
  conn = get_db_connection()
  c = conn.cursor()

  c.execute('''CREATE TABLE IF NOT EXISTS auth (
           id TEXT PRIMARY KEY,
           password TEXT,
           role TEXT)''')
  try:
    c.executemany(
      "INSERT OR IGNORE INTO auth (id, password, role) VALUES (?,?,?)",
      [('Q001', hash_password('quality001'), 'Quality OP')])
    conn.commit()
  except Exception as e:
    print("Auth seed error:", e)

  c.execute('''CREATE TABLE IF NOT EXISTS quality_defects (
           id       INTEGER PRIMARY KEY AUTOINCREMENT,
           sub_batch_id  TEXT NOT NULL,
           pn_sf      TEXT,
           part_sf     TEXT,
           produced_by_op TEXT,
           quality_op_id  TEXT NOT NULL,
           defect_type   TEXT,
           qty_defective  INTEGER DEFAULT 0,
           description   TEXT,
           shift_sp    TEXT,
           station     TEXT,
           reported_at   TEXT NOT NULL,
           status     TEXT DEFAULT 'Open')''')

  c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_quality_defects_sub_batch_id ON quality_defects(sub_batch_id)")

  c.execute('''CREATE TABLE IF NOT EXISTS system_access_logs (
           id INTEGER PRIMARY KEY AUTOINCREMENT,
           event_type TEXT,
           user_id TEXT,
           shift TEXT,
           timestamp TEXT)''')

  # Migrations
  migrations = [
    "ALTER TABLE quality_defects ADD COLUMN action_type TEXT DEFAULT 'Scrap'",
    "ALTER TABLE quality_defects ADD COLUMN is_quarantined INTEGER DEFAULT 0",
    "ALTER TABLE quality_defects ADD COLUMN root_cause TEXT",
    "ALTER TABLE quality_defects ADD COLUMN corrective_action TEXT"
  ]
  for mig in migrations:
    try:
      c.execute(mig)
    except sqlite3.OperationalError:
      pass # Column likely already exists

  conn.commit()
  conn.close()


# ─────────────────────────────────────────────────────────────────────────────
# Application
# ─────────────────────────────────────────────────────────────────────────────
class QualityApp(tk.Tk):
  def __init__(self):
    super().__init__()
    self.title(f"HI-LEX ACT — Quality Control v{APP_VERSION}")
    self.geometry("1280x800")
    self.configure(bg=BG_COLOR)
    
    png_path = persistent_path(os.path.join("assets", "analysis.png"))
    if os.path.exists(png_path):
      try:
        self.iconphoto(True, tk.PhotoImage(file=png_path))
      except Exception:
        pass
      
    try:
      self.state('zoomed')
    except Exception:
      pass
    self.protocol("WM_DELETE_WINDOW", self.on_closing)

    # Windows 11 white title bar
    try:
      import ctypes
      self.update_idletasks()
      hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
      ctypes.windll.dwmapi.DwmSetWindowAttribute(
        hwnd, 35, ctypes.byref(ctypes.c_int(0x00FFFFFF)), 4)
      ctypes.windll.dwmapi.DwmSetWindowAttribute(
        hwnd, 36, ctypes.byref(ctypes.c_int(0x008C5A00)), 4)
    except Exception:
      pass

    self.option_add('*TCombobox*Listbox.background', SURFACE_COLOR)
    self.option_add('*TCombobox*Listbox.foreground', TEXT_COLOR)
    self.option_add('*TCombobox*Listbox.selectBackground', ACCENT_COLOR)
    self.option_add('*TCombobox*Listbox.selectForeground', 'black')

    # Global Right-Click Menu for Entry widgets
    self.entry_menu = tk.Menu(self, tearoff=0, bg=SURFACE_COLOR, fg=TEXT_COLOR, activebackground=ACCENT_COLOR, activeforeground=TEXT_COLOR)
    self.entry_menu.add_command(label="Cut", command=lambda: self.focus_get().event_generate("<<Cut>>") if self.focus_get() else None)
    self.entry_menu.add_command(label="Copy", command=lambda: self.focus_get().event_generate("<<Copy>>") if self.focus_get() else None)
    self.entry_menu.add_command(label="Paste", command=lambda: self.focus_get().event_generate("<<Paste>>") if self.focus_get() else None)
    self.entry_menu.add_separator()
    self.entry_menu.add_command(label="Select All", command=lambda: self.focus_get().select_range(0, 'end') if hasattr(self.focus_get(), 'select_range') else None)

    def show_entry_menu(event):
      try:
        widget = event.widget
        if widget.winfo_class() in ('Entry', 'TEntry'):
          widget.focus()
          self.entry_menu.tk_popup(event.x_root, event.y_root)
      except Exception:
        pass

    self.bind_all("<Button-3>", show_entry_menu)

    # Session
    self.quality_op_id  = ""
    self.quality_op_role = ""

    self._apply_styles()
    self._build_ui()
    self.after(100, self._prompt_login)
    self._tick_clock()

  # ── Styles (mirrors main.py exactly) ────────────────────────────────────
  def _apply_styles(self):
    s = ttk.Style(self)
    s.theme_use('clam')
    s.configure('.', background=BG_COLOR, foreground=TEXT_COLOR,
          fieldbackground=SURFACE_COLOR)
    s.configure('TNotebook', background=BG_COLOR)
    s.configure('TNotebook.Tab', background=SURFACE_COLOR,
          foreground=TEXT_COLOR, padding=[20, 8],
          font=("Segoe UI", 11, "bold"))
    s.map('TNotebook.Tab', background=[('selected', ACCENT_COLOR)])
    s.configure('TFrame', background=BG_COLOR)
    s.configure('TLabel', background=SURFACE_COLOR, foreground=TEXT_COLOR)
    s.configure('TCheckbutton', background=SURFACE_COLOR, foreground=TEXT_COLOR)
    s.configure('TCombobox', fieldbackground=SURFACE_COLOR,
          background=SURFACE_COLOR, foreground=TEXT_COLOR)
    s.map('TCombobox',
       fieldbackground=[('readonly', SURFACE_COLOR)],
       selectbackground=[('readonly', ACCENT_COLOR)],
       selectforeground=[('readonly', '#000000')])
    s.configure('TEntry', fieldbackground=SURFACE_COLOR,
          foreground=TEXT_COLOR, insertcolor='white')
    s.configure('TButton', background=SURFACE_COLOR,
          foreground=TEXT_COLOR, font=HMI_FONT_M, padding=5)
    s.map('TButton', background=[('active', BORDER_COLOR)])

    s.configure('Success.TButton', background=SUCCESS_COLOR,
          foreground="#000000", font=HMI_FONT_M, padding=5)
    s.map('Success.TButton',
       background=[('disabled', '#4A5568'), ('active', '#16A34A')],
       foreground=[('disabled', '#A0AEC0')])

    s.configure('Primary.TButton', background=ACCENT_COLOR,
          foreground="#000000", font=HMI_FONT_M, padding=5)
    s.map('Primary.TButton',
       background=[('disabled', '#4A5568'), ('active', "#0096C7")],
       foreground=[('disabled', '#A0AEC0')])

    s.configure('Secondary.TButton', background=STATUS_IDLE,
          foreground="#FFFFFF", font=HMI_FONT_M, padding=5)
    s.map('Secondary.TButton',
       background=[('disabled', '#4A5568'), ('active', BORDER_COLOR)],
       foreground=[('disabled', '#A0AEC0')])

    s.configure('Header.TButton', background="#005A8C",
          foreground="#FFFFFF", font=HMI_FONT_M, padding=5)
    s.map('Header.TButton',
       background=[('disabled', '#4A5568'), ('active', "#00456B")],
       foreground=[('disabled', '#A0AEC0')])

    s.configure('Warning.TButton', background=WARNING_COLOR,
          foreground="#000000", font=HMI_FONT_M, padding=5)
    s.map('Warning.TButton',
       background=[('disabled', '#4A5568'), ('active', "#D97706")],
       foreground=[('disabled', '#A0AEC0')])

    s.configure('Danger.TButton', background=ERROR_COLOR,
          foreground="#000000", font=HMI_FONT_M, padding=5)
    s.map('Danger.TButton',
       background=[('disabled', '#4A5568'), ('active', "#B91C1C")],
       foreground=[('disabled', '#A0AEC0')])

    s.configure('Treeview', background=SURFACE_COLOR,
          foreground=TEXT_COLOR, fieldbackground=SURFACE_COLOR,
          rowheight=30, font=HMI_FONT_S)
    s.configure('Treeview.Heading', background=BORDER_COLOR,
          foreground=TEXT_COLOR, font=HMI_FONT_M)
    s.map('Treeview.Heading',
       background=[('active', ACCENT_COLOR)],
       foreground=[('active', '#000000')])

    s.configure("Safe.Horizontal.TProgressbar",
          background=SUCCESS_COLOR, troughcolor=BG_COLOR, thickness=8)
    s.configure("Warn.Horizontal.TProgressbar",
          background=WARNING_COLOR, troughcolor=BG_COLOR, thickness=8)
    s.configure("Danger.Horizontal.TProgressbar",
          background=ERROR_COLOR,  troughcolor=BG_COLOR, thickness=8)

  # ── Build UI (mirrors main.py layout) ──────────────────────────────────
  def _build_ui(self):
    # ── WHITE HEADER (identical to main app) ──
    self.header_frame = tk.Frame(self, bg=HEADER_BG, height=70)
    self.header_frame.pack(side=tk.TOP, fill=tk.X)
    self.header_frame.pack_propagate(False)

    # Logo (reuse main app logo if exists)
    try:
      from PIL import Image, ImageTk
      logo_path = resource_path(os.path.join("assets", "logo_en.png"))
      if os.path.exists(logo_path):
        img = Image.open(logo_path)
        ratio = 50 / img.height
        img = img.resize((int(img.width * ratio), 50),
                 Image.Resampling.LANCZOS
                 if hasattr(Image, "Resampling")
                 else Image.ANTIALIAS)
        self._logo_img = ImageTk.PhotoImage(img)
        tk.Label(self.header_frame, image=self._logo_img,
             bg=HEADER_BG).pack(side=tk.LEFT, padx=(20, 10))
    except Exception:
      pass

    tk.Label(self.header_frame,
         text="QUALITY CONTROL",
         bg=HEADER_BG, fg=HEADER_FG,
         font=HMI_FONT_L).pack(side=tk.LEFT)

    # Quality badge
    tk.Label(self.header_frame,
         text=" ⚠ Quality Dept.",
         bg=HEADER_BG, fg=QUALITY_HL,
         font=HMI_FONT_S).pack(side=tk.LEFT, padx=10)

    # Right side
    ttk.Button(self.header_frame, text="Logout",
          style="Danger.TButton",
          command=self._do_logout).pack(side=tk.RIGHT, padx=20)

    self.lbl_header_user = tk.Label(
      self.header_frame,
      text="User: — | Role: —",
      bg=HEADER_BG, fg=TEXT_MUTED, font=HMI_FONT_S)
    self.lbl_header_user.pack(side=tk.RIGHT, padx=10)

    # ── PANED WINDOW (sidebar + content) ──
    self.paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
    self.paned.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # ── SIDEBAR ──
    self.sidebar = tk.Frame(self.paned, bg=SURFACE_COLOR, width=190)
    self.sidebar.pack_propagate(False)
    self.paned.add(self.sidebar, weight=0)

    self.lbl_clock = tk.Label(self.sidebar, text="00:00",
                 bg=SURFACE_COLOR, fg=ACCENT_COLOR,
                 font=("Consolas", 28, "bold"))
    self.lbl_clock.pack(pady=(20, 0))
    self.lbl_date = tk.Label(self.sidebar, text="",
                 bg=SURFACE_COLOR, fg=TEXT_MUTED,
                 font=HMI_FONT_S)
    self.lbl_date.pack(pady=(0, 20))

    tk.Label(self.sidebar, text="QUALITY OVERVIEW",
         bg=SURFACE_COLOR, fg=QUALITY_HL,
         font=HMI_FONT_M).pack()

    def _make_kpi(title):
      f = tk.Frame(self.sidebar, bg=BG_COLOR,
             highlightbackground=BORDER_COLOR,
             highlightthickness=1)
      f.pack(fill=tk.X, padx=15, pady=5)
      tk.Label(f, text=title, bg=BG_COLOR,
           fg=TEXT_MUTED, font=HMI_FONT_S).pack(pady=(5, 0))
      lv = tk.Label(f, text="—", bg=BG_COLOR,
             fg=TEXT_COLOR, font=HMI_FONT_L)
      lv.pack(pady=(0, 5))
      return lv

    self.kpi_total = _make_kpi("Total Defects Today")
    self.kpi_rate  = _make_kpi("Overall Defect Rate %")
    self.kpi_open  = _make_kpi("Open Cases")

    tk.Frame(self.sidebar, bg=BORDER_COLOR, height=1).pack(
      fill=tk.X, padx=15, pady=10)

    ttk.Button(self.sidebar, text="Open Excel",
          style="Secondary.TButton",
          command=self._open_excel).pack(
            fill=tk.X, padx=10, pady=2)

    # ── NOTEBOOK ──
    self.notebook = ttk.Notebook(self.paned)
    self.paned.add(self.notebook, weight=1)

    self.tab_report = ttk.Frame(self.notebook)
    self.tab_log  = ttk.Frame(self.notebook)
    self.tab_analytics = ttk.Frame(self.notebook)

    self.notebook.add(self.tab_report, text="Report Defect")
    self.notebook.add(self.tab_log,  text="Quality Log")
    self.notebook.add(self.tab_analytics, text="Analytics")

    self._build_report_tab()
    self._build_log_tab()
    self._build_analytics_tab()

  # ── create_card (mirrors main.py) ───────────────────────────────────────
  def create_card(self, parent, title, fg_color=TEXT_COLOR):
    card_outer = tk.Frame(parent, bg=SURFACE_COLOR)
    card_outer.pack(fill=tk.X, pady=5, padx=5)
    hdr = tk.Frame(card_outer, bg=SURFACE_COLOR)
    hdr.pack(fill=tk.X)
    tk.Label(hdr, text=title, bg=SURFACE_COLOR,
         fg=fg_color, font=HMI_FONT_M).pack(
           side=tk.LEFT, padx=10, pady=6)
    tk.Frame(card_outer, bg=BORDER_COLOR, height=1).pack(fill=tk.X)
    content = tk.Frame(card_outer, bg=SURFACE_COLOR)
    content.pack(fill=tk.BOTH, expand=True, padx=8, pady=6)
    return card_outer, content

  # ── Login ────────────────────────────────────────────────────────────────
  def _prompt_login(self):
    win = tk.Toplevel(self)
    win.title("HI-LEX — Quality Login")
    win.geometry("400x300")
    win.configure(bg=BG_COLOR)
    win.transient(self)
    win.grab_set()
    win.resizable(False, False)

    win.update_idletasks()
    x = (win.winfo_screenwidth() // 2) - 200
    y = (win.winfo_screenheight() // 2) - 150
    win.geometry(f"+{x}+{y}")
    win.protocol("WM_DELETE_WINDOW",
           lambda: (win.destroy(), self.destroy()))

    tk.Label(win, text="Quality Control Login",
         font=HMI_FONT_L, bg=BG_COLOR,
         fg=TEXT_COLOR).pack(pady=15)

    frm = tk.Frame(win, bg=BG_COLOR)
    frm.pack(pady=10)

    tk.Label(frm, text="Quality OP ID:", bg=BG_COLOR,
         fg=TEXT_COLOR).grid(row=0, column=0,
                   padx=10, pady=10, sticky="e")
    ent_id = ttk.Entry(frm, width=20)
    ent_id.grid(row=0, column=1, padx=10, pady=10)
    ent_id.focus()

    tk.Label(frm, text="Password:", bg=BG_COLOR,
         fg=TEXT_COLOR).grid(row=1, column=0,
                   padx=10, pady=10, sticky="e")
    ent_pw = ttk.Entry(frm, width=20, show="*")
    ent_pw.grid(row=1, column=1, padx=10, pady=10)

    lbl_err = tk.Label(win, text="", bg=BG_COLOR,
              fg=ERROR_COLOR, font=HMI_FONT_S)
    lbl_err.pack()

    def do_login(event=None):
      uid  = ent_id.get().strip()
      upass = ent_pw.get()
      if not uid or not upass:
        lbl_err.config(text="Please enter ID and Password.")
        return
      try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT role FROM auth WHERE id=? AND password=?",
             (uid, hash_password(upass)))
        row = c.fetchone()
        conn.close()
      except Exception as e:
        lbl_err.config(text=f"DB Error: {e}")
        return

      if not row:
        lbl_err.config(text="Invalid ID or Password.")
        return

      role = row[0]
      if role not in ("Quality OP", "Quality Supervisor",
              "Manager", "Supervisor"):
        lbl_err.config(
          text="Access denied. Quality roles only.")
        return

      self.quality_op_id  = uid
      self.quality_op_role = role
      self.lbl_header_user.config(
        text=f"User: {uid} | Role: {role}")
      self.var_qop.set(uid)

      try:
        conn = get_db_connection()
        c = conn.cursor()
        ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute(
          "INSERT INTO system_access_logs "
          "(event_type, user_id, shift, timestamp) VALUES(?,?,?,?)",
          ("QUALITY_LOGIN", uid, "-", ts))
        conn.commit()
        conn.close()
      except Exception:
        pass

      win.destroy()
      self._refresh_log()
      self._refresh_analytics()

    ttk.Button(win, text="Login", style="Success.TButton",
          command=do_login).pack(pady=15)
    win.bind("<Return>", do_login)

  def _do_logout(self):
    if not self.quality_op_id:
      return
    if not messagebox.askyesno("Confirm Logout",
                  "Are you sure you want to logout?"):
      return
    self.quality_op_id  = ""
    self.quality_op_role = ""
    self.lbl_header_user.config(text="User: — | Role: —")
    self.var_qop.set("")
    self._prompt_login()

  # ── Report Defect Tab ────────────────────────────────────────────────────
  def _build_report_tab(self):
    outer = tk.Frame(self.tab_report, bg=BG_COLOR)
    outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # Split: left = scan+info, right = defect details
    left = tk.Frame(outer, bg=BG_COLOR)
    left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
    right = tk.Frame(outer, bg=BG_COLOR)
    right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

    # ── Scan card ──
    _, scan_content = self.create_card(
      left, "Scan Sub-Batch ID", fg_color=ACCENT_COLOR)

    scan_row = tk.Frame(scan_content, bg=SURFACE_COLOR)
    scan_row.pack(fill=tk.X, pady=6)

    self.var_scan = tk.StringVar()
    ent_scan = ttk.Entry(scan_row, textvariable=self.var_scan,
               width=32, font=HMI_FONT_M)
    ent_scan.pack(side=tk.LEFT, padx=(0, 8), ipady=4)
    ent_scan.bind("<Return>", self._on_scan)
    ent_scan.focus()

    ttk.Button(scan_row, text="Fetch", style="Primary.TButton",
          command=self._on_scan).pack(side=tk.LEFT, padx=2)
    ttk.Button(scan_row, text="Clear", style="Secondary.TButton",
          command=self._clear_report_form).pack(side=tk.LEFT, padx=2)

    # ── Box info card ──
    _, info_content = self.create_card(
      left, "Box Information (auto-filled)")

    info_grid = tk.Frame(info_content, bg=SURFACE_COLOR)
    info_grid.pack(fill=tk.X, pady=4)

    fields = [("PN (SF):", "pn"), ("Part Name:", "part"),
         ("Produced by OP:", "op"), ("Station:", "station"),
         ("Shift:", "shift"), ("Date / Time:", "dt")]
    self._info_vars = {}
    for i, (lbl, key) in enumerate(fields):
      row, col = divmod(i, 2)
      tk.Label(info_grid, text=lbl, bg=SURFACE_COLOR,
           fg=TEXT_MUTED, font=HMI_FONT_S,
           anchor="e").grid(row=row, column=col*2,
                   padx=(8, 4), pady=3, sticky="e")
      var = tk.StringVar(value="—")
      self._info_vars[key] = var
      tk.Label(info_grid, textvariable=var, bg=SURFACE_COLOR,
           fg=TEXT_COLOR, font=HMI_FONT_M,
           anchor="w").grid(row=row, column=col*2+1,
                   padx=(0, 16), pady=3, sticky="w")

    # ── Defect details card ──
    _, det_content = self.create_card(
      right, "Defect Details", fg_color=WARNING_COLOR)

    det = tk.Frame(det_content, bg=SURFACE_COLOR)
    det.pack(fill=tk.X, pady=4)

    # Defect type
    tk.Label(det, text="Defect Type ★", bg=SURFACE_COLOR,
         fg=TEXT_MUTED, font=HMI_FONT_S).grid(
           row=0, column=0, padx=8, pady=8, sticky="e")
    self.cb_defect_type = ttk.Combobox(
      det, values=DEFECT_TYPES, state="readonly",
      width=26, font=HMI_FONT_M)
    self.cb_defect_type.grid(row=0, column=1, padx=8, pady=8, sticky="w")
    self.cb_defect_type.set(DEFECT_TYPES[0])

    # Qty defective
    tk.Label(det, text="Qty Defective ★", bg=SURFACE_COLOR,
         fg=TEXT_MUTED, font=HMI_FONT_S).grid(
           row=1, column=0, padx=8, pady=8, sticky="e")
    qty_row = tk.Frame(det, bg=SURFACE_COLOR)
    qty_row.grid(row=1, column=1, padx=8, pady=8, sticky="w")
    vcmd = (self.register(lambda P: P.isdigit() or P == ""), "%P")
    self.var_qty = tk.StringVar()
    ttk.Entry(qty_row, textvariable=self.var_qty, width=8,
         validate="key", validatecommand=vcmd,
         font=HMI_FONT_M).pack(side=tk.LEFT, ipady=3)
    self.lbl_total_qty = tk.Label(qty_row, text="/ Total: —",
                   bg=SURFACE_COLOR, fg=TEXT_MUTED,
                   font=HMI_FONT_S)
    self.lbl_total_qty.pack(side=tk.LEFT, padx=8)

    # Description
    tk.Label(det, text="Description", bg=SURFACE_COLOR,
         fg=TEXT_MUTED, font=HMI_FONT_S).grid(
           row=2, column=0, padx=8, pady=8, sticky="ne")
    self.txt_desc = tk.Text(det, height=4, width=32,
                bg=BG_COLOR, fg=TEXT_COLOR,
                insertbackground="white",
                relief="flat", font=HMI_FONT_S,
                wrap=tk.WORD)
    self.txt_desc.grid(row=2, column=1, padx=8, pady=8, sticky="w")

    # Action Type
    tk.Label(det, text="Action Type ★", bg=SURFACE_COLOR,
         fg=TEXT_MUTED, font=HMI_FONT_S).grid(
           row=3, column=0, padx=8, pady=8, sticky="e")
    act_row = tk.Frame(det, bg=SURFACE_COLOR)
    act_row.grid(row=3, column=1, padx=8, pady=8, sticky="w")
    self.var_action = tk.StringVar(value="Scrap")
    self.cb_action_type = ttk.Combobox(act_row, textvariable=self.var_action,
                      values=["Scrap", "Rework", "Sorting", "Hold for Inspection", "Use As-Is"],
                      state="readonly", width=18, font=HMI_FONT_M)
    self.cb_action_type.pack(side=tk.LEFT)

    # Reported by
    tk.Label(det, text="Reported by ★", bg=SURFACE_COLOR,
         fg=TEXT_MUTED, font=HMI_FONT_S).grid(
           row=4, column=0, padx=8, pady=8, sticky="e")
    self.var_qop = tk.StringVar()
    self.ent_qop = ttk.Entry(det, textvariable=self.var_qop,
                 width=15, font=HMI_FONT_M)
    self.ent_qop.grid(row=4, column=1, padx=8, pady=8, sticky="w")

    # Status + Submit
    self.lbl_report_status = tk.Label(right, text="",
                     bg=BG_COLOR, fg=SUCCESS_COLOR,
                     font=HMI_FONT_M)
    self.lbl_report_status.pack(pady=5)

    ttk.Button(right, text="Submit Defect Report ",
          style="Danger.TButton",
          command=self._submit_defect).pack(pady=6, ipadx=10)

    self._current_record = None

  # ── Quality Log Tab ──────────────────────────────────────────────────────
  def _build_log_tab(self):
    outer = tk.Frame(self.tab_log, bg=BG_COLOR)
    outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    # ── Filters ──
    _, f_content = self.create_card(outer, "Filters")

    frow = tk.Frame(f_content, bg=SURFACE_COLOR)
    frow.pack(fill=tk.X, pady=4)

    tk.Label(frow, text="Date (YYYY-MM-DD):",
         bg=SURFACE_COLOR, fg=TEXT_MUTED,
         font=HMI_FONT_S).pack(side=tk.LEFT, padx=(0, 4))
    self.var_f_date = tk.StringVar()
    ttk.Entry(frow, textvariable=self.var_f_date,
         width=14).pack(side=tk.LEFT, padx=4)

    tk.Label(frow, text="Type:", bg=SURFACE_COLOR,
         fg=TEXT_MUTED, font=HMI_FONT_S).pack(side=tk.LEFT, padx=(12, 4))
    self.cb_f_type = ttk.Combobox(
      frow, values=["All"] + DEFECT_TYPES,
      state="readonly", width=20)
    self.cb_f_type.set("All")
    self.cb_f_type.pack(side=tk.LEFT, padx=4)

    tk.Label(frow, text="Status:", bg=SURFACE_COLOR,
         fg=TEXT_MUTED, font=HMI_FONT_S).pack(side=tk.LEFT, padx=(12, 4))
    self.cb_f_status = ttk.Combobox(
      frow, values=["All", "Open", "Closed"],
      state="readonly", width=10)
    self.cb_f_status.set("All")
    self.cb_f_status.pack(side=tk.LEFT, padx=4)

    ttk.Button(frow, text="Refresh", style="Primary.TButton",
          command=self._refresh_log).pack(side=tk.LEFT, padx=10)
    ttk.Button(frow, text="Reset", style="Secondary.TButton",
          command=self._reset_filters).pack(side=tk.LEFT)

    # ── Treeview ──
    cols = ("id", "sub_batch_id", "pn_sf", "defect_type", "action_type",
        "qty_defective", "reported_at", "status", "is_quarantined")
    hdrs = ("ID", "SB_ID", "PN", "Defect Type", "Action",
        "Qty", "Reported At", "Status", "Quarantine")
    wids = (40, 160, 160, 120, 70, 50, 130, 70, 80)

    tree_frame = tk.Frame(outer, bg=BG_COLOR)
    tree_frame.pack(fill=tk.BOTH, expand=True, pady=5)

    vsb = ttk.Scrollbar(tree_frame, orient="vertical")
    vsb.pack(side=tk.RIGHT, fill=tk.Y)
    hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
    hsb.pack(side=tk.BOTTOM, fill=tk.X)

    self.tree_log = ttk.Treeview(
      tree_frame, columns=cols, show="headings",
      yscrollcommand=vsb.set, xscrollcommand=hsb.set,
      selectmode="browse")
    vsb.config(command=self.tree_log.yview)
    hsb.config(command=self.tree_log.xview)

    for col, hdr, w in zip(cols, hdrs, wids):
      self.tree_log.heading(col, text=hdr,
                 command=lambda c=col: self._sort_log(c))
      self.tree_log.column(col, width=w, minwidth=40, anchor="center")

    self.tree_log.tag_configure("open",  background="#2D1515")
    self.tree_log.tag_configure("closed", background="#152D15")
    self.tree_log.pack(fill=tk.BOTH, expand=True)
    
    self.log_context_menu = tk.Menu(self.tree_log, tearoff=0, bg=SURFACE_COLOR, fg=TEXT_COLOR, activebackground=ACCENT_COLOR, activeforeground=TEXT_COLOR)
    
    def copy_to_clipboard(text):
      self.clipboard_clear()
      self.clipboard_append(text)
      self.update()

    def on_log_right_click(event):
      item = self.tree_log.identify_row(event.y)
      if item:
        self.tree_log.selection_set(item)
        values = self.tree_log.item(item, "values")
        if values:
          self.log_context_menu.delete(0, tk.END)
          self.log_context_menu.add_command(label=f" Copy SB_ID", command=lambda v=values[1]: copy_to_clipboard(v))
          self.log_context_menu.add_command(label=f" Copy PN", command=lambda v=values[2]: copy_to_clipboard(v))
          self.log_context_menu.add_separator()
          self.log_context_menu.add_command(label=f" Copy Row Data", command=lambda v=" | ".join(map(str, values)): copy_to_clipboard(v))
          self.log_context_menu.tk_popup(event.x_root, event.y_root)
          
    self.tree_log.bind("<Button-3>", on_log_right_click)

    # ── Action bar ──
    act = tk.Frame(outer, bg=BG_COLOR)
    act.pack(fill=tk.X, pady=6)

    ttk.Button(act, text="Mark as Closed",
          style="Success.TButton",
          command=lambda: self._change_status("Closed")
          ).pack(side=tk.LEFT, padx=5)

    ttk.Button(act, text="Toggle Defect Quarantine",
          style="Warning.TButton",
          command=self._toggle_quarantine
          ).pack(side=tk.LEFT, padx=5)

    self.lbl_log_count = tk.Label(act, text="",
                   bg=BG_COLOR, fg=TEXT_MUTED,
                   font=HMI_FONT_S)
    self.lbl_log_count.pack(side=tk.RIGHT, padx=10)

    self.lbl_excel_status = tk.Label(act, text="",
                     bg=BG_COLOR, fg=SUCCESS_COLOR,
                     font=HMI_FONT_S)
    self.lbl_excel_status.pack(side=tk.RIGHT, padx=10)

  # ── Analytics Tab ────────────────────────────────────────────────────────
  def _build_analytics_tab(self):
    outer = tk.Frame(self.tab_analytics, bg=BG_COLOR)
    outer.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    top = tk.Frame(outer, bg=BG_COLOR)
    top.pack(side=tk.TOP, fill=tk.X)
    ttk.Button(top, text="Refresh Analytics", style="Primary.TButton", command=self._refresh_analytics).pack(side=tk.LEFT, pady=5)

    left = tk.Frame(outer, bg=BG_COLOR)
    left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
    right = tk.Frame(outer, bg=BG_COLOR)
    right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))

    _, p_content = self.create_card(left, "Pareto Analysis (Defect Types)")
    self.pareto_frame = tk.Frame(p_content, bg=SURFACE_COLOR)
    self.pareto_frame.pack(fill=tk.BOTH, expand=True)

    _, t_content = self.create_card(right, "Top Offenders (By Defect Count)", fg_color=WARNING_COLOR)
    cols = ("Rank", "Operator ID", "Defect Count")
    self.tree_offenders = ttk.Treeview(t_content, columns=cols, show="headings", height=10)
    for c in cols:
      self.tree_offenders.heading(c, text=c)
      self.tree_offenders.column(c, anchor="center")
    self.tree_offenders.pack(fill=tk.BOTH, expand=True, pady=10)

  def _refresh_analytics(self):
    try:
      conn = get_db_connection()
      c = conn.cursor()
      
      c.execute("SELECT produced_by_op, COUNT(*) as cnt FROM quality_defects GROUP BY produced_by_op ORDER BY cnt DESC LIMIT 10")
      offenders = c.fetchall()
      
      for item in self.tree_offenders.get_children():
        self.tree_offenders.delete(item)
        
      for i, row in enumerate(offenders, 1):
        self.tree_offenders.insert("", "end", values=(i, row[0], row[1]))

      c.execute("SELECT defect_type, COUNT(*) as cnt FROM quality_defects GROUP BY defect_type ORDER BY cnt DESC")
      pareto_data = c.fetchall()
      conn.close()
    except Exception as e:
      print("Analytics DB Error:", e)
      return

    for widget in self.pareto_frame.winfo_children():
      widget.destroy()

    if not pareto_data:
      tk.Label(self.pareto_frame, text="No data available for chart.", bg=SURFACE_COLOR, fg=TEXT_MUTED).pack(pady=20)
      return

    types = [r[0] for r in pareto_data]
    counts = [r[1] for r in pareto_data]

    fig, ax = plt.subplots(figsize=(6, 4), facecolor=SURFACE_COLOR)
    ax.set_facecolor(SURFACE_COLOR)
    
    bars = ax.bar(types, counts, color=ACCENT_COLOR)
    ax.tick_params(axis='x', colors=TEXT_COLOR, rotation=15)
    ax.tick_params(axis='y', colors=TEXT_COLOR)
    ax.spines['bottom'].set_color(BORDER_COLOR)
    ax.spines['left'].set_color(BORDER_COLOR)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    
    fig.tight_layout()

    canvas = FigureCanvasTkAgg(fig, master=self.pareto_frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

  # ── Business logic ───────────────────────────────────────────────────────
  def _on_scan(self, event=None):
    sb_id = self.var_scan.get().strip()
    if not sb_id:
      return

    conn = get_db_connection()
    c = conn.cursor()

    c.execute("SELECT id FROM quality_defects WHERE sub_batch_id=?", (sb_id,))
    if c.fetchone():
      conn.close()
      self._set_status(f"SB_ID '{sb_id}' has already been reported.", ERROR_COLOR)
      self._current_record = None
      self._clear_info()
      return

    c.execute(
      "SELECT pn_sf, part_sf, op_id, station, shift_sp, dt_sp, quantity "
      "FROM records WHERE sub_batch_id=?", (sb_id,))
    row = c.fetchone()
    conn.close()

    if not row:
      self._set_status(
        f"SB_ID '{sb_id}' not found.", ERROR_COLOR)
      self._current_record = None
      self._clear_info()
      return

    pn, part, op, station, shift, dt_sp, qty = row
    self._current_record = dict(
      sub_batch_id=sb_id, pn_sf=pn or "—", part_sf=part or "—",
      produced_by_op=op or "—", station=station or "—",
      shift_sp=shift or "—", dt_sp=dt_sp or "—", quantity=qty or 0)

    self._info_vars["pn"].set(pn or "—")
    self._info_vars["part"].set(part or "—")
    self._info_vars["op"].set(op or "—")
    self._info_vars["station"].set(station or "—")
    self._info_vars["shift"].set(shift or "—")
    self._info_vars["dt"].set((dt_sp or "—")[:16])
    self.lbl_total_qty.config(text=f"/ Total in box: {qty or '—'}")
    if not self.var_qop.get():
      self.var_qop.set(self.quality_op_id)
    self._set_status(f"Box found: {pn} — {part}", SUCCESS_COLOR)

  def _submit_defect(self):
    q_op = self.var_qop.get().strip()
    if not q_op:
      messagebox.showerror("Missing Field",
                 "Please enter Quality OP ID.")
      return
    if not self._current_record:
      messagebox.showerror("No Box Scanned",
                 "Please scan a Sub-Batch ID first.")
      return

    defect_type = self.cb_defect_type.get()
    action_type = self.var_action.get()
    qty_str   = self.var_qty.get().strip()
    desc    = self.txt_desc.get("1.0", tk.END).strip()

    if not qty_str:
      messagebox.showerror("Missing Field", "Please enter Qty Defective.")
      return

    qty_def = int(qty_str)
    total  = self._current_record.get("quantity", 0)
    if qty_def < 1 or (total and qty_def > total):
      messagebox.showerror("Invalid Quantity",
                 f"Must be between 1 and {total}.")
      return

    reported_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    rec = self._current_record

    try:
      conn = get_db_connection()
      c = conn.cursor()
      
      c.execute("SELECT id FROM quality_defects WHERE sub_batch_id=?", (rec["sub_batch_id"],))
      if c.fetchone():
        conn.close()
        messagebox.showerror("Duplicate Report", "This Sub-Batch ID has already been reported.")
        return

      # Fetch current status first
      c.execute("SELECT status, quantity FROM records WHERE sub_batch_id=?", (rec["sub_batch_id"],))
      rec_row = c.fetchone()
      current_status = rec_row[0] if rec_row else 'In Rack'
      total_qty = rec_row[1] if rec_row else 0
      
      # Automatically quarantine UNLESS the box is already consumed
      is_quar = 0 if current_status == 'Consumed' else 1
      
      c.execute('''INSERT INTO quality_defects
             (sub_batch_id, pn_sf, part_sf, produced_by_op,
             quality_op_id, defect_type, qty_defective,
             description, shift_sp, station, reported_at, status, action_type, is_quarantined)
             VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
           (rec["sub_batch_id"], rec["pn_sf"],
            rec["part_sf"],   rec["produced_by_op"],
            q_op,        defect_type,
            qty_def,       desc,
            rec["shift_sp"],   rec["station"],
            reported_at,     "Open", action_type, is_quar))
      
      self._recalculate_box_status(c, rec["sub_batch_id"])
      
      conn.commit()
      conn.close()
    except sqlite3.IntegrityError:
      if 'conn' in locals():
        conn.rollback()
        conn.close()
      messagebox.showerror("Duplicate Report", "This Sub-Batch ID has already been reported.")
      return
    except Exception as e:
      if 'conn' in locals():
        conn.rollback()
        conn.close()
      messagebox.showerror("DB Error", f"Failed to save:\n{e}")
      return

    self._set_status(
      f"Report submitted for {rec['sub_batch_id']}.",
      SUCCESS_COLOR)
    self.after(4000, lambda: self._set_status("", SUCCESS_COLOR))
    self._clear_report_form()
    self._refresh_log()
    threading.Thread(target=self._export_excel, daemon=True).start()

  def _clear_report_form(self):
    self.var_scan.set("")
    self.var_qty.set("")
    self.txt_desc.delete("1.0", tk.END)
    self.cb_defect_type.set(DEFECT_TYPES[0])
    self.var_action.set("Scrap")
    self._current_record = None
    self.var_qop.set(self.quality_op_id)
    self._clear_info()
    self.lbl_report_status.config(text="")

  def _clear_info(self):
    for var in self._info_vars.values():
      var.set("—")
    self.lbl_total_qty.config(text="/ Total: —")

  def _set_status(self, msg, color=SUCCESS_COLOR):
    self.lbl_report_status.config(text=msg, fg=color)

  # ── Log ─────────────────────────────────────────────────────────────────
  def _refresh_log(self):
    date_f  = self.var_f_date.get().strip()
    type_f  = self.cb_f_type.get()
    status_f = self.cb_f_status.get()

    q = ("SELECT id, sub_batch_id, pn_sf, defect_type, action_type, qty_defective, "
       "reported_at, status, is_quarantined "
       "FROM quality_defects WHERE 1=1")
    params = []
    if date_f:
      q += " AND reported_at LIKE ?"
      params.append(f"{date_f}%")
    if type_f != "All":
      q += " AND defect_type=?"
      params.append(type_f)
    if status_f != "All":
      q += " AND status=?"
      params.append(status_f)
    q += " ORDER BY id DESC"

    try:
      conn = get_db_connection()
      c = conn.cursor()
      c.execute(q, params)
      rows = c.fetchall()

      today = datetime.datetime.now().strftime("%Y-%m-%d")
      c.execute("SELECT COUNT(*) FROM quality_defects WHERE reported_at LIKE ?",
           (f"{today}%",))
      krow = c.fetchone()
      total_def_today = krow[0] or 0

      # Overall Defect Rate calculation
      c.execute("SELECT SUM(qty_defective) FROM quality_defects")
      sum_def_all = c.fetchone()[0] or 0

      c.execute("SELECT SUM(quantity) FROM records")
      prod_res = c.fetchone()
      total_prod_all = prod_res[0] or 0
      
      defect_rate = (f"{sum_def_all / total_prod_all * 100:.2f}%"
              if total_prod_all else "N/A")

      c.execute("SELECT COUNT(*) FROM quality_defects WHERE status='Open'")
      open_cnt = c.fetchone()[0]
      conn.close()
    except Exception as e:
      messagebox.showerror("DB Error", str(e))
      return

    # Update sidebar KPIs
    self.kpi_total.config(text=str(total_def_today))
    self.kpi_rate.config(text=defect_rate)
    self.kpi_open.config(text=str(open_cnt),
               fg=ERROR_COLOR if open_cnt > 0 else SUCCESS_COLOR)

    # Populate treeview
    for item in self.tree_log.get_children():
      self.tree_log.delete(item)
    for row in rows:
      formatted_row = list(row)
      formatted_row[-1] = " YES" if row[-1] else "NO"
      tag = "open" if row[7] == "Open" else "closed"
      self.tree_log.insert("", "end", values=formatted_row, tags=(tag,))

    self.lbl_log_count.config(
      text=f"{len(rows)} defect(s) displayed")

  def _reset_filters(self):
    self.var_f_date.set("")
    self.cb_f_type.set("All")
    self.cb_f_status.set("All")
    self._refresh_log()

  def _sort_log(self, col):
    rows = [(self.tree_log.set(k, col), k)
        for k in self.tree_log.get_children("")]
    try:
      rows.sort(key=lambda t: int(t[0]))
    except ValueError:
      rows.sort()
    for idx, (_, k) in enumerate(rows):
      self.tree_log.move(k, "", idx)

  def _change_status(self, new_status):
    sel = self.tree_log.selection()
    if not sel:
      messagebox.showinfo("No Selection", "Please select a defect record.")
      return

    vals      = self.tree_log.item(sel[0], "values")
    row_id     = vals[0]
    current_status = vals[7]

    if current_status == "Closed":
      messagebox.showwarning("Ticket Locked", "This defect report is Closed and cannot be reopened.")
      return

    if new_status == "Closed":
      self._prompt_capa(row_id)
    else:
      self._update_status_db(row_id, new_status)

  def _prompt_capa(self, row_id):
    win = tk.Toplevel(self)
    win.title("CAPA - Root Cause & Corrective Action")
    win.geometry("500x400")
    win.configure(bg=BG_COLOR)
    win.transient(self)
    win.grab_set()

    tk.Label(win, text="Root Cause:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(anchor="w", padx=20, pady=(20, 5))
    txt_rc = tk.Text(win, height=4, bg=SURFACE_COLOR, fg=TEXT_COLOR, insertbackground="white", font=HMI_FONT_S)
    txt_rc.pack(fill=tk.X, padx=20)

    tk.Label(win, text="Corrective Action:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(anchor="w", padx=20, pady=(15, 5))
    txt_ca = tk.Text(win, height=4, bg=SURFACE_COLOR, fg=TEXT_COLOR, insertbackground="white", font=HMI_FONT_S)
    txt_ca.pack(fill=tk.X, padx=20)

    def submit():
      rc = txt_rc.get("1.0", tk.END).strip()
      ca = txt_ca.get("1.0", tk.END).strip()
      if not rc or not ca:
        messagebox.showerror("Missing Info", "Both Root Cause and Corrective Action are required.", parent=win)
        return
      win.destroy()
      self._update_status_db(row_id, "Closed", rc, ca)

    ttk.Button(win, text="Submit & Close Ticket", style="Success.TButton", command=submit).pack(pady=20)

  def _update_status_db(self, row_id, new_status, root_cause=None, corrective_action=None):
    try:
      conn = get_db_connection()
      c = conn.cursor()
      if root_cause and corrective_action:
        c.execute("UPDATE quality_defects SET status=?, root_cause=?, corrective_action=? WHERE id=?",
             (new_status, root_cause, corrective_action, row_id))
      else:
        c.execute("UPDATE quality_defects SET status=? WHERE id=?", (new_status, row_id))
        
      # Re-evaluate box status if marked as Closed
      c.execute("SELECT sub_batch_id FROM quality_defects WHERE id=?", (row_id,))
      sb_row = c.fetchone()
      if sb_row:
        self._recalculate_box_status(c, sb_row[0])
      conn.commit()
      conn.close()
    except Exception as e:
      messagebox.showerror("DB Error", str(e))
      return

    self._refresh_log()
    threading.Thread(target=self._export_excel, daemon=True).start()

  def _toggle_quarantine(self):
    sel = self.tree_log.selection()
    if not sel:
      messagebox.showinfo("No Selection", "Please select a defect record.")
      return

    vals = self.tree_log.item(sel[0], "values")
    row_id = vals[0]
    is_quarantined = 1 if vals[-1] == "NO" else 0 # Toggle logic

    try:
      conn = get_db_connection()
      c = conn.cursor()
      c.execute("UPDATE quality_defects SET is_quarantined=? WHERE id=?", (is_quarantined, row_id))
      
      c.execute("SELECT sub_batch_id FROM quality_defects WHERE id=?", (row_id,))
      q_row = c.fetchone()
      if q_row:
        self._recalculate_box_status(c, q_row[0])
        
      conn.commit()
      conn.close()
    except Exception as e:
      messagebox.showerror("DB Error", str(e))
      return

    self._refresh_log()

  # ── Excel Export ─────────────────────────────────────────────────────────
  def _export_excel(self):
    try:
      if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        if "Quality Defects" in wb.sheetnames:
          del wb["Quality Defects"]
      else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
          del wb["Sheet"]

      ws = wb.create_sheet("Quality Defects")

      hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
      hdr_fill = PatternFill(start_color="1B232C", end_color="1B232C",
                  fill_type="solid")
      acc_fill = PatternFill(start_color="F59E0B", end_color="F59E0B",
                  fill_type="solid")
      open_fill = PatternFill(start_color="FFE4E4", end_color="FFE4E4",
                  fill_type="solid")
      cld_fill = PatternFill(start_color="E4FFE4", end_color="E4FFE4",
                  fill_type="solid")
      thin = Side(style="thin", color="334155")
      bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
      center = Alignment(horizontal="center", vertical="center")
      left  = Alignment(horizontal="left",  vertical="center")

      headers = ["#", "Sub-Batch ID", "PN (SF)", "Part Name",
            "Defect Type", "Qty Defective", "Produced by OP",
            "Quality OP", "Shift", "Station", "Reported At", "Status"]
      col_w  = [5, 24, 26, 26, 20, 14, 14, 12, 8, 10, 20, 10]

      ws.merge_cells("A1:L1")
      tc = ws.cell(row=1, column=1,
             value="Quality Defects Report — HI-LEX ACT")
      tc.font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
      tc.fill = acc_fill
      tc.alignment = center
      ws.row_dimensions[1].height = 26

      for ci, hdr in enumerate(headers, 1):
        cell = ws.cell(row=2, column=ci, value=hdr)
        cell.font   = hdr_font
        cell.fill   = hdr_fill
        cell.alignment = center
        cell.border  = bdr
        ws.column_dimensions[cell.column_letter].width = col_w[ci-1]
      ws.row_dimensions[2].height = 20

      conn = get_db_connection()
      c = conn.cursor()
      c.execute("SELECT id, sub_batch_id, pn_sf, part_sf, defect_type, "
           "qty_defective, produced_by_op, quality_op_id, "
           "shift_sp, station, reported_at, status "
           "FROM quality_defects ORDER BY id DESC")
      rows = c.fetchall()
      conn.close()

      for ri, row in enumerate(rows, 3):
        rf = open_fill if row[-1] == "Open" else cld_fill
        for ci, val in enumerate(row, 1):
          cell = ws.cell(row=ri, column=ci, value=val)
          cell.fill   = rf
          cell.border  = bdr
          cell.alignment = center if ci != 4 else left
        ws.row_dimensions[ri].height = 18

      ws.freeze_panes = "A3"
      ws.auto_filter.ref = f"A2:L{max(2, 1+len(rows))}"

      tr = len(rows) + 4
      summary = [
        ("Total Reports",  len(rows)),
        ("Total Defective", sum(r[5] for r in rows)),
        ("Open Cases",   sum(1 for r in rows if r[-1]=="Open")),
        ("Closed Cases",  sum(1 for r in rows if r[-1]=="Closed")),
        ("Exported at",
         datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
      ]
      for si, (lbl, val) in enumerate(summary):
        ws.cell(row=tr+si, column=1, value=lbl).font = Font(
          bold=True, name="Calibri", size=10)
        ws.cell(row=tr+si, column=2, value=val).font = Font(
          name="Calibri", size=10)

      wb.save(EXCEL_FILE)
      # Update UI on main thread
      self.after(0, lambda: self.lbl_excel_status.config(
        text=f" Excel saved {datetime.datetime.now().strftime('%H:%M:%S')}"))
      self.after(5000, lambda: self.lbl_excel_status.config(text=""))
    except Exception as e:
      print(f"[Quality Excel Export Error]: {e}")

  def _open_excel(self):
    if not os.path.exists(EXCEL_FILE):
      self._export_excel()
      
    if os.path.exists(EXCEL_FILE):
      os.startfile(EXCEL_FILE)
    else:
      messagebox.showinfo("Error", "Could not generate Excel file.")

  # ── Clock ────────────────────────────────────────────────────────────────
  def _tick_clock(self):
    now = datetime.datetime.now()
    try:
      self.lbl_clock.config(text=now.strftime("%H:%M"))
      self.lbl_date.config(text=now.strftime("%Y-%m-%d"))
    except Exception:
      return
    self._clock_job = self.after(1000, self._tick_clock)

  # ── Close ────────────────────────────────────────────────────────────────
  def on_closing(self):
    if hasattr(self, '_clock_job'):
      self.after_cancel(self._clock_job)
    self.destroy()


# ─────────────────────────────────────────────────────────────────────────────
  def _recalculate_box_status(self, c, sb_id):
    c.execute("SELECT status, quantity FROM records WHERE sub_batch_id=?", (sb_id,))
    rec_row = c.fetchone()
    if not rec_row or rec_row[0] == 'Consumed': return
    
    total_qty = rec_row[1]
    
    c.execute("SELECT IFNULL(SUM(CASE WHEN status='Closed' AND action_type IN ('Rework', 'Sorting', 'Use As-Is') THEN 0 ELSE qty_defective END), 0) FROM quality_defects WHERE sub_batch_id=?", (sb_id,))
    effective_defects = c.fetchone()[0]
    
    c.execute("SELECT COUNT(*) FROM quality_defects WHERE sub_batch_id=? AND is_quarantined=1 AND NOT (status='Closed' AND action_type IN ('Rework', 'Sorting', 'Use As-Is'))", (sb_id,))
    has_quarantined = c.fetchone()[0] > 0
    
    good_qty = total_qty - effective_defects
    
    if good_qty <= 0:
      new_status = 'Quarantined'
    elif has_quarantined or effective_defects > 0:
      new_status = 'Partial Defect'
    else:
      new_status = 'In Rack'
      
    c.execute("UPDATE records SET status=? WHERE sub_batch_id=?", (new_status, sb_id))

if __name__ == "__main__":
  init_quality_db()
  app = QualityApp()
  app.mainloop()
