import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import sqlite3
import datetime
import os
import webbrowser
import json
import qrcode
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
import sys
import threading
import queue

_excel_queue = queue.Queue()

def _excel_worker():
    while True:
        func, args, kwargs = _excel_queue.get()
        try:
            func(*args, **kwargs)
        except Exception as e:
            print(f"Excel worker error: {e}")
        finally:
            _excel_queue.task_done()

_excel_thread = threading.Thread(target=_excel_worker, daemon=True)
_excel_thread.start()
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import report_generator

def resource_path(relative_path):
  """ Get absolute path to resource, works for dev and for PyInstaller """
  try:
    base_path = sys._MEIPASS
  except Exception:
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
  return os.path.join(base_path, relative_path)

def persistent_path(relative_path):
  """ Get absolute path to persistent data, relative to executable """
  if hasattr(sys, '_MEIPASS'):
    base_path = os.path.dirname(sys.executable)
  else:
    base_path = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
  return os.path.join(base_path, relative_path)

def center_window(win, width, height):
  win.update_idletasks()
  x = (win.winfo_screenwidth() // 2) - (width // 2)
  y = (win.winfo_screenheight() // 2) - (height // 2)
  win.geometry(f"{width}x{height}+{x}+{y}")

try:
  import win32print
  WIN32_PRINT_AVAILABLE = True
except ImportError:
  WIN32_PRINT_AVAILABLE = False

try:
  import matplotlib # type: ignore
  matplotlib.use("Agg")
  from matplotlib.backends.backend_agg import FigureCanvasAgg # type: ignore
  from matplotlib.figure import Figure # type: ignore
  import matplotlib.patches as mpatches
  from PIL import Image, ImageTk
  import numpy as np
  MATPLOTLIB_AVAILABLE = True
except ImportError:
  MATPLOTLIB_AVAILABLE = False

DATA_DIR = persistent_path("data")
os.makedirs(DATA_DIR, exist_ok=True)

# Migration logic for old data files
_base = os.path.abspath(os.path.join(DATA_DIR, ".."))
for _f in ["traceability_config.json", "sf_data.json", "traceability.db", "production_data.xlsx"]:
  _old = os.path.join(_base, _f)
  _new = os.path.join(DATA_DIR, _f)
  if os.path.exists(_old) and not os.path.exists(_new):
    try:
      import shutil
      shutil.move(_old, _new)
    except Exception:
      pass

CONFIG_FILE = os.path.join(DATA_DIR, "traceability_config.json")
APP_VERSION = "1.1.0"
# Premium Industrial HMI Theme

BG_COLOR = "#121212"    # Main Background
SURFACE_COLOR = "#1E1E1E" # Panels & Cards

ACCENT_COLOR = "#00B4D8"  # Primary Action
SUCCESS_COLOR = "#2DC653" # Running / Completed
WARNING_COLOR = "#F59E0B" # Warning
ERROR_COLOR = "#DC2626"  # Error / Stop

BORDER_COLOR = "#334155"  # Borders

TEXT_COLOR = "#F8FAFC"   # Main Text
TEXT_MUTED = "#94A3B8"   # Secondary Text

# Status Colors
STATUS_RUNNING = "#22C55E"
STATUS_IDLE = "#64748B"
STATUS_WARNING = "#F59E0B"
STATUS_STOPPED = "#EF4444"

# Modern HMI Fonts
HMI_FONT_XL = ("Segoe UI", 20, "bold")
HMI_FONT_L = ("Segoe UI", 16, "bold")
HMI_FONT_M = ("Segoe UI", 12, "bold")
HMI_FONT_S = ("Segoe UI", 10)
HMI_FONT_XS = ("Segoe UI", 9)

# Default Hardcoded Data (Fallback)
SF_DATA_DEFAULT = {
  "MOCK-A01-10001-X-SUB": ("Alpha Subsystem Right", [("MOCK-A01-10001-Y-PRT", "Alpha Core Right"), ("MOCK-A01-10002-Z-PRT", "Alpha Buffer Right")]),
  "MOCK-A01-10003-X-SUB": ("Alpha Subsystem Aux Right", [("MOCK-A01-10003-Y-PRT", "Alpha Core (Auxiliary) Right"), ("MOCK-A01-10002-Z-PRT", "Alpha Buffer Right")]),
  "MOCK-A01-10004-X-SUB": ("Alpha Subsystem Left", [("MOCK-A01-10004-Y-PRT", "Alpha Core Left"), ("MOCK-A01-10005-Z-PRT", "Alpha Buffer Left")]),
  "MOCK-A01-10006-X-SUB": ("Alpha Subsystem Aux Left", [("MOCK-A01-10006-Y-PRT", "Alpha Core (Auxiliary) Left"), ("MOCK-A01-10005-Z-PRT", "Alpha Buffer Left")]),
  "MOCK-B02-20001-X-SUB": ("Beta Panel Right Sub", [("MOCK-B02-20001-Y-PRT", "Beta Panel Right"), ("MOCK-B02-20002-Z-PRT", "Beta Sealant")]),
  "MOCK-B02-20003-X-SUB": ("Beta Panel Left Sub", [("MOCK-B02-20003-Y-PRT", "Beta Panel Left"), ("MOCK-B02-20002-Z-PRT", "Beta Sealant")]),
  "MOCK-B02-20004-X-SUB": ("Beta Panel Right Sub V2", [("MOCK-B02-20004-Y-PRT", "Beta Panel (Upgraded) Right"), ("MOCK-B02-20002-Z-PRT", "Beta Sealant")]),
  "MOCK-B02-20005-X-SUB": ("Beta Panel Left Sub V2", [("MOCK-B02-20005-Y-PRT", "Beta Panel (Upgraded) Left"), ("MOCK-B02-20002-Z-PRT", "Beta Sealant")]),
  "MOCK-C03-30001-X-SUB": ("Gamma Switch Right Sub (Short)", [("MOCK-C03-30001-Y-PRT", "Gamma Switch Right"), ("MOCK-C03-30002-Z-PRT", "Gamma Coil Right"), ("MOCK-C03-30003-W-PRT", "Gamma Pin")]),
  "MOCK-C03-30004-X-SUB": ("Gamma Switch Left Sub (Short)", [("MOCK-C03-30004-Y-PRT", "Gamma Switch Left"), ("MOCK-C03-30005-Z-PRT", "Gamma Coil Left"), ("MOCK-C03-30003-W-PRT", "Gamma Pin")]),
  "MOCK-C03-30006-X-SUB": ("Gamma Switch Right Sub (Long)", [("MOCK-C03-30006-Y-PRT", "Gamma Switch Long Right"), ("MOCK-C03-30002-Z-PRT", "Gamma Coil Right"), ("MOCK-C03-30003-W-PRT", "Gamma Pin")]),
  "MOCK-C03-30007-X-SUB": ("Gamma Switch Left Sub (Long)", [("MOCK-C03-30007-Y-PRT", "Gamma Switch Long Left"), ("MOCK-C03-30005-Z-PRT", "Gamma Coil Left"), ("MOCK-C03-30003-W-PRT", "Gamma Pin")]),
  "MOCK-D04-40001-X-SUB": ("Delta Processor Node Right", [("MOCK-D04-40001-Y-PRT", "Delta Primary Node"), ("MOCK-D04-40002-Z-PRT", "Delta Interface Right"), ("MOCK-D04-40003-W-PRT", "Delta Rotor")]),
  "MOCK-D04-40004-X-SUB": ("Delta Processor Node Left", [("MOCK-D04-40001-Y-PRT", "Delta Primary Node"), ("MOCK-D04-40005-Z-PRT", "Delta Interface Left"), ("MOCK-D04-40003-W-PRT", "Delta Rotor")]),
  "MOCK-D04-40006-X-SUB": ("Delta Processor Array Right", [("MOCK-D04-40001-Y-PRT", "Delta Primary Node"), ("MOCK-D04-40007-Y-PRT", "Delta Secondary Node"), ("MOCK-D04-40008-Z-PRT", "Delta Interface Array Right"), ("MOCK-D04-40003-W-PRT", "Delta Rotor")]),
  "MOCK-E05-50001-X-SUB": ("Epsilon Housing Sub Right", [("MOCK-E05-50001-Y-PRT", "Epsilon Housing Right"), ("MOCK-E05-50002-Z-PRT", "Epsilon Axis")]),
  "MOCK-E05-50003-X-SUB": ("Epsilon Housing Sub Left", [("MOCK-E05-50003-Y-PRT", "Epsilon Housing Left"), ("MOCK-E05-50002-Z-PRT", "Epsilon Axis")]),
  "MOCK-F06-60001-X-SUB": ("Zeta Mount Assembly RH", [("MOCK-F06-60001-Y-PRT", "Zeta Mount RH"), ("MOCK-F06-60002-Z-PRT", "Zeta Fastener")]),
  "MOCK-F06-60003-X-SUB": ("Zeta Mount Assembly LH", [("MOCK-F06-60003-Y-PRT", "Zeta Mount LH"), ("MOCK-F06-60002-Z-PRT", "Zeta Fastener")]),
  "MOCK-G07-70001-X-SUB": ("Omega Framework Sub Left", [("MOCK-G07-70001-Y-PRT", "Omega Anchor Left"), ("MOCK-G07-70002-Y-PRT", "Omega Frame Left"), ("MOCK-A01-10005-Z-PRT", "Alpha Buffer Left"), ("MOCK-B02-20002-Z-PRT", "Beta Sealant")])
}

SF_DATA_FILE = os.path.join(DATA_DIR, "sf_data.json")
SF_DATA = {}

def load_sf_data():
  global SF_DATA
  SF_DATA = {}
  try:
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT pn_sf, name_sf, rms_json FROM products")
    rows = c.fetchall()
    for r in rows:
      pn = r[0]
      name = r[1]
      try:
        rms = json.loads(r[2])
      except:
        rms = []
      SF_DATA[pn] = (name, rms)
    conn.close()
    
    if not SF_DATA:
      SF_DATA = dict(SF_DATA_DEFAULT)
      conn = get_db_connection()
      c = conn.cursor()
      for pn, val in SF_DATA.items():
        c.execute("INSERT INTO products (pn_sf, name_sf, rms_json) VALUES (?, ?, ?)", (pn, val[0], json.dumps(val[1])))
      conn.commit()
      conn.close()
      
  except Exception as e:
    print("Error loading SF_DATA from DB:", e)
    SF_DATA = dict(SF_DATA_DEFAULT)


DB_FILE = os.path.join(DATA_DIR, "traceability.db")

def get_db_connection():
  conn = sqlite3.connect(DB_FILE, check_same_thread=False, timeout=10)
  conn.execute('PRAGMA journal_mode=WAL')
  return conn

import hashlib
import binascii

def hash_password(password: str, salt: bytes = None) -> str:
  if salt is None:
    salt = os.urandom(32)
  key = hashlib.pbkdf2_hmac("sha256", password.encode(), salt, 310_000)
  return binascii.hexlify(salt).decode() + ":" + binascii.hexlify(key).decode()

def verify_password(password: str, stored: str) -> bool:
  try:
    if ":" not in stored:
      return False
    salt_hex, key_hex = stored.split(":")
    salt = binascii.unhexlify(salt_hex)
    
    # 1. Normal check (new schema)
    expected = hash_password(password, salt)
    if expected == stored:
      return True
      
    # 2. Migration fallback check (double-hashed from old static salt schema)
    old_salt = b"subproc_trace_salt_2026"
    old_hash_obj = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), old_salt, 100000)
    old_hash_str = binascii.hexlify(old_hash_obj).decode("utf-8")
    
    expected_old = hash_password(old_hash_str, salt)
    if expected_old == stored:
      return True
      
    return False
  except Exception:
    return False

def migrate_passwords():
  conn = get_db_connection()
  c = conn.cursor()
  c.execute("SELECT id, password FROM auth")
  rows = c.fetchall()
  for row in rows:
    uid, pwd = row
    if pwd and ":" not in pwd:
      hashed = hash_password(pwd)
      c.execute("UPDATE auth SET password = ? WHERE id = ?", (hashed, uid))
  conn.commit()
  conn.close()
EXCEL_FILE = os.path.join(DATA_DIR, "production_data.xlsx")

def init_db():
  try:
    from quality_app import init_quality_db
    init_quality_db()
  except ImportError:
    pass
    
  conn = get_db_connection()
  c = conn.cursor()
  
  c.execute('''
    CREATE TABLE IF NOT EXISTS products (
      pn_sf TEXT PRIMARY KEY,
      name_sf TEXT,
      rms_json TEXT
    )
  ''')
  
  c.execute('''
    CREATE TABLE IF NOT EXISTS product_audit_trail (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      action TEXT,
      pn_sf TEXT,
      details TEXT,
      user_id TEXT,
      shift TEXT,
      timestamp TEXT
    )
  ''')
  
  c.execute('''
    CREATE TABLE IF NOT EXISTS system_access_logs (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      event_type TEXT,
      user_id TEXT,
      shift TEXT,
      timestamp TEXT
    )
  ''')
  
  c.execute('''
    CREATE TABLE IF NOT EXISTS auth (
      id TEXT PRIMARY KEY,
      password TEXT,
      role TEXT
    )
  ''')

  c.execute('''
    CREATE TABLE IF NOT EXISTS schema_version (
      version INTEGER PRIMARY KEY,
      applied_at TEXT DEFAULT CURRENT_TIMESTAMP
    )
  ''')
  c.execute("INSERT OR IGNORE INTO schema_version (version) VALUES (1)")
  
  try:
    import os
    default_users = [
      ('mg90', os.environ.get('MANAGER_PASS', 'oi90'), 'Manager'),
      ('sp99', os.environ.get('SUPERVISOR_PASS', 'sp99'), 'Supervisor'),
      ('sl', os.environ.get('SHIFT_LEADER_PASS', 'sl98'), 'Shift Leader'),
      ('op1', os.environ.get('OPERATOR_PASS', 'op1'), 'Operator')
    ]
    for uid, plain_pwd, role in default_users:
      c.execute("INSERT OR IGNORE INTO auth (id, password, role) VALUES (?, ?, ?)", (uid, hash_password(plain_pwd), role))
    
    # Also inserting the 's' supervisor from original if not overridden
    c.execute("INSERT OR IGNORE INTO auth (id, password, role) VALUES (?, ?, ?)", ('s', hash_password(' '), 'Supervisor'))
    
    conn.commit()
  except Exception as e:
    print("Auth seed error:", e)
  
  migrate_passwords()
    
  c.execute("SELECT COUNT(*) FROM products")
  count = c.fetchone()[0]
  if count == 0 and os.path.exists(SF_DATA_FILE):
    try:
      with open(SF_DATA_FILE, "r", encoding="utf-8") as f:
        json_data = json.load(f)
      for pn, val in json_data.items():
        c.execute("INSERT INTO products (pn_sf, name_sf, rms_json) VALUES (?, ?, ?)", (pn, val[0], json.dumps(val[1])))
      conn.commit()
    except Exception:
      pass

  c.execute('''
    CREATE TABLE IF NOT EXISTS records (
      id INTEGER PRIMARY KEY AUTOINCREMENT,
      sub_batch_id TEXT NOT NULL,
      pn_sf TEXT,
      part_sf TEXT,
      rm1_pn TEXT,
      rm1_name TEXT,
      rm2_pn TEXT,
      rm2_name TEXT,
      rm3_pn TEXT,
      rm3_name TEXT,
      rm4_pn TEXT,
      rm4_name TEXT,
      batch1 TEXT,
      batch2 TEXT,
      batch3 TEXT,
      quantity INTEGER,
      shift_sp TEXT,
      op_id TEXT,
      station TEXT,
      dt_sp TEXT,
      dt_line TEXT,
      shift_line TEXT,
      remarks TEXT,
      status TEXT DEFAULT 'In Rack',
      created_at TEXT NOT NULL
    )
  ''')
  
  try:
    c.execute("ALTER TABLE records ADD COLUMN registered_by TEXT DEFAULT ''")
  except sqlite3.OperationalError:
    pass
    
  try:
    c.execute("ALTER TABLE records ADD COLUMN reprint_count INTEGER DEFAULT 0")
  except sqlite3.OperationalError:
    pass
    
  try:
    c.execute("ALTER TABLE records ADD COLUMN last_reprinted_at TEXT")
  except sqlite3.OperationalError:
    pass
    
  try:
    c.execute("ALTER TABLE records ADD COLUMN last_reprinted_by TEXT")
  except sqlite3.OperationalError:
    pass
    
  c.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_records_sub_batch_id ON records(sub_batch_id)")
  
  c.execute('''CREATE TABLE IF NOT EXISTS shift_targets (
           id INTEGER PRIMARY KEY AUTOINCREMENT,
           product_pn TEXT NOT NULL,
           shift TEXT NOT NULL,
           station TEXT,
           target_qty INTEGER NOT NULL,
           effective_date TEXT NOT NULL
         )''')

  c.execute('''CREATE TABLE IF NOT EXISTS part_thresholds (
           pn_sf TEXT PRIMARY KEY,
           min_qty INTEGER NOT NULL DEFAULT 0
         )''')

  c.execute('''CREATE TABLE IF NOT EXISTS inventory_snapshots (
           snapshot_date TEXT NOT NULL,
           pn_sf TEXT NOT NULL,
           boxes_in_rack INTEGER,
           total_qty_in_rack INTEGER,
           oldest_box_age_hours REAL,
           PRIMARY KEY (snapshot_date, pn_sf)
         )''')
    
  c.execute('''CREATE TABLE IF NOT EXISTS downtime_logs (
           id INTEGER PRIMARY KEY AUTOINCREMENT,
           sub_batch_id TEXT,
           station TEXT,
           shift TEXT,
           op_id TEXT,
           duration_min REAL,
           reason TEXT,
           created_at TEXT
         )''')
         
  # Migration: add op_id to downtime_logs if it doesn't exist
  try:
      c.execute("ALTER TABLE downtime_logs ADD COLUMN op_id TEXT")
  except sqlite3.OperationalError:
      pass

  conn.commit()
  conn.close()
  
  load_sf_data()


def rebuild_excel_from_db():
  if os.path.exists(EXCEL_FILE):
    try:
      os.remove(EXCEL_FILE)
    except Exception:
      pass
      
  wb = Workbook()
  ws = wb.active
  ws.title = "Sub-process fill by TL"
  
  headers = [
    "SB ID", "FULL PN Semi fini", "PART NAME (SF)", 
    "RM PN", "RM Name",
    "Batch No. 1", "Batch No. 2", "Batch No. 3", "Quantity",
    "Work in Sub-Process by Shift", "Op ID", "Station",
    "Sub-Process Work Date/Time", "Production Line Entry Date/Time",
    "Work in PROD line by Shift", "Remarks", "Registered by ID"
  ]
  ws.append(headers)
  ws.freeze_panes = "A2"
  
  header_font = Font(name='Calibri', size=10, bold=True, color="000000")
  header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
  header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
  thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

  for col_idx in range(1, 18):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    
  ws.row_dimensions[1].height = 34
  widths = [20, 22, 25, 22, 25, 12, 12, 12, 10, 15, 10, 10, 20, 20, 15, 25, 15]
  from openpyxl.utils import get_column_letter
  for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

  conn = get_db_connection()
  conn.row_factory = sqlite3.Row
  c = conn.cursor()
  c.execute("SELECT * FROM records ORDER BY id ASC")
  rows = c.fetchall()
  conn.close()
  
  for row in rows:
    data = [
      row["sub_batch_id"], row["pn_sf"], row["part_sf"],
      row["rm1_pn"], row["rm1_name"], row["rm2_pn"], row["rm2_name"],
      row["rm3_pn"], row["rm3_name"], row["rm4_pn"], row["rm4_name"],
      row["batch1"], row["batch2"], row["batch3"], row["quantity"],
      row["shift_sp"], row["op_id"], row["station"], row["dt_sp"], row["dt_line"],
      row["shift_line"], row["remarks"], row["registered_by"]
    ]
    _write_record_to_ws(ws, data)
    
  wb.save(EXCEL_FILE)

def save_to_excel(data):
  if not os.path.exists(EXCEL_FILE):
    rebuild_excel_from_db()
    return

  try:
    wb = load_workbook(EXCEL_FILE)
    if "Sub-process fill by TL" in wb.sheetnames:
      ws = wb["Sub-process fill by TL"]
    else:
      rebuild_excel_from_db()
      return
      
    _write_record_to_ws(ws, data)
    wb.save(EXCEL_FILE)
  except Exception as e:
    print(f"Error loading Excel file: {e}")
    return

def _write_record_to_ws(ws, data):
  sb_id = data[0] if data[0] else "-"
  sf_pn = data[1] if data[1] else "-"
  part_sf = data[2] if data[2] else "-"
  rms = []
  for i in range(4):
    rm_pn = data[3 + i*2]
    rm_name = data[4 + i*2]
    if rm_pn:
      rms.append((rm_pn if rm_pn else "-", rm_name if rm_name else "-"))
      
  if not rms:
    rms = [("-", "-")]
    
  num_rms = len(rms)
  rest_data = ["-" if (x == "" or x is None) else x for x in data[11:]]
  
  start_row = ws.max_row + 1
  end_row = start_row + num_rms - 1
  
  for i in range(num_rms):
    row_data = []
    if i == 0:
      row_data.extend([sb_id, sf_pn, part_sf])
      row_data.extend([rms[i][0], rms[i][1]])
      row_data.extend(rest_data)
    else:
      row_data.extend([None, None, None])
      row_data.extend([rms[i][0], rms[i][1]])
      row_data.extend([None] * len(rest_data))
    ws.append(row_data)
    
  if num_rms > 1:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
    ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
    ws.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
    for col in range(6, 18):
      ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
  
  sf_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
  rm_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
  rest_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
  row_font = Font(name='Calibri', size=10, color="000000")
  align = Alignment(horizontal="center", vertical="center", wrap_text=True)
  thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
  
  for r in range(start_row, end_row + 1):
    for c in range(1, 18):
      cell = ws.cell(row=r, column=c)
      cell.font = row_font
      cell.alignment = align
      cell.border = thin_border
      if c in (1, 2, 3):
        cell.fill = sf_fill
      elif c in (4, 5):
        cell.fill = rm_fill
      else:
        cell.fill = rest_fill


def generate_zpl(record_data):
  def format_val(val, default="-"):
    return str(val) if val else default

  zpl = []
  zpl.append("^XA")
  zpl.append("^PW520")
  zpl.append("^LH0,0")
  zpl.append("^CI28")
  
  def draw_layout(offset_y):
    zpl.append(f"^FO30,{20+offset_y}^A0N,24,24^FB460,1,0,C^FDSUB-PROCESS RECORD\\&^FS")
    zpl.append(f"^FO30,{55+offset_y}^GB460,3,3^FS")
    
    zpl.append(f"^FO30,{70+offset_y}^GB460,40,40^FS")
    zpl.append(f"^FO30,{80+offset_y}^A0N,24,24^FR^FB460,1,0,C^FD{format_val(record_data.get('sub_batch_id'))}\\&^FS")
    
    sb_id_val = format_val(record_data.get('sub_batch_id'))
    zpl.append(f"^BY2^FO50,{120+offset_y}^BCN,50,N,N,N,A^FD{sb_id_val}^FS")
    
    y = 185 + offset_y
    def add_row(label, value, font_size=18, y_inc=20):
      nonlocal y
      zpl.append(f"^FO30,{y}^A0N,{font_size},{font_size}^FD{label}^FS")
      zpl.append(f"^FO30,{y}^A0N,{font_size},{font_size}^FB460,1,0,R^FD{value}\\&^FS")
      y += y_inc

    add_row("PN Semi fini", format_val(record_data.get('pn_sf')))
    add_row("Part Name (SF)", format_val(record_data.get('part_sf')))
        
    y += 3
    zpl.append(f"^FO30,{y}^GB460,1,1^FS")
    y += 8
    
    add_row("Batch No. 1", format_val(record_data.get('batch1')))
    add_row("Batch No. 2", format_val(record_data.get('batch2')))
    add_row("Batch No. 3", format_val(record_data.get('batch3')))
    add_row("Quantity", f"{format_val(record_data.get('quantity'))} pcs", font_size=20, y_inc=25)
    
    zpl.append(f"^FO30,{y}^GB460,1,1^FS")
    y += 8
    
    add_row("Shift SP", format_val(record_data.get('shift_sp')))
    add_row("Op ID", format_val(record_data.get('op_id')))
    add_row("Station", format_val(record_data.get('station')))
    dt_sp = format_val(record_data.get('dt_sp'))[:16] if record_data.get('dt_sp') else '-'
    add_row("SP Date/Time", dt_sp)
    dt_line = format_val(record_data.get('dt_line'))[:16] if record_data.get('dt_line') else '-'
    add_row("Line Entry", dt_line)
    
    printed = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    add_row("Printed At", printed)
    
    y += 3
    zpl.append(f"^FO30,{y}^GB460,2,2^FS")
    y += 8
    
    reprints = int(record_data.get('reprint_count') or 0)
    if reprints > 0:
      last_by = record_data.get('last_reprinted_by', 'Unknown')
      last_at = record_data.get('last_reprinted_at', '')[:16] if record_data.get('last_reprinted_at') else ''
      audit_text = f"** REPRINT #{reprints} by {last_by} at {last_at} **"
      zpl.append(f"^FO30,{y}^A0N,14,14^FB460,1,0,C^FD{audit_text}\\&^FS")
    else:
      zpl.append(f"^FO30,{y}^A0N,14,14^FB460,1,0,C^FD-- Original Print --\\&^FS")
    y += 18
    
  draw_layout(0)
  
  zpl.append(f"^FO30,728^A0N,20,20^FB460,1,0,C^FD- - - - - - - - - - - -\\&^FS")
  
  draw_layout(982)

  zpl.append("^PQ1")
  zpl.append("^XZ")
  return "\n".join(zpl).encode("utf-8")

def execute_zpl_print(zpl_string, record_data, silent=False):
  config = {}
  if os.path.exists(CONFIG_FILE):
    try:
      with open(CONFIG_FILE, "r") as f:
        config = json.load(f)
    except: pass
      
  printer_name = config.get("zebra_printer", "")
  
  try:
    qr_dir = persistent_path("qr")
    os.makedirs(qr_dir, exist_ok=True)
    sb_id_safe = str(record_data.get('sub_batch_id', 'unknown')).replace(":", "").replace("-", "")
    file_path = os.path.join(qr_dir, f"{sb_id_safe}.zpl")
    with open(file_path, "wb") as f:
      f.write(zpl_string)
  except Exception as e:
    print("Could not save ZPL file:", e)
    
  if not printer_name:
    if not silent:
      messagebox.showwarning("Setup Required", "ZPL saved to 'qr' folder. Please configure Zebra Printer in Settings to physically print.")
    return
    
  if not WIN32_PRINT_AVAILABLE:
    if not silent:
      messagebox.showerror("Error", "pywin32 is not installed. ZPL saved to 'qr' folder, but cannot print. Please run 'pip install pywin32'.")
    return
  
  try:
    hPrinter = win32print.OpenPrinter(printer_name)
    try:
      job = win32print.StartDocPrinter(hPrinter, 1, ("Sub-Process Label", None, "RAW"))
      win32print.StartPagePrinter(hPrinter)
      win32print.WritePrinter(hPrinter, zpl_string)
      win32print.EndPagePrinter(hPrinter)
      win32print.EndDocPrinter(hPrinter)
      if not silent:
        messagebox.showinfo("Success", f"Label sent to printer '{printer_name}' successfully!")
    finally:
      win32print.ClosePrinter(hPrinter)
  except Exception as e:
    if not silent:
      messagebox.showerror("Printer Error", f"Failed to send to printer '{printer_name}':\n{e}")

def print_html_slip(record_data, silent=False):
  zpl_string = generate_zpl(record_data)
  execute_zpl_print(zpl_string, record_data, silent)

def generate_pick_ticket_zpl(pn, target_qty, accumulated, picked):
  zpl = []
  zpl.append("^XA")
  zpl.append("^PW480")
  zpl.append("^LH20,0")
  zpl.append("^CI28")
  
  zpl.append("^FO0,30^A0N,30,30^FB440,1,0,C^FDFIFO PICK TICKET\\&^FS")
  zpl.append("^FO0,70^GB440,3,3^FS")
  
  zpl.append(f"^FO5,85^A0N,22,22^FDPart: {pn}^FS")
  zpl.append(f"^FO5,115^A0N,22,22^FDTarget: {target_qty} | Actual: {accumulated}^FS")
  
  y = 150
  table_start_y = y
  
  zpl.append(f"^FO0,{y}^GB440,2,2^FS")
  y += 5
  
  zpl.append(f"^FO5,{y}^A0N,20,20^FDSub-Batch ID^FS")
  zpl.append(f"^FO215,{y}^A0N,20,20^FDQty^FS")
  zpl.append(f"^FO280,{y}^A0N,20,20^FDStored Date^FS")
  y += 25
  
  zpl.append(f"^FO0,{y}^GB440,2,2^FS")
  
  for sb, q, dt in picked:
    y += 5
    zpl.append(f"^FO5,{y}^A0N,18,18^FD{sb}^FS")
    zpl.append(f"^FO215,{y}^A0N,18,18^FD{q}^FS")
    dt_short = dt[:16] if dt else ""
    zpl.append(f"^FO280,{y}^A0N,18,18^FD{dt_short}^FS")
    y += 25
    zpl.append(f"^FO0,{y}^GB440,1,1^FS")
    
  zpl.append(f"^FO0,{table_start_y}^GB2,{y - table_start_y},2^FS")
  zpl.append(f"^FO210,{table_start_y}^GB2,{y - table_start_y},2^FS")
  zpl.append(f"^FO270,{table_start_y}^GB2,{y - table_start_y},2^FS")
  zpl.append(f"^FO440,{table_start_y}^GB2,{y - table_start_y},2^FS")
  
  y += 15
  import datetime
  printed = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
  zpl.append(f"^FO5,{y}^A0N,18,18^FDPrinted: {printed}^FS")
  
  zpl.append("^PQ1")
  zpl.append("^XZ")
  return "\n".join(zpl).encode("utf-8")

def execute_ticket_print(zpl_string, pn, silent=False):
  config = {}
  if os.path.exists(CONFIG_FILE):
    try:
      with open(CONFIG_FILE, "r") as f:
        config = json.load(f)
    except: pass

  printer_name = config.get("zebra_printer", "")

  try:
    tickets_dir = persistent_path("tickets")
    os.makedirs(tickets_dir, exist_ok=True)
    safe_pn = str(pn).replace(":", "").replace("-", "")
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = os.path.join(tickets_dir, f"TICKET_{safe_pn}_{timestamp}.zpl")
    with open(file_path, "wb") as f:
      f.write(zpl_string)
  except Exception as e:
    print("Could not save Pick Ticket ZPL file:", e)

  if not printer_name:
    if not silent:
      messagebox.showwarning("Setup Required", "Ticket saved to 'tickets' folder.\\nPlease configure Zebra Printer in Settings to physically print.")
    return

  if not WIN32_PRINT_AVAILABLE:
    if not silent: messagebox.showerror("Error", "win32print module not found. Printing disabled.")
    return

  try:
    hPrinter = win32print.OpenPrinter(printer_name)
    try:
      hJob = win32print.StartDocPrinter(hPrinter, 1, ("Pick Ticket", None, "RAW"))
      try:
        win32print.StartPagePrinter(hPrinter)
        win32print.WritePrinter(hPrinter, zpl_string)
        win32print.EndPagePrinter(hPrinter)
      finally:
        win32print.EndDocPrinter(hPrinter)
      if not silent: messagebox.showinfo("Success", f"Pick ticket sent to printer '{printer_name}'")
    finally:
      win32print.ClosePrinter(hPrinter)
  except Exception as e:
    if not silent: messagebox.showerror("Printer Error", f"Failed to send to printer:\\n{e}")

def add_to_startup():
  try:
    import winreg
    import sys
    import os
    
    key = winreg.HKEY_CURRENT_USER
    key_path = r"Software\Microsoft\Windows\CurrentVersion\Run"
    
    with winreg.OpenKey(key, key_path, 0, winreg.KEY_ALL_ACCESS) as reg_key:
      app_name = "SubProcessTraceability"
      
      if getattr(sys, 'frozen', False):
        exe_path = f'"{sys.executable}"'
      else:
        exe_path = f'"{sys.executable}" "{os.path.abspath(sys.argv[0])}"'
        
      winreg.SetValueEx(reg_key, app_name, 0, winreg.REG_SZ, exe_path)
  except Exception as e:
    print(f"Could not add to startup: {e}")


class TraceabilityApp(tk.Tk):
  def __init__(self):
    super().__init__()
    self.title(f"HI-LEX ACT - Sub-Process Traceability System v{APP_VERSION}")
    self.geometry("1280x800")
    try:
      png_path = resource_path(os.path.join("assets", "new_main_app_logo.png"))
      if os.path.exists(png_path):
        self.iconphoto(True, tk.PhotoImage(file=png_path))
    except: pass
    try:
      self.state('zoomed')
    except:
      pass
    self.configure(bg=BG_COLOR)
    self.protocol("WM_DELETE_WINDOW", self.on_closing)
    self.schedule_daily_snapshot()
    self.after(5000, self.check_and_generate_missed_reports)
    # add_to_startup() removed to prevent automatic registry changes
    
    # Modern Windows 11 Title Bar Color (White background, Blue text)
    try:
      import ctypes
      self.update_idletasks()
      hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
      # DWMWA_CAPTION_COLOR = 35
      ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, ctypes.byref(ctypes.c_int(0x00FFFFFF)), 4)
      # DWMWA_TEXT_COLOR = 36 (HI-LEX Blue #005A8C -> 0x008C5A00)
      ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 36, ctypes.byref(ctypes.c_int(0x008C5A00)), 4)
    except Exception:
      pass
    
    self.option_add('*insertBackground', 'white')
    self.option_add('*Listbox.background', SURFACE_COLOR)
    self.option_add('*Listbox.foreground', TEXT_COLOR)
    self.option_add('*Listbox.selectBackground', ACCENT_COLOR)
    self.option_add('*Listbox.selectForeground', 'black')
    self.option_add('*TCombobox*Listbox.background', SURFACE_COLOR)
    self.option_add('*TCombobox*Listbox.foreground', TEXT_COLOR)
    self.option_add('*TCombobox*Listbox.selectBackground', ACCENT_COLOR)
    self.option_add('*TCombobox*Listbox.selectForeground', 'black')
    
    # Global Right-Click Menu for Entry widgets
    self.entry_menu = tk.Menu(self, tearoff=0, bg=SURFACE_COLOR, fg=TEXT_COLOR, activebackground=ACCENT_COLOR, activeforeground=TEXT_COLOR)
    self.entry_menu.add_command(label="Cut", command=lambda: self.focus_get().event_generate("<<Cut>>") if self.focus_get() else None)
    self.entry_menu.add_command(label="Copy", command=lambda: self.focus_get().event_generate("<<Copy>>") if self.focus_get() else None)
    self.entry_menu.add_command(label="Paste", command=lambda: self.focus_get().event_generate("<<Paste>>") if self.focus_get() else None)
    self.entry_menu.add_separator()
    self.entry_menu.add_command(label="Select All", command=lambda: self.focus_get().select_range(0, 'end') if hasattr(self.focus_get(), 'select_range') else None)

    def show_entry_menu(event):
      try:
        widget = event.widget
        if widget.winfo_class() in ('Entry', 'TEntry'):
          widget.focus()
          self.entry_menu.tk_popup(event.x_root, event.y_root)
      except Exception:
        pass

    self.bind_all("<Button-3>", show_entry_menu)
    
    # Configure Styles
    self.style = ttk.Style(self)
    self.style.theme_use('clam')
    self.style.configure('.', background=BG_COLOR, foreground=TEXT_COLOR, fieldbackground=SURFACE_COLOR)
    self.style.configure('TNotebook', background=BG_COLOR)
    self.style.configure('TNotebook.Tab', background=SURFACE_COLOR, foreground=TEXT_COLOR, padding=[20, 8], font=("Segoe UI", 11, "bold"))
    self.style.map('TNotebook.Tab', background=[('selected', ACCENT_COLOR)])
    self.style.configure('TFrame', background=BG_COLOR)
    self.style.configure('TLabel', background=SURFACE_COLOR, foreground=TEXT_COLOR)
    self.style.configure('TCheckbutton', background=SURFACE_COLOR, foreground=TEXT_COLOR)
    
    self.style.configure('TCombobox', fieldbackground=SURFACE_COLOR, background=SURFACE_COLOR, foreground=TEXT_COLOR)
    self.style.map('TCombobox', 
            fieldbackground=[('readonly', SURFACE_COLOR)], 
            selectbackground=[('readonly', ACCENT_COLOR)],
            selectforeground=[('readonly', '#000000')])
    
    self.style.configure('TEntry', fieldbackground=SURFACE_COLOR, foreground=TEXT_COLOR, insertcolor='white')
    self.style.configure('TButton', background=SURFACE_COLOR, foreground=TEXT_COLOR, font=HMI_FONT_M, padding=5)
    self.style.map('TButton', background=[('active', BORDER_COLOR)])
    
    self.style.configure('Success.TButton', background=SUCCESS_COLOR, foreground="#000000", font=HMI_FONT_M, padding=5)
    self.style.map('Success.TButton', background=[('disabled', '#4A5568'), ('active', STATUS_RUNNING)], foreground=[('disabled', '#A0AEC0')])
    
    self.style.configure('Primary.TButton', background=ACCENT_COLOR, foreground="#000000", font=HMI_FONT_M, padding=5)
    self.style.map('Primary.TButton', background=[('disabled', '#4A5568'), ('active', "#0096C7")], foreground=[('disabled', '#A0AEC0')])
    
    self.style.configure('Secondary.TButton', background=STATUS_IDLE, foreground="#FFFFFF", font=HMI_FONT_M, padding=5)
    self.style.map('Secondary.TButton', background=[('disabled', '#4A5568'), ('active', BORDER_COLOR)], foreground=[('disabled', '#A0AEC0')])
    
    self.style.configure('Header.TButton', background="#005A8C", foreground="#FFFFFF", font=HMI_FONT_M, padding=5)
    self.style.map('Header.TButton', background=[('disabled', '#4A5568'), ('active', "#00456B")], foreground=[('disabled', '#A0AEC0')])
    
    self.style.configure('Warning.TButton', background=WARNING_COLOR, foreground="#000000", font=HMI_FONT_M, padding=5)
    self.style.map('Warning.TButton', background=[('disabled', '#4A5568'), ('active', "#D97706")], foreground=[('disabled', '#A0AEC0')])
    
    self.style.configure('Danger.TButton', background=ERROR_COLOR, foreground="#000000", font=HMI_FONT_M, padding=5)
    self.style.map('Danger.TButton', background=[('disabled', '#4A5568'), ('active', "#B91C1C")], foreground=[('disabled', '#A0AEC0')])
    
    self.style.configure('Treeview', background=SURFACE_COLOR, foreground=TEXT_COLOR, fieldbackground=SURFACE_COLOR, rowheight=30, font=HMI_FONT_S)
    self.style.configure('Treeview.Heading', background=BORDER_COLOR, foreground=TEXT_COLOR, font=HMI_FONT_M)
    self.style.map('Treeview.Heading',
            background=[('active', ACCENT_COLOR)],
            foreground=[('active', '#000000')])
    
    self.app_user_id = ""
    self.app_user_shift = ""
    self.is_admin = False
    
    self.build_ui()
    self.after(100, self.prompt_login)
    self.update_clock()
    self.update_stats()
    self.populate_sf_combobox()
    self.refresh_recent_treeview()
    self.refresh_records_treeview()
    self.update_sub_batch_preview()

  def prompt_login(self):
    login_win = tk.Toplevel(self)
    login_win.title("HI-LEX Login")
    center_window(login_win, 400, 300)
    login_win.configure(bg=BG_COLOR)
    login_win.transient(self)
    login_win.grab_set()
    
    def on_login_close():
      login_win.destroy()
      self.destroy()
      
    login_win.protocol("WM_DELETE_WINDOW", on_login_close)
    
    tk.Label(login_win, text="Operator Login", font=HMI_FONT_L, bg=BG_COLOR, fg=TEXT_COLOR).pack(pady=15)
    
    frame = tk.Frame(login_win, bg=BG_COLOR)
    frame.pack(pady=10)
    
    tk.Label(frame, text="Operator ID:", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=0, column=0, padx=10, pady=10, sticky="e")
    vcmd_login = (login_win.register(lambda P: len(P) <= 7 and all(c.isalnum() or c == '-' for c in P)), '%P')
    entry_id = ttk.Entry(frame, width=20, validate="key", validatecommand=vcmd_login)
    entry_id.grid(row=0, column=1, padx=10, pady=10)
    entry_id.focus()
    
    tk.Label(frame, text="Password:", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=1, column=0, padx=10, pady=10, sticky="e")
    entry_pass = ttk.Entry(frame, width=20, show="*")
    entry_pass.grid(row=1, column=1, padx=10, pady=10)
    
    tk.Label(frame, text="Shift:", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=2, column=0, padx=10, pady=10, sticky="e")
    cb_shift = ttk.Combobox(frame, values=["A", "B", "C"], state="readonly", width=18)
    cb_shift.grid(row=2, column=1, padx=10, pady=10)
    cb_shift.set("A")
    
    def do_login(event=None):
      uid = entry_id.get().strip()
      upass = entry_pass.get()
      ushift = cb_shift.get().strip()
      
      if not uid or not upass or not ushift:
        messagebox.showerror("Error", "Please enter Operator ID, Password, and Shift.", parent=login_win)
        return
        
      try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT password, role FROM auth WHERE id = ?", (uid,))
        result = c.fetchone()
        conn.close()
        
        if result and not verify_password(upass, result[0]):
          result = None
        elif result:
          result = (result[1],)
      except Exception as e:
        messagebox.showerror("DB Error", f"Failed to connect to authentication database:\n{e}", parent=login_win)
        return
        
      if not result:
        messagebox.showerror("Error", "Invalid Operator ID or Password.", parent=login_win)
        return
        
      role = result[0]
      
      if role == "Quality OP":
        messagebox.showerror("Access Denied", "Quality personnel can only access the Quality Application.\n\nPlease use the Quality App.", parent=login_win)
        return
        
      self.is_admin = False
      display_name = uid
      
      if role in ("Supervisor", "Shift Leader", "Manager"):
        self.is_admin = True
        if uid == "mg90":
          display_name = "Taketo Oi (Manager)"
        else:
          display_name = role
        
      self.app_user_id = uid
      self.app_user_shift = ushift
      self.app_user_role = role
      self.lbl_header_user.config(text=f"User: {display_name} | Shift: {ushift}")
      
      try:
        conn = get_db_connection()
        c = conn.cursor()
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute("INSERT INTO system_access_logs (event_type, user_id, shift, timestamp) VALUES (?, ?, ?, ?)",
             ("LOGIN", uid, ushift, timestamp))
        conn.commit()
        conn.close()
      except Exception as e:
        pass
      
      if self.is_admin:
        if hasattr(self, 'lbl_header_warning'):
          self.lbl_header_warning.config(text=f" You are {display_name}, Please don't forget to logout!")
          self.lbl_header_warning.pack(side=tk.RIGHT, padx=20)
        if hasattr(self, 'notebook') and hasattr(self, 'tab3'):
          self.notebook.add(self.tab3, text="KPIs")
          if hasattr(self, 'refresh_kpis'):
            self.refresh_kpis()
        
      login_win.destroy()
      # Initialize default shift for manual entry
      self.clear_form()
      
    ttk.Button(login_win, text="Login", style="Success.TButton", command=do_login).pack(pady=15)
    login_win.bind("<Return>", do_login)

  def create_card(self, parent, title, fg_color=TEXT_COLOR):
    card = tk.Frame(parent, bg=SURFACE_COLOR)
    card.pack(fill=tk.X, pady=5, padx=5)
    
    title_lbl = tk.Label(card, text=title, bg=SURFACE_COLOR, fg=fg_color, font=HMI_FONT_M, anchor="w", padx=10, pady=5)
    title_lbl.pack(fill=tk.X)
    
    tk.Frame(card, bg=BORDER_COLOR, height=1).pack(fill=tk.X)
    
    content = tk.Frame(card, bg=SURFACE_COLOR, padx=10, pady=10)
    content.pack(fill=tk.BOTH, expand=True)
    return card, content

  def open_user_manager(self):
    top = tk.Toplevel(self)
    top.title("User Management (Admin)")
    center_window(top, 800, 500)
    top.resizable(False, False)
    
    top.update_idletasks()
    x = (top.winfo_screenwidth() // 2) - (800 // 2)
    y = (top.winfo_screenheight() // 2) - (500 // 2)
    top.geometry(f"+{x}+{y}")
    
    top.configure(bg=BG_COLOR)
    top.transient(self)
    top.grab_set()
    
    main_frame = tk.Frame(top, bg=BG_COLOR)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    left_frame = tk.Frame(main_frame, bg=BG_COLOR)
    left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 20))
    
    tk.Label(left_frame, text="Add New User", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L).pack(pady=(0, 20))
    
    tk.Label(left_frame, text="Operator ID / Name:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_S).pack(anchor="w")
    ent_id = ttk.Entry(left_frame, width=30)
    ent_id.pack(pady=(0, 15))
    
    tk.Label(left_frame, text="Password:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_S).pack(anchor="w")
    ent_pass = ttk.Entry(left_frame, width=30, show="*")
    ent_pass.pack(pady=(0, 15))
    
    tk.Label(left_frame, text="Role:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_S).pack(anchor="w")
    cb_role = ttk.Combobox(left_frame, values=["Operator", "Quality OP", "Shift Leader", "Supervisor", "Manager", "Admin"], state="readonly", width=28)
    cb_role.set("Operator")
    cb_role.pack(pady=(0, 20))
    
    right_frame = tk.Frame(main_frame, bg=BG_COLOR)
    right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
    
    tk.Label(right_frame, text="Current Users", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L).pack(pady=(0, 15))
    
    cols = ("ID", "Role")
    tree = ttk.Treeview(right_frame, columns=cols, show="headings")
    for col in cols:
      tree.heading(col, text=col)
    tree.column("ID", width=200)
    tree.column("Role", width=150)
    
    sb = ttk.Scrollbar(right_frame, orient="vertical", command=tree.yview)
    tree.configure(yscroll=sb.set)
    sb.pack(side=tk.RIGHT, fill=tk.Y)
    tree.pack(fill=tk.BOTH, expand=True)
    
    def load_users():
      for item in tree.get_children():
        tree.delete(item)
      try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT id, role FROM auth ORDER BY role")
        for row in c.fetchall():
          tree.insert("", "end", values=row)
        conn.close()
      except Exception as e:
        messagebox.showerror("Error", f"Failed to load users: {e}", parent=top)
        
    def add_user():
      uid = ent_id.get().strip()
      upass = ent_pass.get().strip()
      role = cb_role.get()
      if not uid or not upass:
        messagebox.showerror("Error", "ID and Password are required.", parent=top)
        return
      try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("INSERT INTO auth (id, password, role) VALUES (?, ?, ?)", (uid, hash_password(upass), role))
        conn.commit()
        conn.close()
        ent_id.delete(0, tk.END)
        ent_pass.delete(0, tk.END)
        load_users()
        messagebox.showinfo("Success", "User added successfully.", parent=top)
      except sqlite3.IntegrityError:
        messagebox.showerror("Error", "User ID already exists.", parent=top)
      except Exception as e:
        messagebox.showerror("Error", f"Failed to add user: {e}", parent=top)
        
    def delete_user():
      sel = tree.selection()
      if not sel:
        return
      uid = tree.item(sel[0])["values"][0]
      if uid in ["999", "998", "111", "mg90", self.app_user_id]:
        messagebox.showwarning("Restricted", "Cannot delete default or active users.", parent=top)
        return
      if messagebox.askyesno("Confirm Delete", f"Delete user '{uid}'?", parent=top):
        try:
          conn = get_db_connection()
          c = conn.cursor()
          c.execute("DELETE FROM auth WHERE id=?", (uid,))
          conn.commit()
          conn.close()
          load_users()
        except Exception as e:
          messagebox.showerror("Error", f"Failed to delete user: {e}", parent=top)

    btn_add = ttk.Button(left_frame, text="Add User", style="Success.TButton", command=add_user)
    btn_add.pack(fill=tk.X, pady=5)
    
    btn_del = ttk.Button(right_frame, text="Delete Selected", style="Danger.TButton", command=delete_user)
    btn_del.pack(side=tk.RIGHT, pady=10)
    
    load_users()

  def open_settings(self):
    top = tk.Toplevel(self)
    top.title("Settings")
    w = 600 if self.is_admin else 450
    h = 400 if self.is_admin else 200
    center_window(top, w, h)
    
    top.configure(bg=BG_COLOR)
    top.transient(self)
    top.grab_set()
    
    tk.Label(top, text="Printer Settings", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L).pack(pady=(10, 5))
    
    frame = tk.Frame(top, bg=BG_COLOR)
    frame.pack(pady=10)
    
    tk.Label(frame, text="Zebra Printer:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(side=tk.LEFT)
    
    printers = []
    if WIN32_PRINT_AVAILABLE:
      try:
        flags = 2 | 4
        printers = [p[2] for p in win32print.EnumPrinters(flags)]
      except Exception as e:
        pass
    
    cb = ttk.Combobox(frame, values=printers, state="readonly", width=35)
    cb.pack(side=tk.LEFT, padx=10)
    
    config = {}
    if os.path.exists(CONFIG_FILE):
      try:
        with open(CONFIG_FILE, "r") as f:
          config = json.load(f)
      except: pass
      
    if config.get("zebra_printer") in printers:
      cb.set(config.get("zebra_printer"))
    elif printers:
      cb.set(printers[0])
      
    def save():
      config["zebra_printer"] = cb.get()
      with open(CONFIG_FILE, "w") as f:
        json.dump(config, f)
      top.destroy()
      messagebox.showinfo("Saved", "Settings saved successfully.")
      
    ttk.Button(top, text="Save Settings", style="Success.TButton", command=save).pack(pady=10)
    
    if self.is_admin or self.app_user_role in ["Supervisor", "Manager"]:
      tk.Frame(top, bg=BORDER_COLOR, height=2).pack(fill=tk.X, padx=20, pady=10)
      tk.Label(top, text="Management Tools", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L).pack(pady=(5, 5))
      
      admin_frame = tk.Frame(top, bg=BG_COLOR)
      admin_frame.pack(fill=tk.X, padx=20, pady=10)
      
      ttk.Button(admin_frame, text="Manage Targets", style="Success.TButton", command=self.open_targets_manager).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
      if self.is_admin:
        ttk.Button(admin_frame, text="Manage Products", style="Primary.TButton", command=self.open_product_manager).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        ttk.Button(admin_frame, text="Audit Logs", style="Primary.TButton", command=self.open_logs_manager).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)
        ttk.Button(admin_frame, text="Manage Users", style="Warning.TButton", command=self.open_user_manager).pack(side=tk.LEFT, expand=True, fill=tk.X, padx=5)

  def open_targets_manager(self):
    top = tk.Toplevel(self)
    top.title("Production Targets Management")
    center_window(top, 800, 600)
    top.configure(bg=BG_COLOR)
    top.transient(self)
    top.grab_set()
    
    main_frame = tk.Frame(top, bg=BG_COLOR)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    form_frame = tk.LabelFrame(main_frame, text="Add New Target", bg=BG_COLOR, fg=TEXT_COLOR)
    form_frame.pack(fill=tk.X, pady=(0, 20))
    
    tk.Label(form_frame, text="Product PN:", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=0, column=0, padx=5, pady=5)
    var_pn = tk.StringVar()
    cb_pn = ttk.Combobox(form_frame, textvariable=var_pn, values=list(SF_DATA.keys()), state="normal")
    cb_pn.grid(row=0, column=1, padx=5, pady=5)
    
    tk.Label(form_frame, text="Target Qty:", bg=BG_COLOR, fg=TEXT_COLOR).grid(row=0, column=2, padx=5, pady=5)
    var_qty = tk.StringVar()
    entry_qty = ttk.Entry(form_frame, textvariable=var_qty)
    entry_qty.grid(row=0, column=3, padx=5, pady=5)
    
    def refresh_tree():
      for item in tree.get_children(): tree.delete(item)
      conn = get_db_connection()
      c = conn.cursor()
      c.execute("SELECT id, product_pn, target_qty, effective_date FROM shift_targets ORDER BY id DESC")
      for row in c.fetchall():
        tree.insert("", "end", values=row)
      conn.close()
      
    def add_target():
      pn = var_pn.get().strip()
      shift = "All"
      station = "All"
      qty_str = var_qty.get()
      
      if not pn or not qty_str.isdigit():
        messagebox.showerror("Error", "Invalid PN or Quantity")
        return
        
      qty = int(qty_str)
      now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
      
      conn = get_db_connection()
      c = conn.cursor()
      c.execute("INSERT INTO shift_targets (product_pn, shift, station, target_qty, effective_date) VALUES (?, ?, ?, ?, ?)",
           (pn, shift, station, qty, now_str))
      conn.commit()
      conn.close()
      
      var_qty.set("")
      refresh_tree()
      if hasattr(self, 'refresh_kpis'):
        self.refresh_kpis()
      
    def delete_target():
      selected = tree.selection()
      if not selected: return
      item = tree.item(selected[0])
      tid = item['values'][0]
      
      if messagebox.askyesno("Confirm", "Delete selected target?"):
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("DELETE FROM shift_targets WHERE id=?", (tid,))
        conn.commit()
        conn.close()
        refresh_tree()
        if hasattr(self, 'refresh_kpis'):
          self.refresh_kpis()
        
    ttk.Button(form_frame, text="Add Target", style="Success.TButton", command=add_target).grid(row=0, column=4, padx=10, pady=5)
    
    cols = ("ID", "PN", "Target Qty", "Effective Date")
    tree = ttk.Treeview(main_frame, columns=cols, show="headings")
    for c in cols: tree.heading(c, text=c)
    tree.pack(fill=tk.BOTH, expand=True, pady=10)
    
    ttk.Button(main_frame, text="Delete Selected", style="Danger.TButton", command=delete_target).pack(pady=5)
    
    refresh_tree()


  def open_product_manager(self):
    top = tk.Toplevel(self)
    top.title("Product Management (Admin)")
    center_window(top, 900, 600)
    top.resizable(False, False)
    
    top.update_idletasks()
    x = (top.winfo_screenwidth() // 2) - (900 // 2)
    y = (top.winfo_screenheight() // 2) - (600 // 2)
    top.geometry(f"+{x}+{y}")
    
    top.configure(bg=BG_COLOR)
    top.transient(self)
    top.grab_set()
    
    main_frame = tk.Frame(top, bg=BG_COLOR)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    left_frame = tk.Frame(main_frame, bg=BG_COLOR, width=300)
    left_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
    
    right_frame = tk.Frame(main_frame, bg=BG_COLOR)
    right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
    
    tk.Label(left_frame, text="Existing Products (SF_PN)", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(pady=(0, 10))
    
    listbox = tk.Listbox(left_frame, font=HMI_FONT_S, width=35, bg=SURFACE_COLOR, fg=TEXT_COLOR, selectbackground=ACCENT_COLOR)
    listbox.pack(fill=tk.BOTH, expand=True)
    
    def refresh_list():
      listbox.delete(0, tk.END)
      for pn in SF_DATA.keys():
        listbox.insert(tk.END, pn)
        
    refresh_list()
    
    tk.Label(right_frame, text="Product Details", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L).pack(pady=(0, 10))
    
    form_frame = tk.Frame(right_frame, bg=SURFACE_COLOR, padx=20, pady=20)
    form_frame.pack(fill=tk.BOTH, expand=True)
    
    top_form = tk.Frame(form_frame, bg=SURFACE_COLOR)
    top_form.pack(pady=10)
    
    tk.Label(top_form, text="SF Part Number (Key):", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).grid(row=0, column=0, sticky="e", pady=10, padx=5)
    var_sf_pn = tk.StringVar()
    entry_sf_pn = ttk.Entry(top_form, textvariable=var_sf_pn, width=30, font=HMI_FONT_M)
    entry_sf_pn.grid(row=0, column=1, sticky="w", pady=10)
    
    tk.Label(top_form, text="Part Name (SF):", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).grid(row=1, column=0, sticky="e", pady=10, padx=5)
    var_part_name = tk.StringVar()
    entry_part_name = ttk.Entry(top_form, textvariable=var_part_name, width=30, font=HMI_FONT_M)
    entry_part_name.grid(row=1, column=1, sticky="w", pady=10)
    
    tk.Frame(form_frame, height=2, bg=BORDER_COLOR).pack(fill=tk.X, pady=15)
    
    bottom_form = tk.Frame(form_frame, bg=SURFACE_COLOR)
    bottom_form.pack(pady=5)
    
    tk.Label(bottom_form, text="Raw Materials (Up to 4)", bg=SURFACE_COLOR, fg=TEXT_MUTED, font=HMI_FONT_S).grid(row=0, column=0, columnspan=4, sticky="w", pady=5)
    
    rm_vars = []
    rm_entries = []
    for i in range(4):
      tk.Label(bottom_form, text=f"RM {i+1} PN:", bg=SURFACE_COLOR, fg=TEXT_COLOR).grid(row=1+i, column=0, sticky="e", pady=5, padx=5)
      v_pn = tk.StringVar()
      e1 = ttk.Entry(bottom_form, textvariable=v_pn, width=25)
      e1.grid(row=1+i, column=1, sticky="w", pady=5)
      
      tk.Label(bottom_form, text=f"RM {i+1} Name:", bg=SURFACE_COLOR, fg=TEXT_COLOR).grid(row=1+i, column=2, sticky="e", pady=5, padx=5)
      v_name = tk.StringVar()
      e2 = ttk.Entry(bottom_form, textvariable=v_name, width=25)
      e2.grid(row=1+i, column=3, sticky="w", pady=5)
      
      rm_vars.append((v_pn, v_name))
      rm_entries.append((e1, e2))
      
    def set_form_state(state):
      entry_sf_pn.config(state=state)
      entry_part_name.config(state=state)
      for e1, e2 in rm_entries:
        e1.config(state=state)
        e2.config(state=state)
        
    def unlock_form():
      if messagebox.askyesno("Confirm Unlock", "Are you sure you want to unlock and edit this existing product?", parent=top):
        set_form_state("normal")
        btn_unlock.config(state="disabled")
        btn_save.config(state="normal")
        btn_delete.config(state="normal")
        btn_clear.config(state="normal")
      
    def clear_form():
      set_form_state("normal")
      var_sf_pn.set("")
      var_part_name.set("")
      for v_pn, v_name in rm_vars:
        v_pn.set("")
        v_name.set("")
      listbox.selection_clear(0, tk.END)
      btn_unlock.config(state="disabled")
      btn_save.config(state="normal")
      btn_delete.config(state="disabled")
      btn_clear.config(state="normal")
      
    def on_select(evt):
      if not listbox.curselection(): return
      sel_pn = listbox.get(listbox.curselection())
      var_sf_pn.set(sel_pn)
      val = SF_DATA[sel_pn]
      var_part_name.set(val[0])
      rms = val[1]
      for i in range(4):
        if i < len(rms):
          rm_vars[i][0].set(rms[i][0])
          rm_vars[i][1].set(rms[i][1])
        else:
          rm_vars[i][0].set("")
          rm_vars[i][1].set("")
          
      set_form_state("disabled")
      btn_unlock.config(state="normal")
      btn_save.config(state="disabled")
      btn_delete.config(state="disabled")
      btn_clear.config(state="normal")
          
    listbox.bind("<<ListboxSelect>>", on_select)
    
    def save_product():
      pn = var_sf_pn.get().strip()
      name = var_part_name.get().strip()
      if not pn or not name:
        messagebox.showerror("Error", "SF Part Number and Name are required.", parent=top)
        return
        
      is_new = pn not in SF_DATA
      msg = f"Are you sure you want to add the new product '{pn}'?" if is_new else f"Are you sure you want to update the existing product '{pn}'?"
      if not messagebox.askyesno("Confirm Save", msg, parent=top):
        return
      
      rms = []
      for v_pn, v_name in rm_vars:
        r_pn = v_pn.get().strip()
        r_name = v_name.get().strip()
        if r_pn and r_name:
          rms.append((r_pn, r_name))
      
      SF_DATA[pn] = (name, rms)
      
      try:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO products (pn_sf, name_sf, rms_json) VALUES (?, ?, ?)", (pn, name, json.dumps(rms)))
        
        action = "ADD" if is_new else "UPDATE"
        details = f"Name: {name}, RMs: {len(rms)}"
        c.execute("INSERT INTO product_audit_trail (action, pn_sf, details, user_id, shift, timestamp) VALUES (?, ?, ?, ?, ?, ?)",
             (action, pn, details, getattr(self, 'app_user_id', 'Unknown'), getattr(self, 'app_user_shift', 'Unknown'), timestamp))
        conn.commit()
        conn.close()
      except Exception as e:
        messagebox.showerror("DB Error", f"Failed to update database:\n{e}", parent=top)
        
      self.populate_sf_combobox()
      if hasattr(self, 'pl_cb_sf_pn'):
        self.pl_cb_sf_pn['values'] = list(SF_DATA.keys())
      refresh_list()
      messagebox.showinfo("Success", f"Product '{pn}' saved successfully.", parent=top)
      
      # Auto-lock form and allow creating new product
      set_form_state("disabled")
      btn_unlock.config(state="normal")
      btn_save.config(state="disabled")
      btn_delete.config(state="disabled")
      btn_clear.config(state="normal")
      
    def delete_product():
      pn = var_sf_pn.get().strip()
      if not pn: return
      if pn in SF_DATA:
        if messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete product '{pn}'?", parent=top):
          del SF_DATA[pn]
          
          try:
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            conn = get_db_connection()
            c = conn.cursor()
            c.execute("DELETE FROM products WHERE pn_sf = ?", (pn,))
            c.execute("INSERT INTO product_audit_trail (action, pn_sf, details, user_id, shift, timestamp) VALUES (?, ?, ?, ?, ?, ?)",
                 ("DELETE", pn, "", getattr(self, 'app_user_id', 'Unknown'), getattr(self, 'app_user_shift', 'Unknown'), timestamp))
            conn.commit()
            conn.close()
          except Exception as e:
            messagebox.showerror("DB Error", f"Failed to delete from database:\n{e}", parent=top)
            
          self.populate_sf_combobox()
          if hasattr(self, 'pl_cb_sf_pn'):
            self.pl_cb_sf_pn['values'] = list(SF_DATA.keys())
          refresh_list()
          clear_form()
          messagebox.showinfo("Deleted", f"Product '{pn}' deleted.", parent=top)
      
    btn_frame = tk.Frame(right_frame, bg=BG_COLOR)
    btn_frame.pack(fill=tk.X, pady=15)
    
    btn_clear = ttk.Button(btn_frame, text="Add New Product ", style="Secondary.TButton", command=clear_form)
    btn_clear.pack(side=tk.LEFT, padx=5)
    
    btn_unlock = ttk.Button(btn_frame, text="Unlock ", style="Danger.TButton", command=unlock_form)
    btn_unlock.pack(side=tk.LEFT, padx=5)
    btn_unlock.config(state="disabled")
    
    btn_delete = ttk.Button(btn_frame, text="Delete Product", style="Warning.TButton", command=delete_product)
    btn_delete.pack(side=tk.RIGHT, padx=5)
    btn_delete.config(state="disabled")
    
    btn_save = ttk.Button(btn_frame, text="Save Product", style="Success.TButton", command=save_product)
    btn_save.pack(side=tk.RIGHT, padx=5)

  def open_logs_manager(self):
    top = tk.Toplevel(self)
    top.title("System Audit Logs")
    center_window(top, 1000, 600)
    top.resizable(False, False)
    
    top.update_idletasks()
    x = (top.winfo_screenwidth() // 2) - (1000 // 2)
    y = (top.winfo_screenheight() // 2) - (600 // 2)
    top.geometry(f"+{x}+{y}")
    
    top.configure(bg=BG_COLOR)
    top.transient(self)
    top.grab_set()
    
    main_frame = tk.Frame(top, bg=BG_COLOR)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    tk.Label(main_frame, text="System Audit Logs", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L).pack(pady=(0, 15))
    
    log_notebook = ttk.Notebook(main_frame)
    log_notebook.pack(fill=tk.BOTH, expand=True)
    
    tab_products = ttk.Frame(log_notebook)
    tab_access = ttk.Frame(log_notebook)
    
    log_notebook.add(tab_products, text="Product Modifications")
    log_notebook.add(tab_access, text="System Access Logs")
    
    cols_prod = ("ID", "Action", "Part Number", "Details", "User ID", "Shift", "Timestamp")
    tree_prod = ttk.Treeview(tab_products, columns=cols_prod, show="headings")
    for col in cols_prod:
      tree_prod.heading(col, text=col)
      tree_prod.column(col, width=120)
    tree_prod.column("Details", width=250)
    tree_prod.column("ID", width=50)
    
    sb_prod = ttk.Scrollbar(tab_products, orient="vertical", command=tree_prod.yview)
    tree_prod.configure(yscroll=sb_prod.set)
    sb_prod.pack(side=tk.RIGHT, fill=tk.Y)
    tree_prod.pack(fill=tk.BOTH, expand=True)
    
    cols_access = ("ID", "Event", "User ID", "Shift", "Timestamp")
    tree_access = ttk.Treeview(tab_access, columns=cols_access, show="headings")
    for col in cols_access:
      tree_access.heading(col, text=col)
      tree_access.column(col, width=150)
    tree_access.column("ID", width=50)
    
    sb_access = ttk.Scrollbar(tab_access, orient="vertical", command=tree_access.yview)
    tree_access.configure(yscroll=sb_access.set)
    sb_access.pack(side=tk.RIGHT, fill=tk.Y)
    tree_access.pack(fill=tk.BOTH, expand=True)
    
    def refresh_logs():
      for item in tree_prod.get_children():
        tree_prod.delete(item)
      for item in tree_access.get_children():
        tree_access.delete(item)
        
      try:
        conn = get_db_connection()
        c = conn.cursor()
        c.execute("SELECT id, action, pn_sf, details, user_id, shift, timestamp FROM product_audit_trail ORDER BY id DESC")
        for row in c.fetchall():
          tree_prod.insert("", "end", values=row)
          
        try:
          c.execute("SELECT id, event_type, user_id, shift, timestamp FROM system_access_logs ORDER BY id DESC")
          for row in c.fetchall():
            tree_access.insert("", "end", values=row)
        except sqlite3.OperationalError:
          pass
        conn.close()
      except Exception as e:
        messagebox.showerror("Error", f"Failed to fetch logs: {e}", parent=top)
        
    btn_frame = tk.Frame(main_frame, bg=BG_COLOR)
    btn_frame.pack(fill=tk.X, pady=(10,0))
    ttk.Button(btn_frame, text="Refresh", style="Secondary.TButton", command=refresh_logs).pack(side=tk.RIGHT)
    
    refresh_logs()

  def build_ui(self):
    HEADER_BG = "#FFFFFF"
    HEADER_FG = "#005A8C"
    
    self.header_frame = tk.Frame(self, bg=HEADER_BG, height=70)
    self.header_frame.pack(side=tk.TOP, fill=tk.X)
    self.header_frame.pack_propagate(False)
    
    try:
      logo_path = resource_path(os.path.join("assets", "logo_en.png"))
      if os.path.exists(logo_path):
        try:
          from PIL import Image, ImageTk
          img = Image.open(logo_path)
          
          ratio = 50 / img.height
          new_size = (int(img.width * ratio), 50)
          if hasattr(Image, "Resampling"):
            res_filter = Image.Resampling.LANCZOS
          else:
            res_filter = Image.ANTIALIAS
          img = img.resize(new_size, res_filter)
          self.logo_img = ImageTk.PhotoImage(img)
          tk.Label(self.header_frame, image=self.logo_img, bg=HEADER_BG).pack(side=tk.LEFT, padx=(20, 10))
        except Exception as pil_e:
          self.logo_img = tk.PhotoImage(file=logo_path)
          tk.Label(self.header_frame, image=self.logo_img, bg=HEADER_BG).pack(side=tk.LEFT, padx=(20, 10))
      
      tk.Label(self.header_frame, text="SUB-PROCESS TRACEABILITY", bg=HEADER_BG, fg=HEADER_FG, font=HMI_FONT_L).pack(side=tk.LEFT)
    except Exception as e:
      tk.Label(self.header_frame, text="HI-LEX ACT - SUB-PROCESS TRACEABILITY", bg=HEADER_BG, fg=HEADER_FG, font=HMI_FONT_L).pack(side=tk.LEFT, padx=20)
    
    self.settings_btn = ttk.Button(self.header_frame, text="Settings", style="Header.TButton", command=self.open_settings)
    self.settings_btn.pack(side=tk.RIGHT, padx=20)
    
    def do_logout():
      self.perform_logout(force=False)
        
    self.logout_btn = ttk.Button(self.header_frame, text="Logout", style="Danger.TButton", command=do_logout)
    self.logout_btn.pack(side=tk.RIGHT, padx=5)
    
    self.lbl_header_user = tk.Label(self.header_frame, text="User: - | Shift: -", bg=HEADER_BG, fg=TEXT_MUTED, font=HMI_FONT_S)
    self.lbl_header_user.pack(side=tk.RIGHT, padx=10)
    
    self.lbl_header_warning = tk.Label(self.header_frame, text=" Please don't forget to logout!", bg=HEADER_BG, fg=ERROR_COLOR, font=HMI_FONT_M)
    
    self.lbl_quality_alert = tk.Label(self.header_frame, text=" QUALITY ALERT: High Defect Rate!", bg=HEADER_BG, fg=ERROR_COLOR, font=HMI_FONT_M)
    
    #tk.Label(self.header_frame, text="Developed by Naoufal El Hlou", bg=HEADER_BG, fg="#94A3B8", font=("Segoe UI", 9, "italic")).pack(side=tk.RIGHT, padx=10)
    
    try:
      settings_path = resource_path(os.path.join("assets", "settings_icon.png"))
      if os.path.exists(settings_path):
        from PIL import Image, ImageTk
        s_img = Image.open(settings_path).convert("RGBA")
        data = s_img.getdata()
        new_data = []
        for item in data:
          if item[3] > 10:
            new_data.append((255, 255, 255, item[3]))
          else:
            new_data.append((255, 255, 255, 0))
        s_img.putdata(new_data)
        
        if hasattr(Image, "Resampling"):
          res_filter = Image.Resampling.LANCZOS
        else:
          res_filter = Image.ANTIALIAS
        s_img = s_img.resize((20, 20), res_filter)
        self.settings_icon = ImageTk.PhotoImage(s_img)
        self.settings_btn.config(image=self.settings_icon, compound=tk.LEFT, text=" Settings")
    except Exception as e:
      print("Could not load settings icon:", e)
    
    # PanedWindow
    self.paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
    self.paned.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    # Sidebar
    self.sidebar = tk.Frame(self.paned, bg=SURFACE_COLOR, width=180)
    self.sidebar.pack_propagate(False)
    self.paned.add(self.sidebar, weight=0)
    
    self.lbl_clock = tk.Label(self.sidebar, text="00:00", bg=SURFACE_COLOR, fg=ACCENT_COLOR, font=("Consolas", 28, "bold"))
    self.lbl_clock.pack(pady=(20,0))
    self.lbl_date = tk.Label(self.sidebar, text="YYYY-MM-DD", bg=SURFACE_COLOR, fg=TEXT_MUTED, font=HMI_FONT_S)
    self.lbl_date.pack(pady=(0, 20))

    
    tk.Label(self.sidebar, text="SHIFT OVERVIEW", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack()
    
    def create_side_kpi(parent, title):
      f = tk.Frame(parent, bg=BG_COLOR, highlightbackground=BORDER_COLOR, highlightthickness=1)
      f.pack(fill=tk.X, padx=15, pady=5)
      lbl_title = tk.Label(f, text=title, bg=BG_COLOR, fg=TEXT_MUTED, font=HMI_FONT_S)
      lbl_title.pack(pady=(5,0))
      lbl_val = tk.Label(f, text="0", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L)
      lbl_val.pack(pady=(0,5))
      return lbl_title, lbl_val
      
    self.lbl_title_recs, self.side_stat_recs = create_side_kpi(self.sidebar, "Records (Shift -)")
    self.lbl_title_qty, self.side_stat_qty = create_side_kpi(self.sidebar, "Qty (Shift -)")
    _, self.side_stat_today = create_side_kpi(self.sidebar, "Total Today")
    
    tk.Label(self.sidebar, text="RECENT PNs", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(pady=(10,0))
    self.recent_pns_listbox = tk.Listbox(self.sidebar, bg=BG_COLOR, fg=TEXT_COLOR, bd=0, relief="flat", height=8, font=HMI_FONT_S)
    self.recent_pns_listbox.pack(fill=tk.X, padx=15, pady=10)
    
    ttk.Button(self.sidebar, text="Open Excel", style="Secondary.TButton", command=self.open_excel).pack(fill=tk.X, padx=10, pady=2)
    ttk.Button(self.sidebar, text="Print Last Slip", style="Warning.TButton", command=self.print_last_slip).pack(fill=tk.X, padx=10, pady=2)
    
    # Notebook
    self.notebook = ttk.Notebook(self.paned)
    self.paned.add(self.notebook, weight=1)
    
    self.tab1 = ttk.Frame(self.notebook)
    self.tab5 = ttk.Frame(self.notebook)
    self.tab4 = ttk.Frame(self.notebook)
    self.tab2 = ttk.Frame(self.notebook)
    self.tab_inventory = ttk.Frame(self.notebook)
    self.tab3 = ttk.Frame(self.notebook)
    
    self.notebook.add(self.tab1, text="New Entry")
    self.notebook.add(self.tab5, text="Consume to Line")
    self.notebook.add(self.tab4, text="Deep Traceability")
    self.notebook.add(self.tab2, text="Records")
    self.notebook.add(self.tab_inventory, text="Live Inventory")
    self.notebook.add(self.tab3, text="KPIs")
    self.notebook.hide(self.tab3)
    
    self.build_tab1()
    self.build_tab5()
    self.build_tab4()
    self.build_tab2()
    self.build_tab_inventory()
    self.build_tab3()
    
    self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_changed)
    self._check_quality_rate()

  def perform_logout(self, force=False):
    if not force:
      if not messagebox.askyesno("Confirm Logout", "Are you sure you want to log out?", parent=self):
        return
        
    if hasattr(self, 'app_user_id') and self.app_user_id:
      try:
        conn = get_db_connection()
        c = conn.cursor()
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c.execute("INSERT INTO system_access_logs (event_type, user_id, shift, timestamp) VALUES (?, ?, ?, ?)",
             ("LOGOUT", self.app_user_id, self.app_user_shift, timestamp))
        conn.commit()
        conn.close()
      except Exception as e:
        pass

    self.app_user_id = ""
    self.app_user_shift = ""
    self.is_admin = False
    self.lbl_header_user.config(text="User: - | Shift: -")
    if hasattr(self, 'lbl_header_warning'):
      self.lbl_header_warning.pack_forget()
    if hasattr(self, 'notebook') and hasattr(self, 'tab3'):
      self.notebook.hide(self.tab3)
    self.prompt_login()
    
  def _on_tab_changed(self, event):
    try:
      selected_tab = self.notebook.tab(self.notebook.select(), "text")
      if selected_tab == "Records":
        self.refresh_records_treeview()
      elif selected_tab == "Live Inventory":
        self.refresh_inventory()
      elif selected_tab == "KPIs":
        self.refresh_kpis()
    except Exception:
      pass
      
  def _check_quality_rate(self):
    def _fetch_data():
      try:
        conn = get_db_connection()
        c = conn.cursor()
        today = datetime.datetime.now().strftime("%Y-%m-%d")
        
        c.execute("SELECT COUNT(*) FROM quality_defects WHERE reported_at LIKE ?", (f"{today}%",))
        krow = c.fetchone()
        total_def_today = krow[0] or 0
        
        c.execute("SELECT SUM(quantity) FROM records WHERE dt_sp LIKE ?", (f"{today}%",))
        prod_res = c.fetchone()
        total_prod = prod_res[0] or 0
        
        conn.close()
        return total_def_today, total_prod
      except Exception:
        return 0, 0

    def _update_ui(result):
      total_def_today, total_prod = result
      if total_prod > 0:
        rate = (total_def_today / total_prod) * 100
        if rate > 3.0:
          self.lbl_quality_alert.pack(side=tk.LEFT, padx=20)
        else:
          self.lbl_quality_alert.pack_forget()
      else:
        self.lbl_quality_alert.pack_forget()

    def _run():
      result = _fetch_data()
      self.after(0, lambda: _update_ui(result))

    import threading
    threading.Thread(target=_run, daemon=True).start()

    try:
      self.after(60000, self._check_quality_rate)
    except Exception:
      pass
  def build_tab1(self):
    # Container Split
    top_panel = tk.Frame(self.tab1, bg=BG_COLOR)
    top_panel.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    
    bottom_panel = tk.Frame(self.tab1, bg=BG_COLOR)
    bottom_panel.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=5)

    # 2-Column Layout
    self.t1_left = tk.Frame(top_panel, bg=BG_COLOR, width=400)
    self.t1_left.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=5)
    self.t1_left.pack_propagate(False) # Keep fixed width for left panel
    
    separator = ttk.Separator(top_panel, orient='vertical')
    separator.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=10)
    
    self.t1_right = tk.Frame(top_panel, bg=BG_COLOR)
    self.t1_right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=5)

    # ---------------- LEFT PANEL ----------------
    # Shift Target Tracker
    card_tracker, tracker_frame = self.create_card(self.t1_left, "Shift Target Tracker", fg_color=ACCENT_COLOR)
    card_tracker.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=5, padx=5)
    
    self.tracker_canvas = tk.Canvas(tracker_frame, bg=SURFACE_COLOR, highlightthickness=0)
    tracker_scrollbar = ttk.Scrollbar(tracker_frame, orient="vertical", command=self.tracker_canvas.yview)
    
    self.tracker_scroll_frame = tk.Frame(self.tracker_canvas, bg=SURFACE_COLOR)
    
    self.tracker_scroll_frame.bind(
      "<Configure>",
      lambda e: self.tracker_canvas.configure(scrollregion=self.tracker_canvas.bbox("all"))
    )
    
    self.tracker_canvas.create_window((0, 0), window=self.tracker_scroll_frame, anchor="nw", width=330)
    self.tracker_canvas.configure(yscrollcommand=tracker_scrollbar.set)
    
    self.tracker_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=2, pady=5)
    tracker_scrollbar.pack(side=tk.RIGHT, fill=tk.Y, pady=5)
    
    style = ttk.Style()
    style.configure("TProgressbar", thickness=15)
    style.configure("Safe.Horizontal.TProgressbar", background=SUCCESS_COLOR, troughcolor=BG_COLOR, thickness=15)
    style.configure("Warn.Horizontal.TProgressbar", background=WARNING_COLOR, troughcolor=BG_COLOR, thickness=15)
    style.configure("Danger.Horizontal.TProgressbar", background=ERROR_COLOR, troughcolor=BG_COLOR, thickness=15)
    
    self.lbl_scroll_indicator = tk.Label(self.t1_left, text=" Scroll down for more refs ", bg=BG_COLOR, fg=TEXT_MUTED, font=HMI_FONT_S)
    self.lbl_scroll_indicator.pack(side=tk.TOP)
    
    # Quick Search
    card_search, search_frame = self.create_card(self.t1_left, "Quick Search")
    card_search.pack(side=tk.BOTTOM, fill=tk.X, pady=5, padx=5)
    self.sv_search = tk.StringVar()
    self.sv_search.trace_add("write", self.filter_sf_combobox)
    search_entry = ttk.Entry(search_frame, textvariable=self.sv_search, width=20)
    search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
    self.lbl_search_count = tk.Label(search_frame, text="30", fg=ACCENT_COLOR, bg=SURFACE_COLOR)
    self.lbl_search_count.pack(side=tk.LEFT, padx=5)
    # ttk.Button(search_frame, text="​", width=3, command=lambda: self.sv_search.set("")).pack(side=tk.LEFT)
    
    # Buttons moved to t1_right
    
    # ---------------- BOTTOM PANEL ----------------
    # Recent Records Tree
    _, tree_frame = self.create_card(bottom_panel, "Last 5 Records (Double-click to print)")
    
    cols = ("SB_ID", "SF_PN", "Qty", "Shift", "Station", "DateTime")
    self.tree_recent = ttk.Treeview(tree_frame, columns=cols, show="headings", height=5)
    for c in cols:
      self.tree_recent.heading(c, text=c)
      self.tree_recent.column(c, width=100)
    self.tree_recent.pack(fill=tk.X)
    self.tree_recent.bind("<Double-1>", self.on_recent_double_click)
    
    self.tree_recent.tag_configure("even", background="#141A20")
    self.tree_recent.tag_configure("odd", background="#1B232C")
    

    # ---------------- RIGHT PANEL (FORM) ----------------
    self.form_frame = tk.Frame(self.t1_right, bg=BG_COLOR)
    self.form_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=20, pady=10)
    
    self.form_col_left = tk.Frame(self.form_frame, bg=BG_COLOR)
    self.form_col_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
    
    self.form_col_right = tk.Frame(self.form_frame, bg=BG_COLOR)
    self.form_col_right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
    
    # Section 1: Part Reference (Left Column)
    _, self.lf1 = self.create_card(self.form_col_left, "Part Reference")
    
    ttk.Label(self.lf1, text="Scan Main RM Barcode", width=20, font=HMI_FONT_S).grid(row=0, column=0, sticky="w", pady=2, padx=5)
    self.var_scan_rm_t1 = tk.StringVar()
    self.entry_scan_rm_t1 = ttk.Entry(self.lf1, textvariable=self.var_scan_rm_t1)
    self.entry_scan_rm_t1.grid(row=0, column=1, sticky="ew", pady=2, padx=5)
    self.entry_scan_rm_t1.bind("<Return>", self.on_rm_scanned_t1)
    
    self.lbl_scan_rm_status = tk.Label(self.lf1, text="", bg=SURFACE_COLOR, font=HMI_FONT_S)
    self.lbl_scan_rm_status.grid(row=1, column=0, columnspan=2, sticky="w", padx=10)
    
    ttk.Label(self.lf1, text="FULL PN Semi fini", width=20, font=HMI_FONT_S).grid(row=2, column=0, sticky="w", pady=2, padx=5)
    self.cb_sf_pn = ttk.Combobox(self.lf1, state="readonly")
    self.cb_sf_pn.grid(row=2, column=1, sticky="ew", pady=2, padx=5)
    self.cb_sf_pn.bind("<<ComboboxSelected>>", self.on_sf_selected)
    
    ttk.Label(self.lf1, text="PART NAME (SF)", width=20, font=HMI_FONT_S).grid(row=3, column=0, sticky="w", pady=2, padx=5)
    self.var_part_sf = tk.StringVar()
    ttk.Entry(self.lf1, textvariable=self.var_part_sf, state="disabled").grid(row=3, column=1, sticky="ew", pady=2, padx=5)
    
    self.rm_widgets = []
    self.rm_vars_t1 = []
    
    self.lf1.columnconfigure(1, weight=1)
    
    # Section 2: Batch & Qty
    _, lf2 = self.create_card(self.form_col_right, "Batch Numbers & Quantity")
    
    ttk.Label(lf2, text="Batch No. (1/2/3)", font=HMI_FONT_S, width=20).grid(row=0, column=0, sticky="w", pady=2, padx=5)
    b_frame = tk.Frame(lf2, bg=SURFACE_COLOR)
    b_frame.grid(row=0, column=1, sticky="ew", pady=2, padx=5)
    self.var_b1 = tk.StringVar()
    self.var_b2 = tk.StringVar()
    self.var_b3 = tk.StringVar()
    self.var_b1.trace_add("write", lambda *args: self.var_b1.set(self.var_b1.get().upper()))
    self.var_b2.trace_add("write", lambda *args: self.var_b2.set(self.var_b2.get().upper()))
    self.var_b3.trace_add("write", lambda *args: self.var_b3.set(self.var_b3.get().upper()))
    ttk.Entry(b_frame, textvariable=self.var_b1).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))
    ttk.Entry(b_frame, textvariable=self.var_b2).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
    ttk.Entry(b_frame, textvariable=self.var_b3).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
    
    ttk.Label(lf2, text="Quantity", font=HMI_FONT_S, width=20).grid(row=1, column=0, sticky="w", pady=2, padx=5)
    q_frame = tk.Frame(lf2, bg=SURFACE_COLOR)
    q_frame.grid(row=1, column=1, sticky="ew", pady=2, padx=5)
    self.var_qty = tk.StringVar()
    qty_entry = ttk.Entry(q_frame, textvariable=self.var_qty, font=HMI_FONT_M)
    qty_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2))
    ttk.Label(q_frame, text="pcs", font=HMI_FONT_S).pack(side=tk.LEFT, padx=5)
    
    lf2.columnconfigure(1, weight=1)
    
    # Section 3: Operation Details
    _, lf3 = self.create_card(self.form_col_right, "Operation Details")
    
    ttk.Label(lf3, text="Shift", font=HMI_FONT_S, width=20).grid(row=0, column=0, sticky="w", pady=2, padx=5)
    
    op_frame = tk.Frame(lf3, bg=SURFACE_COLOR)
    op_frame.grid(row=0, column=1, sticky="ew", pady=2, padx=5)
    
    self.cb_shift_sp = ttk.Combobox(op_frame, values=["A", "B", "C"], state="readonly", width=5)
    self.cb_shift_sp.pack(side=tk.LEFT, padx=(0, 10))
    self.cb_shift_sp.set("")
    self.cb_shift_sp.bind("<<ComboboxSelected>>", lambda e: [self.update_sub_batch_preview(), self.update_stats()])
    
    ttk.Label(op_frame, text="Op ID", font=HMI_FONT_S).pack(side=tk.LEFT, padx=(5, 5))
    self.var_op_id = tk.StringVar()
    vcmd_op = (self.register(lambda P: len(P) <= 7 and all(c.isalnum() or c == '-' for c in P)), '%P')
    ttk.Entry(op_frame, textvariable=self.var_op_id, width=12, validate="key", validatecommand=vcmd_op).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5)
    
    ttk.Label(op_frame, text="Station", font=HMI_FONT_S).pack(side=tk.LEFT, padx=(10, 5))
    self.cb_station = ttk.Combobox(op_frame, values=["S06", "S07", "S10", "S11"], state="readonly", width=5)
    self.cb_station.pack(side=tk.LEFT, padx=5)
    self.cb_station.bind("<<ComboboxSelected>>", lambda e: self.update_sub_batch_preview())
    
    lf3.columnconfigure(1, weight=1)
    self.cb_station.bind("<<ComboboxSelected>>", lambda e: self.update_sub_batch_preview())
    
    # Section 4: Date & Time
    _, lf4 = self.create_card(self.form_col_right, "Date & Time")
    
    def create_dt_picker(parent, label_text):
      ttk.Label(parent, text=label_text, font=HMI_FONT_S, width=20).grid(row=0, column=0, sticky="w", pady=2, padx=5)
      
      frame = tk.Frame(parent, bg=SURFACE_COLOR)
      frame.grid(row=0, column=1, sticky="ew", pady=2, padx=5)
      
      de = DateEntry(frame, width=12, background=ACCENT_COLOR, foreground='white', borderwidth=2)
      de.pack(side=tk.LEFT, padx=2)
      h_spin = tk.Spinbox(frame, from_=0, to=23, wrap=True, width=3, format="%02.0f", bg=SURFACE_COLOR, fg=TEXT_COLOR)
      h_spin.pack(side=tk.LEFT, padx=2)
      ttk.Label(frame, text=":").pack(side=tk.LEFT)
      m_spin = tk.Spinbox(frame, from_=0, to=59, wrap=True, width=3, format="%02.0f", bg=SURFACE_COLOR, fg=TEXT_COLOR)
      m_spin.pack(side=tk.LEFT, padx=2)
      
      def set_now():
        now = datetime.datetime.now()
        de.set_date(now.date())
        h_spin.delete(0, "end")
        h_spin.insert(0, f"{now.hour:02d}")
        m_spin.delete(0, "end")
        m_spin.insert(0, f"{now.minute:02d}")
        
      live_var = tk.BooleanVar(value=True)
      cb_live = ttk.Checkbutton(frame, text="Live", variable=live_var)
      cb_live.pack(side=tk.LEFT, padx=10)
      
      after_id = [None]
      def auto_update():
        if not parent.winfo_exists():
          return
        if live_var.get():
          set_now()
        after_id[0] = parent.after(1000, auto_update)
        
      def stop_live(*args):
        live_var.set(False)
        if after_id[0]:
          parent.after_cancel(after_id[0])
          after_id[0] = None

      parent.bind("<Destroy>", lambda e: stop_live(), add="+")
      de.bind("<<DateEntrySelected>>", stop_live)
      h_spin.bind("<Button-1>", stop_live)
      h_spin.bind("<Key>", stop_live)
      m_spin.bind("<Button-1>", stop_live)
      m_spin.bind("<Key>", stop_live)
      
      auto_update()
      return frame, de, h_spin, m_spin

    self.f_dt_sp, self.de_sp, self.h_sp, self.m_sp = create_dt_picker(lf4, "Work Date / Time")
    lf4.columnconfigure(1, weight=1)
    
    # Section 5: Additional Info
    card_lf5, lf5 = self.create_card(self.form_col_right, "Additional Info")
    card_lf5.pack(fill=tk.BOTH, expand=True, pady=5, padx=5)
    
    ttk.Label(lf5, text="Remarks", font=HMI_FONT_S, width=20).grid(row=0, column=0, sticky="nw", pady=2, padx=5)
    self.txt_remarks = tk.Text(lf5, height=2, bg=SURFACE_COLOR, fg=TEXT_COLOR)
    self.txt_remarks.grid(row=0, column=1, sticky="nsew", pady=2, padx=5)
    
    lf5.columnconfigure(1, weight=1)
    lf5.rowconfigure(0, weight=1)
    
    # ---------------- ACTION BUTTONS ----------------
    self.action_frame = tk.Frame(self.t1_right, bg=SURFACE_COLOR, highlightbackground=BORDER_COLOR, highlightthickness=1)
    self.action_frame.pack(side=tk.TOP, fill=tk.X, padx=20, pady=(10, 10))
    
    btn_container = tk.Frame(self.action_frame, bg=SURFACE_COLOR)
    btn_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
    
    self.lbl_status = tk.Label(btn_container, text="", bg=SURFACE_COLOR, fg=SUCCESS_COLOR, font=HMI_FONT_M)
    self.lbl_status.pack(side=tk.LEFT, padx=10)
    
    ttk.Button(btn_container, text="Save Record", style="Success.TButton", command=self.save_record).pack(side=tk.RIGHT, padx=(10, 0), ipadx=20, ipady=5)
    ttk.Button(btn_container, text="Same as Last", style="Primary.TButton", command=self.same_as_last).pack(side=tk.RIGHT, padx=10, ipadx=10, ipady=5)
    ttk.Button(btn_container, text="Clear Form", style="Secondary.TButton", command=self.clear_form).pack(side=tk.RIGHT, padx=(0, 10), ipadx=10, ipady=5)
    
  def build_tab5(self):
    # Consume to Line UI
    main_frame = tk.Frame(self.tab5, bg=BG_COLOR)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
    
    tk.Label(main_frame, text="Consume Box to Production Line", bg=BG_COLOR, fg=WARNING_COLOR, font=HMI_FONT_L).pack(pady=(0, 20))
    
    _, card = self.create_card(main_frame, "Scan Box Label")
    
    ttk.Label(card, text="Scan Sub-Batch ID (SB_ID) ", font=HMI_FONT_M).pack(pady=10)
    self.var_consume_scan = tk.StringVar()
    entry_scan = ttk.Entry(card, textvariable=self.var_consume_scan, width=40, font=HMI_FONT_L)
    entry_scan.pack(pady=10)
    entry_scan.bind("<Return>", self.on_consume_scan)
    
    shift_frame = tk.Frame(card, bg=SURFACE_COLOR)
    shift_frame.pack(pady=5)
    
    self.var_manual_shift = tk.BooleanVar(value=False)
    def toggle_manual():
      if self.var_manual_shift.get():
        self.cb_consume_shift.config(state="readonly")
      else:
        self.cb_consume_shift.config(state="disabled")
        self.cb_consume_shift.set(getattr(self, 'app_user_shift', ''))
        
    cb_manual = ttk.Checkbutton(shift_frame, text="Manual Shift", variable=self.var_manual_shift, command=toggle_manual)
    cb_manual.pack(side=tk.LEFT, padx=5)
    
    self.cb_consume_shift = ttk.Combobox(shift_frame, values=["A", "B", "C"], state="disabled", width=5, font=HMI_FONT_M)
    self.cb_consume_shift.pack(side=tk.LEFT, padx=5)
    
    self.lbl_consume_status = tk.Label(card, text="", bg=SURFACE_COLOR, font=HMI_FONT_M)
    self.lbl_consume_status.pack(pady=20)

  def on_consume_scan(self, event=None):
    scan_data = self.var_consume_scan.get().strip()
    if hasattr(self, 'var_manual_shift') and self.var_manual_shift.get():
      shift = self.cb_consume_shift.get()
    else:
      shift = getattr(self, 'app_user_shift', '')
      
    if not scan_data: return
    if not shift:
      self.lbl_consume_status.config(text="Please select shift first.", fg=ERROR_COLOR)
      return
      
    sb_id = scan_data
    if scan_data.startswith("{") and scan_data.endswith("}"):
      try:
        import json
        data = json.loads(scan_data)
        sb_id = data.get("sub_batch_id", "")
        if not sb_id:
          dt_sp = data.get("sub_process_datetime", "")
          station = data.get("station", "")
          shift_sp = data.get("sub_process_shift", "")
          if dt_sp and station and shift_sp:
            sb_id = f"SB{dt_sp.replace('-', '').replace(':', '').replace(' ', '')}{station}{shift_sp}"
      except Exception:
        pass
        
    if not sb_id:
      self.lbl_consume_status.config(text="Invalid barcode format.", fg=ERROR_COLOR)
      return

    dt_line = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    
    conn = get_db_connection()
    c = conn.cursor()
    
    c.execute("SELECT id, status, dt_sp, pn_sf, quantity FROM records WHERE sub_batch_id = ?", (sb_id,))
    row = c.fetchone()
    
    if not row:
      self.lbl_consume_status.config(text="Error: Sub-Batch ID not found in database!", fg=ERROR_COLOR)
      conn.close()
      return
      
    if row[1] == 'Consumed':
      self.lbl_consume_status.config(text=f"Warning: Box {sb_id} is already Consumed!", fg=WARNING_COLOR)
      conn.close()
      self.var_consume_scan.set("")
      return
      
    orig_qty = row[4]
    
    # Check Quality Quarantine Status
    try:
      c.execute("SELECT IFNULL(SUM(CASE WHEN status='Closed' AND action_type IN ('Rework', 'Sorting', 'Use As-Is') THEN 0 ELSE qty_defective END), 0) FROM quality_defects WHERE sub_batch_id=? AND is_quarantined=1", (sb_id,))
      q_row = c.fetchone()
      defects = q_row[0] if q_row else 0
      
      if defects > 0:
        self.lbl_consume_status.config(text="")
        messagebox.showerror("Quarantined Box", f" BLOCKED!\n\nBox {sb_id} has {defects} quarantined part(s).\nIt CANNOT be consumed.", parent=self)
        conn.close()
        self.var_consume_scan.set("")
        return
    except sqlite3.OperationalError:
      pass
      
    dt_sp = row[2]
    pn_sf = row[3]
    
    # FIFO Check
    c.execute("SELECT COUNT(*) FROM records WHERE pn_sf=? AND status='In Rack' AND dt_sp < ?", (pn_sf, dt_sp))
    older_count = c.fetchone()[0]
    
    if older_count > 0:
      msg = f"FIFO WARNING: There are {older_count} OLDER box(es) of {pn_sf} in the rack.\n\nAre you sure you want to consume this newer box?"
      if not messagebox.askyesno("FIFO Alert", msg, parent=self):
        conn.close()
        self.var_consume_scan.set("")
        self.lbl_consume_status.config(text="Consume cancelled due to FIFO violation.", fg=WARNING_COLOR)
        return
        
    c.execute("UPDATE records SET status = 'Consumed', dt_line = ?, shift_line = ? WHERE sub_batch_id = ?", (dt_line, shift, sb_id))
    conn.commit()
    conn.close()
    
    self.update_excel_record(sb_id, dt_line, shift)
    
    self.lbl_consume_status.config(text=f"Success! {sb_id} consumed at {dt_line}", fg=SUCCESS_COLOR)
    self.var_consume_scan.set("")
    self.refresh_records_treeview()
    self.after(5000, lambda: self.lbl_consume_status.config(text=""))

  def update_excel_record(self, sb_id, dt_line, shift_line):
    if not os.path.exists(EXCEL_FILE): return
    try:
      from openpyxl import load_workbook
      wb = load_workbook(EXCEL_FILE)
      if "Sub-process fill by TL" not in wb.sheetnames: return
      ws = wb["Sub-process fill by TL"]
      
      for row in range(2, ws.max_row + 1):
        cell_sb_id = ws.cell(row=row, column=1).value
        if cell_sb_id == sb_id:
          ws.cell(row=row, column=14).value = dt_line
          ws.cell(row=row, column=15).value = shift_line
      wb.save(EXCEL_FILE)
    except Exception as e:
      print("Error updating excel:", e)
      
  def build_tab4(self):
    # Container Split
    self.tab4.columnconfigure(0, weight=1)
    self.tab4.columnconfigure(1, weight=2)
    self.tab4.rowconfigure(0, weight=1)
    
    # --- Left Pane: Profile & Actions ---
    left_pane = tk.Frame(self.tab4, bg=BG_COLOR)
    left_pane.grid(row=0, column=0, sticky="nsew", padx=(20, 10), pady=20)
    
    tk.Label(left_pane, text="Deep Traceability", bg=BG_COLOR, fg=ACCENT_COLOR, font=HMI_FONT_L).pack(pady=(0, 20), anchor="w")
    
    _, search_card = self.create_card(left_pane, "Search Box")
    ttk.Label(search_card, text="Scan Bar Code or Type SB_ID ", font=HMI_FONT_S).pack(pady=5, anchor="w")
    self.var_trace_scan = tk.StringVar()
    entry_scan = ttk.Entry(search_card, textvariable=self.var_trace_scan, width=35, font=HMI_FONT_M)
    entry_scan.pack(pady=5, fill=tk.X)
    entry_scan.bind("<Return>", self.on_trace_scan)
    
    self.lbl_trace_status = tk.Label(search_card, text="", bg=SURFACE_COLOR, font=HMI_FONT_S)
    self.lbl_trace_status.pack(pady=5)
    
    _, profile_card = self.create_card(left_pane, "Box Profile")
    self.txt_trace_profile = tk.Text(profile_card, height=12, width=40, bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M, state="disabled", relief="flat", padx=15, pady=15)
    self.txt_trace_profile.pack(pady=10, fill=tk.BOTH, expand=True)
    
    self.btn_do_reprint = ttk.Button(profile_card, text=" Reprint Label", style="Warning.TButton", command=self.do_reprint_action)
    self.btn_do_reprint.pack(pady=10, fill=tk.X)
    self.btn_do_reprint.config(state="disabled")

    _, stats_card = self.create_card(left_pane, "Today's Part Stats")
    self.lbl_stat_produced = tk.Label(stats_card, text="Produced Today: --", bg=SURFACE_COLOR, fg=TEXT_MUTED, font=HMI_FONT_M, anchor="w")
    self.lbl_stat_produced.pack(fill=tk.X, pady=(10, 2), padx=15)
    self.lbl_stat_defects = tk.Label(stats_card, text="Defects Today: --", bg=SURFACE_COLOR, fg=TEXT_MUTED, font=HMI_FONT_M, anchor="w")
    self.lbl_stat_defects.pack(fill=tk.X, pady=2, padx=15)
    self.lbl_stat_rate = tk.Label(stats_card, text="Defect Rate: --", bg=SURFACE_COLOR, fg=WARNING_COLOR, font=HMI_FONT_M, anchor="w")
    self.lbl_stat_rate.pack(fill=tk.X, pady=(2, 10), padx=15)

    # --- Right Pane: Lifecycle Timeline ---
    right_pane = tk.Frame(self.tab4, bg=BG_COLOR)
    right_pane.grid(row=0, column=1, sticky="nsew", padx=(10, 20), pady=20)
    
    tk.Label(right_pane, text="Lifecycle Timeline", bg=BG_COLOR, fg=ACCENT_COLOR, font=HMI_FONT_L).pack(pady=(0, 20), anchor="w")
    
    timeline_card = tk.Frame(right_pane, bg=SURFACE_COLOR)
    timeline_card.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
    
    # Scrollable Text for Timeline
    scroll = ttk.Scrollbar(timeline_card)
    scroll.pack(side=tk.RIGHT, fill=tk.Y)
    
    self.txt_timeline = tk.Text(timeline_card, yscrollcommand=scroll.set, bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M, state="disabled", relief="flat", wrap="word", padx=20, pady=20)
    self.txt_timeline.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
    scroll.config(command=self.txt_timeline.yview)
    
    # Profile Tags
    self.txt_trace_profile.tag_config("label", font=("Segoe UI", 10, "bold"), foreground=TEXT_MUTED)
    self.txt_trace_profile.tag_config("value", font=("Segoe UI", 12, "bold"), foreground=TEXT_COLOR, spacing3=10)
    self.txt_trace_profile.tag_config("rm_title", font=("Segoe UI", 11, "bold"), foreground=ACCENT_COLOR, spacing1=10, spacing3=5)
    self.txt_trace_profile.tag_config("rm_val", font=("Segoe UI", 10), foreground=TEXT_COLOR)

    # Timeline Tags
    self.txt_timeline.tag_config("title", font=("Segoe UI", 16, "bold"), foreground=ACCENT_COLOR, spacing3=15)
    self.txt_timeline.tag_config("time", font=("Segoe UI", 10, "bold"), foreground=TEXT_MUTED)
    self.txt_timeline.tag_config("event_title", font=("Segoe UI", 13, "bold"))
    self.txt_timeline.tag_config("event_details", font=("Segoe UI", 11), foreground=TEXT_MUTED, lmargin1=25, lmargin2=25, spacing3=15)
    self.txt_timeline.tag_config("success", foreground=SUCCESS_COLOR)
    self.txt_timeline.tag_config("error", foreground=ERROR_COLOR)
    self.txt_timeline.tag_config("warning", foreground=WARNING_COLOR)
    self.txt_timeline.tag_config("info", foreground="#60A5FA") # Light Blue
    self.txt_timeline.tag_config("center", justify='center')
    
    # Initialize empty state placeholders
    self.txt_trace_profile.config(state="normal")
    self.txt_trace_profile.insert(tk.END, "\nNo Box Selected", "time")
    self.txt_trace_profile.tag_add("center", "1.0", "end")
    self.txt_trace_profile.config(state="disabled")
    
    self.txt_timeline.config(state="normal")
    self.txt_timeline.insert(tk.END, "\n\n\n\n\nWaiting for Box ID scan...\n\nScan a Barcode or type the SB_ID in the Search Box\nto load the lifecycle timeline.", "time")
    self.txt_timeline.tag_add("center", "1.0", "end")
    self.txt_timeline.config(state="disabled")

  def on_trace_scan(self, event=None):
    raw_input = self.var_trace_scan.get().strip()
    if not raw_input: return
    
    sb_id = raw_input
    if raw_input.startswith('{') and raw_input.endswith('}'):
      try:
        import json
        data = json.loads(raw_input)
        sb_id = data.get("sub_batch_id", "")
      except json.JSONDecodeError:
        pass
        
    if not sb_id:
      self._set_trace_error("Could not determine Sub-Batch ID!")
      return
      
    self.var_trace_scan.set(sb_id)
    
    # Fetch Data
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    
    # 1. Fetch Production Record
    c.execute("SELECT * FROM records WHERE sub_batch_id = ?", (sb_id,))
    record = c.fetchone()
    
    if not record:
      conn.close()
      self._set_trace_error(f"Record '{sb_id}' not found in Database!")
      return
      
    # 2. Fetch Quality Defects
    # Make sure quality_defects table exists
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='quality_defects'")
    if c.fetchone():
      c.execute("SELECT * FROM quality_defects WHERE sub_batch_id = ? ORDER BY reported_at ASC", (sb_id,))
      defects = c.fetchall()
    else:
      defects = []
    
    self.lbl_trace_status.config(text="Traceability generated successfully.", fg=SUCCESS_COLOR)
    
    # --- Update Quick Stats ---
    import datetime
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    
    c.execute("SELECT SUM(quantity) FROM records WHERE pn_sf = ? AND dt_sp LIKE ?", (record['pn_sf'], f"{today}%"))
    total_produced = c.fetchone()[0] or 0
    
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='quality_defects'")
    total_defects = 0
    if c.fetchone():
      c.execute("SELECT SUM(CASE WHEN status='Closed' AND action_type IN ('Rework', 'Sorting', 'Use As-Is') THEN 0 ELSE qty_defective END) FROM quality_defects WHERE pn_sf = ? AND reported_at LIKE ?", (record['pn_sf'], f"{today}%"))
      total_defects = c.fetchone()[0] or 0
      
    conn.close()
    
    rate = (total_defects / total_produced * 100) if total_produced > 0 else 0.0
    
    self.lbl_stat_produced.config(text=f"Produced Today: {total_produced} pcs", fg=TEXT_COLOR)
    self.lbl_stat_defects.config(text=f"Defects Today: {total_defects} pcs", fg=TEXT_COLOR)
    self.lbl_stat_rate.config(text=f"Defect Rate: {rate:.1f}%")
    
    if rate > 5.0:
      self.lbl_stat_rate.config(fg=ERROR_COLOR)
    elif rate > 0.0:
      self.lbl_stat_rate.config(fg=WARNING_COLOR)
    else:
      self.lbl_stat_rate.config(fg=SUCCESS_COLOR)
    
    # Update Profile
    self.txt_trace_profile.config(state="normal")
    self.txt_trace_profile.delete("1.0", tk.END)
    
    self.txt_trace_profile.insert(tk.END, "Sub-Batch ID:\n", "label")
    self.txt_trace_profile.insert(tk.END, f"{record['sub_batch_id']}\n", "value")
    
    self.txt_trace_profile.insert(tk.END, "SF Part Number:\n", "label")
    self.txt_trace_profile.insert(tk.END, f"{record['pn_sf']} ({record['part_sf']})\n", "value")
    
    box_defects = sum(d['qty_defective'] for d in defects if not (d['status'] == 'Closed' and d['action_type'] in ('Rework', 'Sorting', 'Use As-Is'))) if defects else 0
    orig_qty = record['quantity']
    good_qty = orig_qty - box_defects
    
    self.txt_trace_profile.insert(tk.END, "Quantity:\n", "label")
    if box_defects > 0:
        self.txt_trace_profile.insert(tk.END, f"{orig_qty} pcs ", "value")
        self.txt_trace_profile.insert(tk.END, f"({good_qty} OK, {box_defects} Defective)\n", "rm_val")
    else:
        self.txt_trace_profile.insert(tk.END, f"{orig_qty} pcs\n", "value")
    
    self.txt_trace_profile.insert(tk.END, "Current Status:\n", "label")
    self.txt_trace_profile.insert(tk.END, f"{record['status']}\n", "value")
    
    self.txt_trace_profile.insert(tk.END, "--- RM Traceability ---\n", "rm_title")
    
    self.txt_trace_profile.insert(tk.END, f"RM1: {record['rm1_pn']}\n", "rm_val")
    rm1_batches = [b for b in (record['batch1'], record['batch2'], record['batch3']) if b]
    if rm1_batches:
      batches_str = " | ".join(rm1_batches)
      self.txt_trace_profile.insert(tk.END, f" ↳ Batches: {batches_str}\n", "rm_val")
      
    if record['rm2_pn']: self.txt_trace_profile.insert(tk.END, f"RM2: {record['rm2_pn']}\n", "rm_val")
    if record['rm3_pn']: self.txt_trace_profile.insert(tk.END, f"RM3: {record['rm3_pn']}\n", "rm_val")
    if record['rm4_pn']: self.txt_trace_profile.insert(tk.END, f"RM4: {record['rm4_pn']}\n", "rm_val")
    
    self.txt_trace_profile.config(state="disabled")
    
    self.btn_do_reprint.config(state="normal")
    self.current_trace_sb_id = sb_id
    
    # Generate Timeline
    self.txt_timeline.config(state="normal")
    self.txt_timeline.delete("1.0", tk.END)
    
    self.txt_timeline.insert(tk.END, f"Lifecycle Journey for {sb_id}\n", "title")
    
    try:
      events = []
      
      # Event 1: Creation
      events.append({
        "time": record['dt_sp'],
        "type": "creation",
        "title": f" Created in Sub-Process",
        "details": f"Station: {record['station']}  |  Operator: {record['op_id']}  |  Shift: {record['shift_sp']}",
        "tag": "success"
      })
      
      # Event 2: Quality Defects
      for df in defects:
        status_emoji = "" if df['status'] == "Open" else ""
        q_title = f"{status_emoji} Quality Alert: {df['defect_type']}"
        q_details = f"Qty Defective: {df['qty_defective']}  |  Status: {df['status']}  |  Inspector: {df['quality_op_id']}"
        if df['is_quarantined']:
          q_details += "\n Box was Quarantined."
        if df['action_type']:
          q_details += f"\nAction: {df['action_type']}"
          
        events.append({
          "time": df['reported_at'],
          "type": "quality",
          "title": q_title,
          "details": q_details,
          "tag": "error" if df['status'] == "Open" else "warning"
        })
        
      # Event 3: Consumption
      if record['dt_line']:
        events.append({
          "time": record['dt_line'],
          "type": "consume",
          "title": f" Consumed to Final Production Line",
          "details": f"Shift: {record['shift_line']}  |  Registered by: {record['registered_by']}",
          "tag": "info"
        })
        
      # Sort events by time
      events.sort(key=lambda x: x['time'] if x['time'] else "")
      
      for ev in events:
        if not ev['time']: continue
        self.txt_timeline.insert(tk.END, f"[{ev['time']}] ", "time")
        self.txt_timeline.insert(tk.END, f"{ev['title']}\n", ("event_title", ev['tag']))
        self.txt_timeline.insert(tk.END, f"↳ {ev['details']}\n", "event_details")
        
    except Exception as e:
      self.txt_timeline.insert(tk.END, f"\n\n Error generating timeline: {str(e)}", "error")
      import traceback
      print(traceback.format_exc())
      
    self.txt_timeline.config(state="disabled")

  def _set_trace_error(self, msg):
    self.lbl_trace_status.config(text=msg, fg=ERROR_COLOR)
    
    self.txt_trace_profile.config(state="normal")
    self.txt_trace_profile.delete("1.0", tk.END)
    self.txt_trace_profile.insert(tk.END, "\nNo Box Selected", "time")
    self.txt_trace_profile.config(state="disabled")
    
    self.lbl_stat_produced.config(text="Produced Today: --", fg=TEXT_MUTED)
    self.lbl_stat_defects.config(text="Defects Today: --", fg=TEXT_MUTED)
    self.lbl_stat_rate.config(text="Defect Rate: --", fg=WARNING_COLOR)
    
    self.txt_timeline.config(state="normal")
    self.txt_timeline.delete("1.0", tk.END)
    self.txt_timeline.insert(tk.END, "\n\n\n\n\nWaiting for Box ID scan...\n\nScan a Barcode or type the SB_ID in the Search Box\nto load the lifecycle timeline.", "time")
    self.txt_timeline.tag_add("center", "1.0", "end")
    self.txt_timeline.config(state="disabled")
    
    self.btn_do_reprint.config(state="disabled")
    self.var_trace_scan.set("")

  def do_reprint_action(self):
    if hasattr(self, 'current_trace_sb_id'):
      self.do_print(self.current_trace_sb_id)
      self.lbl_trace_status.config(text="Label sent to printer.", fg=SUCCESS_COLOR)
      self.after(3000, lambda: getattr(self, 'lbl_trace_status', tk.Label()).config(text=""))
      

      
  def on_rm_scanned_t1(self, event=None):
    scanned_rm = self.var_scan_rm_t1.get().strip()
    if hasattr(self, 'lbl_scan_rm_status'):
      self.lbl_scan_rm_status.config(text="")
    if not scanned_rm: return
    
    matched_sf_pn = None
    for sf_pn, (sf_name, rm_list) in SF_DATA.items():
      if rm_list and rm_list[0][0] == scanned_rm:
        matched_sf_pn = sf_pn
        break
        
    if matched_sf_pn:
      self.cb_sf_pn.config(state="readonly")
      self.cb_sf_pn.set(matched_sf_pn)
      if hasattr(self, 'lbl_scan_rm_status'):
        self.lbl_scan_rm_status.config(text="Match Found!", fg=SUCCESS_COLOR)
        self.after(3000, lambda: getattr(self, 'lbl_scan_rm_status', tk.Label()).config(text=""))
      self.on_sf_selected(None)
      self.cb_sf_pn.config(state="disabled")
    else:
      if hasattr(self, 'lbl_scan_rm_status'):
        self.lbl_scan_rm_status.config(text="Mismatch! Invalid RM Barcode.", fg=ERROR_COLOR)
        self.after(4000, lambda: getattr(self, 'lbl_scan_rm_status', tk.Label()).config(text=""))
      self.var_scan_rm_t1.set("")
    


  def build_tab2(self):
    filter_frame = tk.Frame(self.tab2, bg=SURFACE_COLOR)
    filter_frame.pack(fill=tk.X, padx=10, pady=10)
    
    ttk.Label(filter_frame, text="Search:").pack(side=tk.LEFT, padx=2)
    self.var_rec_search = tk.StringVar()
    ttk.Entry(filter_frame, textvariable=self.var_rec_search, width=20).pack(side=tk.LEFT, padx=2)
    
    ttk.Label(filter_frame, text="Shift:").pack(side=tk.LEFT, padx=(10,2))
    self.cb_rec_shift = ttk.Combobox(filter_frame, values=["All", "A", "B", "C"], state="readonly", width=5)
    self.cb_rec_shift.set("All")
    self.cb_rec_shift.pack(side=tk.LEFT, padx=2)
    
    ttk.Label(filter_frame, text="Station:").pack(side=tk.LEFT, padx=(10,2))
    self.cb_rec_station = ttk.Combobox(filter_frame, values=["All", "S06", "S07", "S10", "S11"], state="readonly", width=5)
    self.cb_rec_station.set("All")
    self.cb_rec_station.pack(side=tk.LEFT, padx=2)
    
    ttk.Button(filter_frame, text="Search", style="Primary.TButton", command=self.refresh_records_treeview).pack(side=tk.LEFT, padx=10)
    ttk.Button(filter_frame, text="All", style="Secondary.TButton", command=self.reset_records_filter).pack(side=tk.LEFT, padx=2)
    ttk.Button(filter_frame, text="Print Selected", style="Warning.TButton", command=self.print_selected_record).pack(side=tk.RIGHT, padx=5)
    
    self.lbl_rec_count = tk.Label(filter_frame, text="0 records", bg=SURFACE_COLOR, fg=TEXT_COLOR)
    self.lbl_rec_count.pack(side=tk.RIGHT, padx=10)
    
    tree_frame = tk.Frame(self.tab2, bg=BG_COLOR)
    tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
    
    cols = ("SB_ID", "SF_PN", "SF_Name", "Qty", "Shift", "Station", "Op_ID", "DateTime", "Status", "Reprints")
    self.tree_records = ttk.Treeview(tree_frame, columns=cols, show="headings")
    for c in cols:
      self.tree_records.heading(c, text=c, command=lambda _c=c: self.sort_treeview(self.tree_records, _c, False))
      self.tree_records.column(c, width=90)
      
    hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_records.xview)
    self.tree_records.configure(xscrollcommand=hsb.set)
    
    self.tree_records.grid(row=0, column=0, sticky="nsew")
    hsb.grid(row=1, column=0, sticky="ew")
    tree_frame.rowconfigure(0, weight=1)
    tree_frame.columnconfigure(0, weight=1)
    
    self.tree_records.tag_configure("pending", background=WARNING_COLOR, foreground="black")
    self.tree_records.tag_configure("partial_defect", background="#A16207", foreground="white") # Dark Yellow/Brown for partial defect
    self.tree_records.tag_configure("even", background="#141A20")
    self.tree_records.tag_configure("odd", background="#1B232C")
    self.tree_records.tag_configure("reprint_highlight", background="#B45309", foreground="white") # Orange background for reprinted labels
    self.tree_records.bind("<Double-1>", lambda e: self.print_selected_record())
    
    self.rec_context_menu = tk.Menu(self.tree_records, tearoff=0, bg=SURFACE_COLOR, fg=TEXT_COLOR, activebackground=ACCENT_COLOR, activeforeground=TEXT_COLOR)
    
    def copy_to_clipboard(text):
      self.clipboard_clear()
      self.clipboard_append(text)
      self.update()

    def on_rec_right_click(event):
      item = self.tree_records.identify_row(event.y)
      if item:
        self.tree_records.selection_set(item)
        values = self.tree_records.item(item, "values")
        if values:
          self.rec_context_menu.delete(0, tk.END)
          
          user_role = getattr(self, 'app_user_role', 'Operator')
          if user_role in ["Supervisor", "Manager", "Admin"]:
            self.rec_context_menu.add_command(label=" View Operator Scorecard", command=self.open_operator_scorecard)
            self.rec_context_menu.add_separator()
            
          self.rec_context_menu.add_command(label=f" Copy SB_ID", command=lambda v=values[0]: copy_to_clipboard(v))
          self.rec_context_menu.add_command(label=f" Copy Op_ID", command=lambda v=values[6]: copy_to_clipboard(v))
          self.rec_context_menu.add_command(label=f" Copy PN", command=lambda v=values[1]: copy_to_clipboard(v))
          self.rec_context_menu.add_separator()
          self.rec_context_menu.add_command(label=f" Copy Row Data", command=lambda v=" | ".join(map(str, values)): copy_to_clipboard(v))
          
          self.rec_context_menu.tk_popup(event.x_root, event.y_root)

    self.tree_records.bind("<Button-3>", on_rec_right_click)
    self.page_frame = tk.Frame(self.tab2, bg=BG_COLOR)
    self.page_frame.pack(fill=tk.X, pady=5)

  def get_dashboard_data(self):
    conn = get_db_connection()
    c = conn.cursor()
    
    now = datetime.datetime.now()
    today_date = now.date()
    
    start_today_str = f"{today_date.strftime('%Y-%m-%d')} 00:00:00"
    
    start_week_date = today_date - datetime.timedelta(days=today_date.weekday())
    start_week_str = f"{start_week_date.strftime('%Y-%m-%d')} 00:00:00"
    
    start_month_date = today_date.replace(day=1)
    start_month_str = f"{start_month_date.strftime('%Y-%m-%d')} 00:00:00"
    
    def get_stats(start_dt):
      c.execute("SELECT shift_sp, SUM(quantity) FROM records WHERE dt_sp >= ? GROUP BY shift_sp", (start_dt,))
      rows = c.fetchall()
      stats = {'A': 0, 'B': 0, 'C': 0}
      for r in rows:
        if r[0] in stats:
          stats[r[0]] = r[1] if r[1] else 0
      total = sum(stats.values())
      return total, stats['A'], stats['B'], stats['C']
      
    today_stats = get_stats(start_today_str)
    week_stats = get_stats(start_week_str)
    month_stats = get_stats(start_month_str)
    
    conn.close()
    return today_stats, week_stats, month_stats


  def update_clock(self):
    try:
      now = datetime.datetime.now()
      self.lbl_clock.config(text=now.strftime("%H:%M:%S"))
      self.lbl_date.config(text=now.strftime("%Y-%m-%d"))
      self._check_shift_end_reports(now)
    except Exception:
      pass
    self.after(1000, self.update_clock)

  def _check_shift_end_reports(self, now):
    try:
      wd = now.weekday() # 0=Mon, 4=Fri, 5=Sat, 6=Sun
      h = now.hour
      m = now.minute
      
      shift_ended = False
      
      if wd in [0, 1, 2, 3]:
        if (h == 14 and m == 0) or (h == 22 and m == 0) or (h == 6 and m == 0):
          shift_ended = True
      elif wd == 4: # Friday
        if (h == 12 and m == 0) or (h == 18 and m == 30) or (h == 6 and m == 0):
          shift_ended = True
      elif wd == 5: # Saturday
        if h == 6 and m == 0: # end of Friday night shift
          shift_ended = True
          
      if shift_ended:
        if hasattr(self, 'app_user_id') and self.app_user_id:
          print(f"Shift ended! Forcing logout for {self.app_user_id}...")
          self.perform_logout(force=True)

    except Exception as e:
      print(f"Shift End Check Error: {e}")

  def check_and_generate_missed_reports(self):
    try:
      now = datetime.datetime.now()
      # PDF reports are strictly based on the calendar day (00:00 to 23:59).
      # Since the day ends at midnight, we can generate yesterday's report anytime today.
      report_date = now - datetime.timedelta(days=1)
      report_date_str = report_date.strftime("%Y-%m-%d")
      
      # Determine expected file path
      reports_dir = os.path.join(DATA_DIR, "reports", report_date_str)
      os.makedirs(reports_dir, exist_ok=True)
      expected_pdf = os.path.join(reports_dir, f"Report_{report_date_str}_Daily.pdf")
      
      if not os.path.exists(expected_pdf):
        # Prevent concurrent triggers
        if getattr(self, '_is_generating_report', False):
            return
            
        self._is_generating_report = True
        
        start_dt = report_date.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = report_date.replace(hour=23, minute=59, second=59, microsecond=999999)
        
        print(f"Detected missed report for {report_date_str}. Triggering generation...")
        
        import threading
        def _run_gen():
          try:
            import report_generator
            report_generator.generate_daily_pdf_report(start_dt, end_dt)
            print(f"Successfully generated missed report for {report_date_str}")
          except Exception as e:
            print(f"Failed to generate report: {e}")
          finally:
            self._is_generating_report = False

        threading.Thread(target=_run_gen, daemon=True).start()
        
    except Exception as e:
      print(f"Error checking missed reports: {e}")
      self._is_generating_report = False
      
    # Check again in 5 minutes
    self.after(300000, self.check_and_generate_missed_reports)


  def update_stats(self):
    try:
      conn = get_db_connection()
      c = conn.cursor()
      c.execute("SELECT product_pn, AVG(target_qty) FROM shift_targets GROUP BY product_pn")
      PROD_RATES = {row[0]: int(row[1] or 120) for row in c.fetchall()}
      conn.close()
    except Exception:
      PROD_RATES = {}
    DEFAULT_RATE = 120 # default hourly rate if PN is not in PROD_RATES

    now = datetime.datetime.now()
    # Production day is purely calendar day
    start_dt = now.strftime("%Y-%m-%d 00:00:00")
    end_dt = now.strftime("%Y-%m-%d 23:59:59")
      
    shift_val = getattr(self, 'app_user_shift', "")
    if not shift_val:
      shift_val = "-"
      
    current_pn = ""
    if hasattr(self, 'cb_sf_pn'):
      current_pn = self.cb_sf_pn.get()
    
    conn = get_db_connection()
    c = conn.cursor()
    
    # Shift stats (Overall for shift)
    c.execute("SELECT COUNT(*), SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ? AND shift_sp = ?", (start_dt, end_dt, shift_val))
      
    res_shift = c.fetchone()
    shift_count = res_shift[0] if res_shift[0] else 0
    shift_qty = res_shift[1] if res_shift[1] else 0
    
    # Today stats (Overall)
    c.execute("SELECT SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ?", (start_dt, end_dt))
    today_qty = c.fetchone()[0] or 0
    
    # Get all PNs produced in this shift
    c.execute("SELECT pn_sf, SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ? AND shift_sp = ? GROUP BY pn_sf", (start_dt, end_dt, shift_val))
    shift_pns_data = {row[0]: (row[1] or 0) for row in c.fetchall() if row[0]}
    
    conn.close()
    
    # Calculate Work Mins
    wd = now.weekday()
    if wd == 6: # Sunday
      work_mins = 0
    elif wd == 4: # Friday
      work_mins = 290
    else:
      work_mins = 440
      
    # Update Shift Target Tracker on Tab 1
    if hasattr(self, 'tracker_scroll_frame'):
      for widget in self.tracker_scroll_frame.winfo_children():
        widget.destroy()
        
      conn = get_db_connection()
      c = conn.cursor()
      c.execute("SELECT product_pn, target_qty FROM shift_targets ORDER BY id ASC")
      db_targets = {}
      for row in c.fetchall():
        db_targets[row[0]] = row[1]
      conn.close()
      all_pns_to_track = set([p for p, q in shift_pns_data.items() if q > 0])
        
      sorted_pns = sorted(list(all_pns_to_track), key=lambda x: (x not in db_targets, x))
      
      for pn in sorted_pns:
        qty = shift_pns_data.get(pn, 0)
        tgt = db_targets.get(pn, 0)
        pct = min(100, (qty / tgt) * 100) if tgt > 0 else 0
        
        f = tk.Frame(self.tracker_scroll_frame, bg=SURFACE_COLOR)
        f.pack(fill=tk.X, pady=(0, 10))
        
        part_name = SF_DATA.get(pn, ("", []))[0]
        pn_display = f"{pn} ({part_name})" if part_name else pn
        lbl = tk.Label(f, text=f"{pn_display}\nProduced: {qty:,} / Target: {tgt:,}", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_S, justify=tk.LEFT)
        lbl.pack(anchor="w")
        
        if pct >= 80:
          pb_style = "Safe.Horizontal.TProgressbar"
        elif pct >= 40:
          pb_style = "Warn.Horizontal.TProgressbar"
        else:
          pb_style = "Danger.Horizontal.TProgressbar"
        
        pb = ttk.Progressbar(f, style=pb_style, orient="horizontal", mode="determinate")
        pb.pack(fill=tk.X, pady=(2, 0))
        pb['value'] = pct
        
        # Bind mousewheel to children too
        def _on_mw(e, cvs=self.tracker_canvas): 
          cvs.yview_scroll(int(-1*(e.delta/120)), "units")
          return "break"
          
        lbl.bind("<MouseWheel>", _on_mw)
        pb.bind("<MouseWheel>", _on_mw)
        f.bind("<MouseWheel>", _on_mw)

      def _on_canvas_mw(e):
        self.tracker_canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        return "break"
        
      self.tracker_canvas.bind("<MouseWheel>", _on_canvas_mw)
      self.tracker_scroll_frame.bind("<MouseWheel>", _on_canvas_mw)
      
      self.tracker_scroll_frame.update_idletasks()
      self.tracker_canvas.config(scrollregion=self.tracker_canvas.bbox("all"))
      
      if len(sorted_pns) > 4:
        self.lbl_scroll_indicator.pack(pady=(0, 5), before=self.t1_left.winfo_children()[2])
      else:
        self.lbl_scroll_indicator.pack_forget()
    
    self.lbl_title_recs.config(text=f"Records (Shift {shift_val})")
    self.side_stat_recs.config(text=str(shift_count))
    self.lbl_title_qty.config(text=f"Qty (Shift {shift_val})")
    self.side_stat_qty.config(text=str(shift_qty))
    self.side_stat_today.config(text=str(today_qty))

  def populate_sf_combobox(self):
    sf_list = list(SF_DATA.keys())
    self.cb_sf_pn['values'] = sf_list
    self.lbl_search_count.config(text=f"{len(sf_list)} matches", fg=SUCCESS_COLOR)

  def filter_sf_combobox(self, *args):
    search_term = self.sv_search.get().lower()
    filtered = [pn for pn in SF_DATA.keys() if search_term in pn.lower()]
    self.cb_sf_pn['values'] = filtered
    if filtered:
      self.lbl_search_count.config(text=f"{len(filtered)} matches", fg=SUCCESS_COLOR)
      self.cb_sf_pn.set(filtered[0])
      self.on_sf_selected(None)
    else:
      self.lbl_search_count.config(text="0 matches", fg="red")
      self.cb_sf_pn.set("")

  def on_sf_selected(self, event):
    sf_pn = self.cb_sf_pn.get()
    for w in self.rm_widgets:
      w.destroy()
    self.rm_widgets.clear()
    self.rm_vars_t1 = []

    if sf_pn in SF_DATA:
      sf_name, rm_list = SF_DATA[sf_pn]
      self.var_part_sf.set(sf_name)
      
      start_row = 4
      for idx, (rm_id, rm_name) in enumerate(rm_list):
        row_idx = start_row + (idx * 2)
        lbl_ref = ttk.Label(self.lf1, text=f"RM {idx+1} Ref", font=HMI_FONT_S, width=20)
        lbl_ref.grid(row=row_idx, column=0, sticky="w", padx=5, pady=2)
        cb_var = tk.StringVar(value=rm_id)
        cb = ttk.Entry(self.lf1, textvariable=cb_var, state="readonly")
        cb.grid(row=row_idx, column=1, sticky="ew", padx=5, pady=2)
        
        lbl_name = ttk.Label(self.lf1, text=f"  └ Name", font=HMI_FONT_S, width=20)
        lbl_name.grid(row=row_idx+1, column=0, sticky="w", padx=5, pady=2)
        name_var = tk.StringVar(value=rm_name)
        en_name = ttk.Entry(self.lf1, textvariable=name_var, state="readonly")
        en_name.grid(row=row_idx+1, column=1, sticky="ew", padx=5, pady=2)
        
        self.rm_widgets.extend([lbl_ref, cb, lbl_name, en_name])
        self.rm_vars_t1.append((cb_var, name_var))

  def update_sub_batch_preview(self):
    self.update_stats()
    pass

  def get_dt_string(self, de, h, m):
    d = de.get_date()
    return f"{d.strftime('%Y-%m-%d')} {h.get()}:{m.get()}"

  def clear_form(self):
    if hasattr(self, 'var_scan_rm_t1'):
      self.var_scan_rm_t1.set("")
    self.cb_sf_pn.config(state="readonly")
    self.cb_sf_pn.set("")
    self.var_part_sf.set("")
    for w in self.rm_widgets:
      w.destroy()
    self.rm_widgets.clear()
    self.rm_vars_t1 = []
    self.var_b1.set("")
    self.var_b2.set("")
    self.var_b3.set("")
    self.var_qty.set("")
    self.cb_shift_sp.set("")
    self.var_op_id.set("")
    self.cb_station.set("")
    
    now = datetime.datetime.now()
    self.de_sp.set_date(now.date())
    self.h_sp.delete(0, "end"); self.h_sp.insert(0, f"{now.hour:02d}")
    self.m_sp.delete(0, "end"); self.m_sp.insert(0, f"{now.minute:02d}")
    
    self.txt_remarks.delete("1.0", tk.END)
    if hasattr(self, 'cb_consume_shift') and hasattr(self, 'var_manual_shift'):
      if not self.var_manual_shift.get():
        self.cb_consume_shift.set(getattr(self, 'app_user_shift', ''))
    self.update_sub_batch_preview()

  def same_as_last(self):
    if hasattr(self, 'var_scan_rm_t1'):
      self.var_scan_rm_t1.set("")
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT * FROM records ORDER BY id DESC LIMIT 1")
    rec = c.fetchone()
    conn.close()
    if rec:
      self.cb_sf_pn.set(rec[2])
      self.on_sf_selected(None)
      self.var_qty.set(rec[15])
      self.cb_shift_sp.set(rec[16])
      if hasattr(self, 'var_op_id'):
        self.var_op_id.set(rec[17])
      self.cb_station.set(rec[18])
      self.update_sub_batch_preview()
      self.lbl_status.config(text="Pre-filled with last record data.", fg=SUCCESS_COLOR)
      self.after(3000, lambda: self.lbl_status.config(text=""))

  def background_excel_save(self, excel_data):
    try:
      _excel_queue.put((save_to_excel, (excel_data,), {}))
      _excel_queue.put((self.export_kpis_to_excel, (), {}))
    except Exception as e:
      print(f"Background Excel Save Error: {e}")

  def save_record(self):
    sf_pn = self.cb_sf_pn.get()
    qty_str = self.var_qty.get()
    op_id = self.var_op_id.get()
    station = self.cb_station.get()
    shift_sp = self.cb_shift_sp.get()
    
    if not all([sf_pn, qty_str, op_id, station, shift_sp]) or not self.rm_vars_t1:
      messagebox.showerror("Error", "Please fill all required fields ().")
      return
      
    if not any([self.var_b1.get().strip(), self.var_b2.get().strip(), self.var_b3.get().strip()]):
      messagebox.showerror("Error", "Please provide at least one Batch Number.")
      return
      
    rm_pns = [""] * 4
    rm_names = [""] * 4
    for idx, (cb_var, name_var) in enumerate(self.rm_vars_t1):
      if not cb_var.get():
        messagebox.showerror("Error", "Please fill all RM Reference fields.")
        return
      if idx < 4:
        rm_pns[idx] = cb_var.get()
        rm_names[idx] = name_var.get()

    try:
      qty = int(qty_str)
    except ValueError:
      messagebox.showerror("Error", "Quantity must be a valid integer.")
      return

    shift_sp = self.cb_shift_sp.get()
    dt_sp = self.get_dt_string(self.de_sp, self.h_sp, self.m_sp)
    dt_line = ""
    shift_line = ""
    
    conn = get_db_connection()
    c = conn.cursor()
    
    import uuid
    base_sb_id = f"SB{dt_sp.replace('-', '').replace(':', '').replace(' ', '')}{station}{shift_sp[0] if shift_sp else ''}"
    uid = uuid.uuid4().hex[:4].upper()
    c.execute("SELECT COUNT(*) FROM records WHERE sub_batch_id LIKE ?", (base_sb_id + "%",))
    count = c.fetchone()[0]
    sb_id = f"{base_sb_id}{count+1:02d}{uid}"

    # Custom Premium Routing Dialog (Safe Mode)
    dialog = tk.Toplevel(self)
    dialog.title("Routing Destination")
    dialog.geometry("500x220")
    dialog.transient(self)
    dialog.grab_set()
    dialog.configure(bg=BG_COLOR)
    
    # Center the dialog
    dialog.update_idletasks()
    x = self.winfo_x() + (self.winfo_width() - 500) // 2
    y = self.winfo_y() + (self.winfo_height() - 220) // 2
    dialog.geometry(f"+{x}+{y}")
    
    result = [None]
    
    tk.Label(dialog, text="Where should this record be routed?", bg=BG_COLOR, fg=TEXT_COLOR, font=("Segoe UI", 16, "bold")).pack(pady=(25, 15))
    
    btn_frame = tk.Frame(dialog, bg=BG_COLOR)
    btn_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=(0, 25))
    
    def set_routing(status):
        result[0] = status
        dialog.destroy()
        
    def create_hover_btn(parent, text, subtext, color, command):
        f = tk.Frame(parent, bg=SURFACE_COLOR, cursor="hand2", highlightbackground=BORDER_COLOR, highlightthickness=1)
        
        lbl_main = tk.Label(f, text=text, bg=SURFACE_COLOR, fg=color, font=("Segoe UI", 14, "bold"), cursor="hand2")
        lbl_main.pack(pady=(20, 5))
        lbl_sub = tk.Label(f, text=subtext, bg=SURFACE_COLOR, fg=TEXT_MUTED, font=HMI_FONT_S, cursor="hand2")
        lbl_sub.pack(pady=(0, 20))
        
        def on_enter(e):
            f.config(bg="#1c242d", highlightbackground=color)
            lbl_main.config(bg="#1c242d")
            lbl_sub.config(bg="#1c242d", fg=TEXT_COLOR)
        def on_leave(e):
            f.config(bg=SURFACE_COLOR, highlightbackground=BORDER_COLOR)
            lbl_main.config(bg=SURFACE_COLOR)
            lbl_sub.config(bg=SURFACE_COLOR, fg=TEXT_MUTED)
            
        for w in (f, lbl_main, lbl_sub):
            w.bind("<Enter>", on_enter)
            w.bind("<Leave>", on_leave)
            w.bind("<Button-1>", lambda e: command())
            
        return f

    btn_rack = create_hover_btn(btn_frame, "Stored in Rack", "Prints Barcode Label", ACCENT_COLOR, lambda: set_routing('In Rack'))
    btn_rack.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=(0, 10))
    
    btn_line = create_hover_btn(btn_frame, "Direct to Line", "Fast / No Label", WARNING_COLOR, lambda: set_routing('Direct to Line'))
    btn_line.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=(10, 0))
    
    self.wait_window(dialog)
    
    record_status = result[0]
    if not record_status:
        return # User closed dialog without selecting
        
    created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    registered_by = f"{getattr(self, 'app_user_id', '')}"
    
    # --- DOWNTIME TRACKING LOGIC ---
    def prompt_downtime(is_smart_check=False, deficit_minutes=0):
        dt_dialog = tk.Toplevel(self)
        dt_dialog.title("Downtime Detected" if is_smart_check else "Manual Downtime Entry")
        dt_dialog.geometry("500x380")
        dt_dialog.transient(self)
        dt_dialog.grab_set()
        dt_dialog.configure(bg=BG_COLOR)
        
        dt_dialog.update_idletasks()
        dx = self.winfo_x() + (self.winfo_width() - 500) // 2
        dy = self.winfo_y() + (self.winfo_height() - 380) // 2
        dt_dialog.geometry(f"+{dx}+{dy}")
        
        dt_result = {"duration": 0, "reason": None}
        
        # Premium Header
        header_bg = WARNING_COLOR if is_smart_check else ACCENT_COLOR
        header_frame = tk.Frame(dt_dialog, bg=header_bg, height=60)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text="Downtime Log" if is_smart_check else "Manual Downtime Log", 
                 bg=header_bg, fg="#ffffff", font=("Segoe UI", 16, "bold")).pack(side=tk.LEFT, padx=20, pady=15)
        
        content_frame = tk.Frame(dt_dialog, bg=BG_COLOR)
        content_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=20)
        
        if is_smart_check:
            msg = f"A production deficit equivalent to {deficit_minutes:.1f} minutes\nwas detected in this hour. What was the reason?"
            dur_val = tk.StringVar(value=str(round(deficit_minutes)))
        else:
            msg = "Did you experience any downtime before this entry?\nPlease enter the duration and reason."
            dur_val = tk.StringVar(value="0")
            
        tk.Label(content_frame, text=msg, bg=BG_COLOR, fg=TEXT_MUTED, font=("Segoe UI", 11), justify=tk.LEFT).pack(anchor="w", pady=(0, 15))
        
        # Duration row
        f_dur = tk.Frame(content_frame, bg=BG_COLOR)
        f_dur.pack(fill=tk.X, pady=5)
        tk.Label(f_dur, text="Duration (Minutes):", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_S, width=18, anchor="w").pack(side=tk.LEFT)
        entry_dur = ttk.Entry(f_dur, textvariable=dur_val, width=15, font=HMI_FONT_M)
        entry_dur.pack(side=tk.LEFT)
        
        # Reason row
        f_reason = tk.Frame(content_frame, bg=BG_COLOR)
        f_reason.pack(fill=tk.X, pady=10)
        tk.Label(f_reason, text="Primary Reason:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_S, width=18, anchor="w").pack(side=tk.LEFT)
        reasons = ["Material Shortage", "Quality Issue", "Changeover", "Data Entry Delay", "Deplacement", "Other"]
        cb_reason = ttk.Combobox(f_reason, values=reasons, state="readonly", width=25, font=HMI_FONT_M)
        if reasons: cb_reason.current(0)
        cb_reason.pack(side=tk.LEFT)
        
        # Other reason row (hidden by default)
        f_other = tk.Frame(content_frame, bg=BG_COLOR)
        tk.Label(f_other, text="Specify Other:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_S, width=18, anchor="w").pack(side=tk.LEFT)
        txt_other = tk.Entry(f_other, font=HMI_FONT_M, bg=SURFACE_COLOR, fg=TEXT_COLOR, insertbackground=TEXT_COLOR)
        txt_other.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        def on_reason_change(e):
            if cb_reason.get() == "Other":
                f_other.pack(fill=tk.X, pady=5)
            else:
                f_other.pack_forget()
                
        cb_reason.bind("<<ComboboxSelected>>", on_reason_change)
        
        # Footer buttons
        btn_frame = tk.Frame(dt_dialog, bg=BG_COLOR)
        btn_frame.pack(fill=tk.X, padx=30, pady=(0, 20))
        
        def on_submit():
            try:
                mins = float(dur_val.get())
                reason_val = cb_reason.get()
                if reason_val == "Other":
                    other_text = txt_other.get().strip()
                    if not other_text:
                        messagebox.showerror("Error", "Please specify the 'Other' reason.", parent=dt_dialog)
                        return
                    reason_val = f"Other: {other_text}"
                    
                if mins > 0 and reason_val:
                    dt_result["duration"] = mins
                    dt_result["reason"] = reason_val
                dt_dialog.destroy()
            except ValueError:
                messagebox.showerror("Error", "Please enter a valid number for minutes.", parent=dt_dialog)
                
        def on_skip():
            dt_dialog.destroy()
            
        ttk.Button(btn_frame, text="Save Downtime", style="Success.TButton", command=on_submit).pack(side=tk.RIGHT, ipadx=15, ipady=5)
        if not is_smart_check:
            ttk.Button(btn_frame, text="No Downtime (Skip)", command=on_skip).pack(side=tk.RIGHT, padx=10, ipadx=5, ipady=5)
            
        self.wait_window(dt_dialog)
        return dt_result["duration"], dt_result["reason"]

    if record_status == 'Direct to Line':
        dur, reason = prompt_downtime(is_smart_check=False)
        if dur > 0:
            c.execute("INSERT INTO downtime_logs (sub_batch_id, station, shift, op_id, duration_min, reason, created_at) VALUES (?,?,?,?,?,?,?)",
                      (sb_id, station, shift_sp, op_id, dur, reason, created_at))
    else:
        # Smart Check (Hour by Hour based on computer clock)
        current_time = datetime.datetime.now()
        hour_start = current_time.replace(minute=0, second=0, microsecond=0)
        
        elapsed_mins = (current_time - hour_start).total_seconds() / 60
        
        # Get target from settings (shift_targets table)
        c.execute("SELECT target_qty FROM shift_targets WHERE product_pn=? ORDER BY id DESC LIMIT 1", (sf_pn,))
        target_row = c.fetchone()
        
        if target_row:
            db_target = target_row[0]
            # If the user enters a full shift target (e.g., 1920), divide by 8 hours to get hourly target.
            # If they enter the hourly target directly (e.g., 240), use it directly.
            hourly_target = (db_target / 8.0) if db_target > 500 else db_target
        else:
            hourly_target = 240 # fallback
        
        # Get actual production THIS HOUR
        c.execute("SELECT SUM(quantity) FROM records WHERE station=? AND shift_sp=? AND created_at >= ?", 
                  (station, shift_sp, hour_start.strftime("%Y-%m-%d %H:%M:%S")))
        actual_qty = c.fetchone()[0] or 0
        actual_qty += qty # include the current box
        
        # Get logged downtime THIS HOUR
        c.execute("SELECT SUM(duration_min) FROM downtime_logs WHERE station=? AND shift=? AND created_at >= ?", 
                  (station, shift_sp, hour_start.strftime("%Y-%m-%d %H:%M:%S")))
        logged_dt = c.fetchone()[0] or 0
        
        effective_elapsed = elapsed_mins - logged_dt
        if effective_elapsed > 0:
            expected_qty = (effective_elapsed / 60) * hourly_target
            
            if actual_qty < expected_qty:
                deficit_qty = expected_qty - actual_qty
                deficit_mins = (deficit_qty / hourly_target) * 60
                
                if deficit_mins > 3.75: # 3 mins 45 seconds grace period
                    dur, reason = prompt_downtime(is_smart_check=True, deficit_minutes=deficit_mins)
                    if dur > 0:
                        c.execute("INSERT INTO downtime_logs (sub_batch_id, station, shift, op_id, duration_min, reason, created_at) VALUES (?,?,?,?,?,?,?)",
                                  (sb_id, station, shift_sp, op_id, dur, reason, created_at))
    # --- END DOWNTIME TRACKING LOGIC ---

    data = (
      sb_id, sf_pn, self.var_part_sf.get(), 
      rm_pns[0], rm_names[0], rm_pns[1], rm_names[1],
      rm_pns[2], rm_names[2], rm_pns[3], rm_names[3],
      self.var_b1.get(), self.var_b2.get(), self.var_b3.get(), qty,
      shift_sp, op_id, station, dt_sp, dt_line,
      shift_line, self.txt_remarks.get("1.0", tk.END).strip(),
      record_status, created_at, registered_by
    )

    try:
      c.execute('''
        INSERT INTO records (
          sub_batch_id, pn_sf, part_sf, 
          rm1_pn, rm1_name, rm2_pn, rm2_name,
          rm3_pn, rm3_name, rm4_pn, rm4_name,
          batch1, batch2, batch3, quantity,
          shift_sp, op_id, station, dt_sp, dt_line,
          shift_line, remarks, status, created_at, registered_by
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
      ''', data)
      conn.commit()
    except sqlite3.IntegrityError:
      conn.rollback()
      conn.close()
      messagebox.showerror("Error", "Duplicate Sub-Batch ID detected due to concurrent scans. Please try saving again.")
      return
    except Exception as e:
      conn.rollback()
      conn.close()
      messagebox.showerror("Error", f"Failed to save record: {e}")
      return
      
    conn.close()
    
    excel_data = [
      sb_id, sf_pn, self.var_part_sf.get(), 
      rm_pns[0], rm_names[0], rm_pns[1], rm_names[1],
      rm_pns[2], rm_names[2], rm_pns[3], rm_names[3],
      self.var_b1.get(), self.var_b2.get(), self.var_b3.get(), qty,
      shift_sp, op_id, station, dt_sp, dt_line,
      shift_line, self.txt_remarks.get("1.0", tk.END).strip(),
      registered_by
    ]
    
    self.background_excel_save(excel_data)

    self.lbl_status.config(text=f"Saved successfully: {sb_id}", fg=SUCCESS_COLOR)
    self.after(7000, lambda: self.lbl_status.config(text=""))
    
    self.update_stats()
    self.refresh_recent_treeview()
    self.refresh_records_treeview()
    self.update_sub_batch_preview()
    
    # Auto-print without preview
    if record_status == 'Direct to Line':
      return
      
    try:
      conn_print = get_db_connection()
      conn_print.row_factory = sqlite3.Row
      c_print = conn_print.cursor()
      c_print.execute("SELECT * FROM records WHERE sub_batch_id=?", (sb_id,))
      row_print = c_print.fetchone()
      conn_print.close()
      if row_print:
        print_html_slip(dict(row_print), silent=True)
    except Exception as e:
      print("Auto-print failed:", e)
    
    # update recent PNs sidebar
    pns = self.recent_pns_listbox.get(0, tk.END)
    if sf_pn not in pns:
      self.recent_pns_listbox.insert(0, sf_pn)
      if self.recent_pns_listbox.size() > 8:
        self.recent_pns_listbox.delete(tk.END)
        
    self.clear_form()

  def refresh_recent_treeview(self):
    for item in self.tree_recent.get_children():
      self.tree_recent.delete(item)
      
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("""SELECT r.sub_batch_id, r.pn_sf, r.quantity, r.shift_sp, r.station, r.dt_sp, IFNULL(SUM(CASE WHEN q.status='Closed' AND q.action_type IN ('Rework', 'Sorting', 'Use As-Is') THEN 0 ELSE q.qty_defective END), 0)
                 FROM records r 
                 LEFT JOIN quality_defects q ON r.sub_batch_id = q.sub_batch_id 
                 GROUP BY r.id ORDER BY r.id DESC LIMIT 5""")
    rows = c.fetchall()
    conn.close()
    
    for i, row in enumerate(rows):
      orig_qty = row[2]
      defects = row[6]
      qty_display = f"{orig_qty} ({orig_qty - defects} OK)" if defects > 0 else orig_qty
      display_tuple = (row[0], row[1], qty_display, row[3], row[4], row[5])
      
      tag = "even" if i % 2 == 0 else "odd"
      self.tree_recent.insert("", "end", values=display_tuple, tags=(tag,))

  def refresh_records_treeview(self, keep_page=False):
    if not keep_page:
      self.current_page = 1

    for item in self.tree_records.get_children():
      self.tree_records.delete(item)
      
    base_query = """SELECT r.sub_batch_id, r.pn_sf, r.part_sf, r.quantity, r.shift_sp, r.station, r.op_id, r.dt_sp, r.status, r.reprint_count, IFNULL(SUM(CASE WHEN q.status='Closed' AND q.action_type IN ('Rework', 'Sorting', 'Use As-Is') THEN 0 ELSE q.qty_defective END), 0)
               FROM records r LEFT JOIN quality_defects q ON r.sub_batch_id = q.sub_batch_id WHERE 1=1"""
    filters, params = [], []
    
    search = self.var_rec_search.get()
    if search:
      filters.append("(r.sub_batch_id LIKE ? OR r.pn_sf LIKE ? OR r.part_sf LIKE ? OR r.op_id LIKE ?)")
      params.extend([f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"])
      
    shift = self.cb_rec_shift.get()
    if shift != "All":
      filters.append("r.shift_sp = ?")
      params.append(shift)
      
    station = self.cb_rec_station.get()
    if station != "All":
      filters.append("r.station = ?")
      params.append(station)
      
    if filters:
      base_query += " AND " + " AND ".join(filters)
      
    base_query += " GROUP BY r.id ORDER BY r.id DESC"
    
    conn = get_db_connection()
    c = conn.cursor()
    c.execute(base_query, params)
    rows = c.fetchall()
    conn.close()
    
    total_count = len(rows)
    self.records_per_page = 20
    self.total_pages = (total_count + self.records_per_page - 1) // self.records_per_page if total_count > 0 else 1
    
    self.current_page = getattr(self, 'current_page', 1)
    if self.current_page > self.total_pages:
      self.current_page = self.total_pages
      
    start_idx = (self.current_page - 1) * self.records_per_page
    end_idx = start_idx + self.records_per_page
    display_rows = rows[start_idx:end_idx]
    
    for i, row in enumerate(display_rows):
      orig_qty = row[3]
      status_val = row[8]
      reprint_count = row[9] if row[9] is not None else 0
      defects = row[10]
      
      qty_display = f"{orig_qty} ({orig_qty - defects} OK)" if defects > 0 else orig_qty
      display_tuple = (row[0], row[1], row[2], qty_display, row[4], row[5], row[6], row[7], status_val, reprint_count)
      
      tags = ()
      if status_val == 'Partial Defect':
        tags = ("partial_defect",)
      elif reprint_count >= 1:
        tags = ("reprint_highlight",)
      elif status_val == 'pending':
        tags = ("pending",)
      else:
        tags = ("even" if i % 2 == 0 else "odd",)
      self.tree_records.insert("", "end", values=display_tuple, tags=tags)
      
    start_display = start_idx + 1 if total_count > 0 else 0
    end_display = min(end_idx, total_count)
    self.lbl_rec_count.config(text=f"Showing {start_display}-{end_display} of {total_count} records")
    
    self.render_pagination()

  def render_pagination(self):
    for widget in getattr(self, 'page_frame', tk.Frame()).winfo_children():
      widget.destroy()
      
    if not hasattr(self, 'page_frame') or not hasattr(self, 'total_pages') or self.total_pages <= 1:
      return
      
    btn_frame = tk.Frame(self.page_frame, bg=BG_COLOR)
    btn_frame.pack(anchor="center")
    
    def go_page(p):
      self.current_page = p
      self.refresh_records_treeview(keep_page=True)
      
    if self.current_page > 1:
      ttk.Button(btn_frame, text="< Prev", command=lambda: go_page(self.current_page - 1), width=6).pack(side=tk.LEFT, padx=2)
      
    start_p = max(1, self.current_page - 3)
    end_p = min(self.total_pages, self.current_page + 3)
    
    if start_p > 1:
      ttk.Button(btn_frame, text="1", command=lambda: go_page(1), width=3).pack(side=tk.LEFT, padx=1)
      if start_p > 2:
        tk.Label(btn_frame, text="...", bg=BG_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT, padx=1)
        
    for p in range(start_p, end_p + 1):
      if p == self.current_page:
        lbl = tk.Label(btn_frame, text=str(p), bg=ACCENT_COLOR, fg="white", font=("Segoe UI", 10, "bold"), width=3)
        lbl.pack(side=tk.LEFT, padx=2)
      else:
        ttk.Button(btn_frame, text=str(p), command=lambda p=p: go_page(p), width=3).pack(side=tk.LEFT, padx=1)
        
    if end_p < self.total_pages:
      if end_p < self.total_pages - 1:
        tk.Label(btn_frame, text="...", bg=BG_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT, padx=1)
      ttk.Button(btn_frame, text=str(self.total_pages), command=lambda: go_page(self.total_pages), width=3).pack(side=tk.LEFT, padx=1)
      
    if self.current_page < self.total_pages:
      ttk.Button(btn_frame, text="Next >", command=lambda: go_page(self.current_page + 1), width=6).pack(side=tk.LEFT, padx=2)

  def reset_records_filter(self):
    self.var_rec_search.set("")
    self.cb_rec_shift.set("All")
    self.cb_rec_station.set("All")
    self.refresh_records_treeview()

  def sort_treeview(self, tree, col, reverse):
    l = [(tree.set(k, col), k) for k in tree.get_children('')]
    try:
      l.sort(key=lambda t: int(t[0]), reverse=reverse)
    except ValueError:
      l.sort(reverse=reverse)
      
    for index, (val, k) in enumerate(l):
      tree.move(k, '', index)
    tree.heading(col, command=lambda: self.sort_treeview(tree, col, not reverse))

  def on_recent_double_click(self, event):
    item = self.tree_recent.selection()
    if not item: return
    sb_id = self.tree_recent.item(item[0], "values")[0]
    self.do_print(sb_id)

  def open_operator_scorecard(self):
    selected = self.tree_records.selection()
    if not selected: return
    item = self.tree_records.item(selected[0])
    op_id = item['values'][6]
    
    top = tk.Toplevel(self)
    top.title(f"Operator Scorecard: {op_id}")
    center_window(top, 600, 400)
    top.configure(bg=BG_COLOR)
    top.transient(self)
    
    # Header
    header = tk.Frame(top, bg=SURFACE_COLOR)
    header.pack(fill=tk.X, padx=10, pady=10)
    
    tk.Label(header, text="Performance Scorecard", font=("Inter", 14, "bold"), bg=SURFACE_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT, padx=10, pady=10)
    tk.Label(header, text=f"Operator: {op_id}", font=("Inter", 12), bg=SURFACE_COLOR, fg=ACCENT_COLOR).pack(side=tk.RIGHT, padx=10, pady=10)
    
    # Filter
    filter_frame = tk.Frame(top, bg=BG_COLOR)
    filter_frame.pack(fill=tk.X, padx=10, pady=5)
    
    tk.Label(filter_frame, text="Time Window:", bg=BG_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT)
    var_window = tk.StringVar(value="Today")
    
    def refresh_scorecard(*args):
      window = var_window.get()
      now = datetime.datetime.now()
      today_date = now.date()
      
      if window == "Today":
        start_dt = f"{today_date.strftime('%Y-%m-%d')} 00:00:00"
      elif window == "This Week":
        start_week = today_date - datetime.timedelta(days=today_date.weekday())
        start_dt = f"{start_week.strftime('%Y-%m-%d')} 00:00:00"
      else: # This Month
        start_month = today_date.replace(day=1)
        start_dt = f"{start_month.strftime('%Y-%m-%d')} 00:00:00"
        
      conn = get_db_connection()
      c = conn.cursor()
      
      # 1. Total Output & Shift Avg
      c.execute("SELECT op_id, SUM(quantity) FROM records WHERE dt_sp >= ? GROUP BY op_id", (start_dt,))
      op_data = c.fetchall()
      
      total_output = 0
      shift_total = 0
      op_count = len(op_data)
      for r in op_data:
        q = r[1] or 0
        shift_total += q
        if r[0] == str(op_id):
          total_output = q
          
      shift_avg = (shift_total / op_count) if op_count > 0 else 0
      
      # 2. Quality Rate & Quarantine Events
      c.execute("SELECT qty_defective, is_quarantined FROM quality_defects WHERE produced_by_op = ? AND reported_at >= ?", (str(op_id), start_dt))
      defects = c.fetchall()
      total_defects = sum(d[0] for d in defects if d[0])
      quarantine_events = sum(1 for d in defects if d[1] == 1)
      
      qual_rate = ((total_output - total_defects) / total_output * 100) if total_output > 0 else (100 if total_output > 0 else 0)
      
      conn.close()
      
      # Update UI
      lbl_out_val.config(text=f"{total_output}")
      lbl_avg_val.config(text=f"{shift_avg:.1f}")
      lbl_qual_val.config(text=f"{qual_rate:.1f}%" if total_output > 0 else "N/A")
      lbl_quar_val.config(text=f"{quarantine_events}")
      
    cb_window = ttk.Combobox(filter_frame, textvariable=var_window, values=["Today", "This Week", "This Month"], state="readonly", width=15)
    cb_window.pack(side=tk.LEFT, padx=10)
    cb_window.bind("<<ComboboxSelected>>", refresh_scorecard)
    
    # Metrics Grid
    metrics_frame = tk.Frame(top, bg=BG_COLOR)
    metrics_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    def create_metric_card(parent, title, row, col, fg=TEXT_COLOR):
      f = tk.Frame(parent, bg=SURFACE_COLOR, bd=1, relief="ridge")
      f.grid(row=row, column=col, sticky="nsew", padx=10, pady=10)
      tk.Label(f, text=title, font=("Inter", 10), bg=SURFACE_COLOR, fg=TEXT_MUTED).pack(pady=(15,5))
      lbl_val = tk.Label(f, text="0", font=("Inter", 24, "bold"), bg=SURFACE_COLOR, fg=fg)
      lbl_val.pack(pady=(0,15))
      return lbl_val
      
    metrics_frame.columnconfigure(0, weight=1)
    metrics_frame.columnconfigure(1, weight=1)
    metrics_frame.rowconfigure(0, weight=1)
    metrics_frame.rowconfigure(1, weight=1)
    
    lbl_out_val = create_metric_card(metrics_frame, "Total Output", 0, 0, SUCCESS_COLOR)
    lbl_avg_val = create_metric_card(metrics_frame, "Shift Average (All Ops)", 0, 1, ACCENT_COLOR)
    lbl_qual_val = create_metric_card(metrics_frame, "Quality Rate", 1, 0, SUCCESS_COLOR)
    lbl_quar_val = create_metric_card(metrics_frame, "Quarantine Events", 1, 1, WARNING_COLOR)
    
    refresh_scorecard()

  def print_selected_record(self):
    item = self.tree_records.selection()
    if not item: return
    sb_id = self.tree_records.item(item[0], "values")[0]
    self.do_print(sb_id)

  def print_last_slip(self):
    conn = get_db_connection()
    c = conn.cursor()
    c.execute("SELECT sub_batch_id FROM records ORDER BY id DESC LIMIT 1")
    row = c.fetchone()
    conn.close()
    if row:
      self.do_print(row[0])
    else:
      messagebox.showinfo("Info", "No records found.")

  def do_print(self, sb_id):
    conn = get_db_connection()
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM records WHERE sub_batch_id=?", (sb_id,))
    row = c.fetchone()
    conn.close()
    
    if row:
      record_data = dict(row)
      
      # Update reprint audit trail silently
      try:
        conn_update = get_db_connection()
        c_update = conn_update.cursor()
        now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        c_update.execute('''UPDATE records 
                 SET reprint_count = reprint_count + 1, 
                   last_reprinted_at = ?, 
                   last_reprinted_by = ? 
                 WHERE sub_batch_id = ?''', 
                (now_str, getattr(self, 'app_user_id', ''), sb_id))
        conn_update.commit()
        conn_update.close()
        self.refresh_records_treeview() # Update UI to show new reprint count
        
        record_data['reprint_count'] = int(record_data.get('reprint_count') or 0) + 1
        record_data['last_reprinted_by'] = getattr(self, 'app_user_id', '')
        record_data['last_reprinted_at'] = now_str
      except Exception as e:
        print(f"Failed to update reprint audit trail: {e}")
        
      # Print silently
      print_html_slip(record_data, silent=True)
      
    else:
      messagebox.showerror("Error", f"Record not found: {sb_id}")

  def build_tab_inventory(self):
    main_frame = tk.Frame(self.tab_inventory, bg=BG_COLOR)
    main_frame.pack(fill=tk.BOTH, expand=True)
    tk.Label(main_frame, text="Live Inventory Dashboard", bg=BG_COLOR, fg=ACCENT_COLOR, font=HMI_FONT_L).pack(pady=(10, 5))
    self.stats_frame = tk.Frame(main_frame, bg=BG_COLOR)
    self.stats_frame.pack(fill=tk.X, padx=20, pady=5)
    self.lbl_stat_total = self.create_stat_card(self.stats_frame, "Total Boxes in Rack", "0", ACCENT_COLOR)
    self.lbl_stat_oldest = self.create_stat_card(self.stats_frame, "Oldest Box Age", "0 days", ERROR_COLOR)
    self.lbl_stat_low = self.create_stat_card(self.stats_frame, "Low WIP Alerts", "0", WARNING_COLOR)
    
    # Control Panel Row
    control_frame = tk.Frame(main_frame, bg=SURFACE_COLOR, bd=1, relief="solid")
    control_frame.pack(fill=tk.X, padx=20, pady=10)
    
    # Filter
    filter_frame = tk.Frame(control_frame, bg=SURFACE_COLOR)
    filter_frame.pack(side=tk.LEFT, padx=20, pady=10)
    tk.Label(filter_frame, text="Filter by PN:", font=HMI_FONT_S, bg=SURFACE_COLOR, fg=TEXT_MUTED).pack(side=tk.LEFT)
    self.inv_search_var = tk.StringVar()
    self.inv_search_var.trace_add('write', lambda *args: self.refresh_inventory())
    tk.Entry(filter_frame, textvariable=self.inv_search_var, font=HMI_FONT_M, bg=BG_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT, padx=10)
    
    # Pick List Generator
    pick_frame = tk.Frame(control_frame, bg=SURFACE_COLOR)
    pick_frame.pack(side=tk.RIGHT, padx=20, pady=10)
    tk.Label(pick_frame, text="Pick-List Generator | PN:", font=HMI_FONT_S, bg=SURFACE_COLOR, fg=TEXT_MUTED).pack(side=tk.LEFT)
    self.pick_pn_var = tk.StringVar()
    self.pick_pn_combo = ttk.Combobox(pick_frame, textvariable=self.pick_pn_var, font=HMI_FONT_M, width=18, state="readonly")
    self.pick_pn_combo.pack(side=tk.LEFT, padx=5)
    tk.Label(pick_frame, text="Qty:", font=HMI_FONT_S, bg=SURFACE_COLOR, fg=TEXT_MUTED).pack(side=tk.LEFT)
    self.pick_qty_var = tk.StringVar()
    tk.Entry(pick_frame, textvariable=self.pick_qty_var, font=HMI_FONT_M, width=8, bg=BG_COLOR, fg=TEXT_COLOR).pack(side=tk.LEFT, padx=5)
    ttk.Button(pick_frame, text="Generate Ticket", style="Accent.TButton", command=self.generate_pick_list).pack(side=tk.LEFT, padx=10)
    
    # PanedWindow for the tables
    paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
    paned.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)

    left_pane = tk.Frame(paned, bg=BG_COLOR, width=560)
    paned.add(left_pane, weight=0)
    
    right_pane = tk.Frame(paned, bg=BG_COLOR)
    paned.add(right_pane, weight=1)
    
    table_paned = ttk.PanedWindow(left_pane, orient=tk.VERTICAL)
    table_paned.pack(fill=tk.BOTH, expand=True, padx=(0, 10))

    # Row 1: Summary Table
    row1 = tk.Frame(table_paned, bg=BG_COLOR)
    table_paned.add(row1, weight=1)

    _, agg_card = self.create_card(row1, "Summary by Part Number (Right-Click to Set Min Threshold)")
    cols_agg = ("PN", "Part Name", "Total", "Min", "Boxes")
    self.tree_inv_agg = ttk.Treeview(agg_card, columns=cols_agg, show="headings", height=8)
    for col in cols_agg: self.tree_inv_agg.heading(col, text=col)
    self.tree_inv_agg.column("PN", width=170); self.tree_inv_agg.column("Part Name", width=190); self.tree_inv_agg.column("Total", width=60); self.tree_inv_agg.column("Min", width=50); self.tree_inv_agg.column("Boxes", width=55)
    self.tree_inv_agg.pack(fill=tk.BOTH, expand=True, pady=10, padx=10)
    
    self.tree_inv_agg.tag_configure("low_wip", background="#FFEBEE", foreground="#C62828")
    
    def set_min_threshold(event):
      item = self.tree_inv_agg.identify_row(event.y)
      if not item: return
      self.tree_inv_agg.selection_set(item)
      pn = self.tree_inv_agg.item(item, "values")[0]
      qty_str = simpledialog.askstring("Set Threshold", f"Enter minimum required quantity for {pn}:", parent=self)
      if qty_str and qty_str.isdigit():
        try:
          conn = get_db_connection()
          c = conn.cursor()
          c.execute("INSERT OR REPLACE INTO part_thresholds (pn_sf, min_qty) VALUES (?, ?)", (pn, int(qty_str)))
          conn.commit()
          conn.close()
          self.refresh_inventory()
        except Exception as e:
          messagebox.showerror("Error", str(e))
    self.tree_inv_agg.bind("<Button-3>", set_min_threshold)
    
    # Row 2: Details Table
    row2 = tk.Frame(table_paned, bg=BG_COLOR)
    table_paned.add(row2, weight=1)
    
    _, det_card = self.create_card(row2, "Detailed Rack Content (FIFO)")
    cols_det = ("SB_ID", "PN", "Qty", "Age", "Status")
    self.tree_inv_det = ttk.Treeview(det_card, columns=cols_det, show="headings", height=10)
    for col in cols_det: self.tree_inv_det.heading(col, text=col)
    self.tree_inv_det.column("SB_ID", width=160); self.tree_inv_det.column("PN", width=159); self.tree_inv_det.column("Qty", width=52); self.tree_inv_det.column("Age", width=90); self.tree_inv_det.column("Status", width=65)
    
    self.tree_inv_det.tag_configure("age_green", background="#E8F5E9", foreground="#2E7D32")
    self.tree_inv_det.tag_configure("age_yellow", background="#FFFDE7", foreground="#F9A825")
    self.tree_inv_det.tag_configure("age_red", background="#FFEBEE", foreground="#C62828")
    
    scrolldet = ttk.Scrollbar(det_card, orient=tk.VERTICAL, command=self.tree_inv_det.yview)
    self.tree_inv_det.configure(yscroll=scrolldet.set)
    scrolldet.pack(side=tk.RIGHT, fill=tk.Y)
    self.tree_inv_det.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
    
    # Right Pane: WIP Chart
    _, chart1_card = self.create_card(right_pane, "WIP Level by Part Number")
    
    if MATPLOTLIB_AVAILABLE:
      self.fig_wip = Figure(figsize=(9, 6.5), dpi=100, facecolor=SURFACE_COLOR)
      self.ax_wip = self.fig_wip.add_subplot(111)
      self.style_ax(self.ax_wip, "")
      self.lbl_chart_wip = tk.Label(chart1_card, bg=SURFACE_COLOR)
      self.lbl_chart_wip.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

  def create_stat_card(self, parent, title, initial_value, color):
    frame = tk.Frame(parent, bg=SURFACE_COLOR, bd=1, relief="solid", padx=20, pady=10)
    frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)
    tk.Label(frame, text=title, font=HMI_FONT_S, bg=SURFACE_COLOR, fg=TEXT_MUTED).pack()
    lbl = tk.Label(frame, text=initial_value, font=HMI_FONT_XL, bg=SURFACE_COLOR, fg=color)
    lbl.pack()
    return lbl

  def generate_pick_list(self):
    pn = self.pick_pn_var.get().strip()
    qty_str = self.pick_qty_var.get().strip()
    if not pn or not qty_str.isdigit():
      messagebox.showerror("Error", "Please enter a valid Part Number and numeric Quantity.")
      return
    target_qty = int(qty_str)
    
    conn = get_db_connection()
    c = conn.cursor()
    c.execute('''SELECT sub_batch_id, quantity, dt_sp 
           FROM records 
           WHERE status="In Rack" AND pn_sf=? 
           ORDER BY dt_sp ASC''', (pn,))
    rows = c.fetchall()
    conn.close()
    
    if not rows:
      messagebox.showinfo("Pick List", f"No inventory found in rack for {pn}")
      return
      
    total_rack_qty = sum(r[1] for r in rows)
    
    picked = []
    accumulated = 0
    for r in rows:
      picked.append((r[0], r[1], r[2]))
      accumulated += r[1]
      if accumulated >= target_qty:
        break
        
    if total_rack_qty < target_qty:
      messagebox.showwarning("Shortage", f"Only {total_rack_qty} units available in the rack. Cannot fulfill {target_qty}.")
      
    # Display the Pick List
    popup = tk.Toplevel(self)
    popup.title(f"FIFO Pick Ticket - {pn}")
    center_window(popup, 500, 600)
    popup.configure(bg=BG_COLOR)
    popup.transient(self)
    popup.grab_set()
    
    tk.Label(popup, text=f"FIFO Pick Ticket", font=HMI_FONT_L, bg=BG_COLOR, fg=ACCENT_COLOR).pack(pady=10)
    tk.Label(popup, text=f"Part: {pn} | Target: {target_qty} | Actual: {total_rack_qty}", font=HMI_FONT_M, bg=BG_COLOR, fg=TEXT_COLOR).pack(pady=5)
    
    frame = tk.Frame(popup, bg=SURFACE_COLOR)
    frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    
    for i, (sb, q, dt) in enumerate(picked):
      tk.Label(frame, text=f"{i+1}. {sb} (Qty: {q})", font=HMI_FONT_M, bg=SURFACE_COLOR, fg=TEXT_COLOR, anchor="w").pack(fill=tk.X, padx=10, pady=2)
      tk.Label(frame, text=f"  Stored: {dt}", font=HMI_FONT_S, bg=SURFACE_COLOR, fg=TEXT_MUTED, anchor="w").pack(fill=tk.X, padx=10)

    btn_frame = tk.Frame(popup, bg=BG_COLOR)
    btn_frame.pack(pady=20)
    
    def print_ticket():
      zpl_data = generate_pick_ticket_zpl(pn, target_qty, total_rack_qty, picked)
      execute_ticket_print(zpl_data, pn)
      
    ttk.Button(btn_frame, text="Print Ticket", style="Accent.TButton", command=print_ticket).pack(side=tk.LEFT, padx=10)
    ttk.Button(btn_frame, text="Close", style="Secondary.TButton", command=popup.destroy).pack(side=tk.LEFT, padx=10)

  def refresh_inventory(self):
    search_q = getattr(self, 'inv_search_var', tk.StringVar()).get().strip().lower()
    
    conn = get_db_connection()
    c = conn.cursor()
    
    # Aggregated with threshold
    c.execute('''SELECT r.pn_sf, r.part_sf, SUM(r.quantity), COUNT(r.id), COALESCE(t.min_qty, 0)
           FROM records r
           LEFT JOIN part_thresholds t ON r.pn_sf = t.pn_sf
           WHERE r.status="In Rack"
           GROUP BY r.pn_sf ORDER BY SUM(r.quantity) DESC''')
    agg_rows = c.fetchall()
    
    # Detailed
    c.execute('''SELECT sub_batch_id, pn_sf, quantity, dt_sp, status 
           FROM records WHERE status="In Rack" ORDER BY pn_sf, dt_sp ASC''')
    det_rows = c.fetchall()
    conn.close()
    
    for item in self.tree_inv_agg.get_children(): self.tree_inv_agg.delete(item)
    for item in self.tree_inv_det.get_children(): self.tree_inv_det.delete(item)
    
    pn_labels = []
    qty_values = []
    min_values = []
    part_labels = []
    full_pns = []
    low_wip_count = 0
    total_boxes = 0
    
    for row in agg_rows:
      pn, part, total_qty, box_count, min_qty = row
      if search_q and search_q not in pn.lower() and search_q not in part.lower():
        continue
      
      tag = ""
      if total_qty < min_qty:
        tag = "low_wip"
        low_wip_count += 1
        
      self.tree_inv_agg.insert("", "end", values=(pn, part, total_qty, min_qty, box_count), tags=(tag,))
      pn_labels.append(pn[:15])
      part_labels.append(part[:25])
      full_pns.append(pn)
      qty_values.append(total_qty or 0)
      min_values.append(min_qty or 0)
      total_boxes += box_count
      
    now = datetime.datetime.now()
    age_counts = {"< 2 Days": 0, "2-7 Days": 0, "> 7 Days": 0}
    max_age_days = 0
    max_age_hours = 0
    
    # Track oldest box per PN to add star
    pn_oldest_dt = {}
    for row in det_rows:
      pn = row[1]
      dt_str = row[3]
      try:
        try:
          dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
          dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
        if pn not in pn_oldest_dt or dt_obj < pn_oldest_dt[pn]:
          pn_oldest_dt[pn] = dt_obj
      except: pass
      
    for row in det_rows:
      pn = row[1]
      if search_q and search_q not in pn.lower():
        continue
        
      dt_str = row[3]
      try:
        try:
          dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
          dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
        diff = now - dt_obj
        hours = diff.total_seconds() / 3600
        days = diff.days
        
        if days > max_age_days or (days == max_age_days and int(hours % 24) > max_age_hours):
          max_age_days = days
          max_age_hours = int(hours % 24)
        
        age_str = f"{days}d {int(hours % 24)}h"
        if dt_obj == pn_oldest_dt.get(pn):
          age_str += " ⭐"
        
        if days < 2:
          tag = "age_green"
          age_counts["< 2 Days"] += 1
        elif days <= 7:
          tag = "age_yellow"
          age_counts["2-7 Days"] += 1
        else:
          tag = "age_red"
          age_counts["> 7 Days"] += 1
      except Exception:
        age_str = "Unknown"
        tag = "age_green"
        age_counts["< 2 Days"] += 1
        
      vals = (row[0], pn, row[2], age_str, row[4])
      self.tree_inv_det.insert("", "end", values=vals, tags=(tag,))
      
    # Update Quick Stats
    if hasattr(self, 'lbl_stat_total'):
      self.lbl_stat_total.config(text=str(total_boxes))
      self.lbl_stat_oldest.config(text=f"{max_age_days}d {max_age_hours}h")
      self.lbl_stat_low.config(text=str(low_wip_count))
      
    # Update Charts
    if MATPLOTLIB_AVAILABLE and hasattr(self, 'ax_wip'):
      self.ax_wip.clear()
      self.style_ax(self.ax_wip, "")
      if part_labels:
        import textwrap
        top_part = [textwrap.fill(lbl, width=15) for lbl in part_labels[:8]]
        top_qty = qty_values[:8]
        top_min = min_values[:8]
        
        x = np.arange(len(top_part))
        width = 0.35
        
        rects1 = self.ax_wip.bar(x - width/2, top_qty, width, color='#38C958')
        rects2 = self.ax_wip.bar(x + width/2, top_min, width, color='#FFA000')
        
        for rect in rects1 + rects2:
          height = rect.get_height()
          if height > 0:
            self.ax_wip.annotate(f"{int(height)}",
                       xy=(rect.get_x() + rect.get_width() / 2, height),
                       xytext=(0, 3),
                       textcoords="offset points",
                       ha='center', va='bottom', color=TEXT_COLOR, fontsize=8)
        
        self.ax_wip.set_xticks(x)
        self.ax_wip.set_xticklabels(top_part, rotation=20, ha='right', fontsize=8)
        self.ax_wip.tick_params(axis='y', labelsize=8)
      else:
        self.ax_wip.text(0.5, 0.5, "No Data", color=TEXT_MUTED, ha="center")
      self.fig_wip.subplots_adjust(left=0.08, right=0.96, top=0.92, bottom=0.22)
      self.render_chart_to_label(self.fig_wip, self.lbl_chart_wip)

    if hasattr(self, 'pick_pn_combo'):
      self.pick_pn_combo['values'] = full_pns


  def render_chart_to_label(self, fig, label_widget):
    if not MATPLOTLIB_AVAILABLE: return
    canvas = FigureCanvasAgg(fig)
    canvas.draw()
    rgba = np.asarray(canvas.buffer_rgba())
    img = Image.fromarray(rgba)
    photo = ImageTk.PhotoImage(img)
    label_widget.configure(image=photo)
    label_widget.image = photo

  def style_ax(self, ax, title):
    ax.set_facecolor(SURFACE_COLOR)
    ax.tick_params(colors=TEXT_COLOR)
    ax.set_title(title, color=TEXT_COLOR, pad=15)
    for spine in ax.spines.values():
      spine.set_color(BORDER_COLOR)

  def build_tab3(self):
    main_frame = tk.Frame(self.tab3, bg=BG_COLOR)
    main_frame.pack(fill=tk.BOTH, expand=True)

    if not MATPLOTLIB_AVAILABLE:
      tk.Label(main_frame, text=" Matplotlib is required to view KPIs.\nPlease run 'pip install matplotlib' and restart.", bg=BG_COLOR, fg=ERROR_COLOR, font=HMI_FONT_L).pack(pady=50, padx=50)
      return

    # Top Cards Row
    self.kpi_cards_frame = tk.Frame(main_frame, bg=BG_COLOR)
    self.kpi_cards_frame.pack(fill=tk.X, pady=10, padx=20)
    
    self.lbl_kpi_total = tk.Label(self.kpi_cards_frame, text="Daily Total: 0", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L, padx=20, pady=10)
    self.lbl_kpi_total.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=5)
    
    self.lbl_kpi_top_op = tk.Label(self.kpi_cards_frame, text="Top Operator: -", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L, padx=20, pady=10)
    self.lbl_kpi_top_op.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=5)
    
    self.lbl_kpi_oee = tk.Label(self.kpi_cards_frame, text="Quality: - | Perf: -", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L, padx=20, pady=10)
    self.lbl_kpi_oee.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=5)

    self.lbl_kpi_dtl = tk.Label(self.kpi_cards_frame, text="Direct to Line: 0", bg=SURFACE_COLOR, fg=WARNING_COLOR, font=HMI_FONT_L, padx=20, pady=10)
    self.lbl_kpi_dtl.pack(side=tk.LEFT, expand=True, fill=tk.BOTH, padx=5)
    
    controls_frame = tk.Frame(main_frame, bg=BG_COLOR)
    controls_frame.pack(pady=5)

    tk.Label(controls_frame, text="Select Date:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(side=tk.LEFT, padx=10)
    self.kpi_date_entry = DateEntry(controls_frame, width=12, background=ACCENT_COLOR, foreground='white', borderwidth=2)
    self.kpi_date_entry.pack(side=tk.LEFT, padx=10)

    tk.Label(controls_frame, text="Shift:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(side=tk.LEFT, padx=10)
    self.cb_kpi_shift = ttk.Combobox(controls_frame, values=["All", "A", "B", "C"], state="readonly", width=8, font=HMI_FONT_M)
    self.cb_kpi_shift.pack(side=tk.LEFT, padx=5)

    user_role = getattr(self, 'app_user_role', 'Operator')
    if user_role in ["Admin", "Supervisor", "Quality OP"]:
      self.cb_kpi_shift.set("All")
    else:
      self.cb_kpi_shift.set(getattr(self, 'app_user_shift', 'A'))

    ttk.Button(controls_frame, text="Refresh KPIs", style="Primary.TButton", command=self.refresh_kpis).pack(side=tk.LEFT, padx=10)

    # Notebook for Charts
    self.kpi_notebook = ttk.Notebook(main_frame)
    self.kpi_notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
    
    self.kpi_tab_daily = tk.Frame(self.kpi_notebook, bg=BG_COLOR)
    self.kpi_tab_trends = tk.Frame(self.kpi_notebook, bg=BG_COLOR)
    self.kpi_tab_downtime = tk.Frame(self.kpi_notebook, bg=BG_COLOR)
    
    self.kpi_notebook.add(self.kpi_tab_daily, text="Daily Production KPIs")
    self.kpi_notebook.add(self.kpi_tab_downtime, text="Downtime Analysis")
    self.kpi_notebook.add(self.kpi_tab_trends, text="30-Day WIP Trends")
    
    # --- DOWNTIME TAB CHARTS ---
    dt_top_frame = tk.Frame(self.kpi_tab_downtime, bg=BG_COLOR)
    dt_top_frame.pack(fill=tk.BOTH, expand=True, pady=5, padx=20)
    
    self.fig_dt_reasons = Figure(figsize=(6, 3.2), dpi=100)
    self.ax_dt_reasons = self.fig_dt_reasons.add_subplot(111)
    self.lbl_chart_dt_reasons = tk.Label(dt_top_frame, bg=BG_COLOR)
    self.lbl_chart_dt_reasons.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

    self.fig_dt_station = Figure(figsize=(6, 3.2), dpi=100)
    self.ax_dt_station = self.fig_dt_station.add_subplot(111)
    self.lbl_chart_dt_station = tk.Label(dt_top_frame, bg=BG_COLOR)
    self.lbl_chart_dt_station.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
    
    dt_bottom_frame = tk.Frame(self.kpi_tab_downtime, bg=BG_COLOR)
    dt_bottom_frame.pack(fill=tk.BOTH, expand=True, pady=5, padx=20)
    
    self.fig_dt_pn = Figure(figsize=(6, 3.2), dpi=100)
    self.ax_dt_pn = self.fig_dt_pn.add_subplot(111)
    self.lbl_chart_dt_pn = tk.Label(dt_bottom_frame, bg=BG_COLOR)
    self.lbl_chart_dt_pn.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

    self.fig_dt_op = Figure(figsize=(6, 3.2), dpi=100)
    self.ax_dt_op = self.fig_dt_op.add_subplot(111)
    self.lbl_chart_dt_op = tk.Label(dt_bottom_frame, bg=BG_COLOR)
    self.lbl_chart_dt_op.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)

    # Charts Area (Daily)
    self.charts_frame_1 = tk.Frame(self.kpi_tab_daily, bg=BG_COLOR)
    self.charts_frame_1.pack(fill=tk.X, pady=5, padx=20)
    
    self.fig_shift = Figure(figsize=(15, 3.5), dpi=100)
    self.ax_shift = self.fig_shift.add_subplot(111)
    self.lbl_chart_shift = tk.Label(self.charts_frame_1, bg=BG_COLOR)
    self.lbl_chart_shift.pack(fill=tk.BOTH, expand=True)

    self.charts_frame_2 = tk.Frame(self.kpi_tab_daily, bg=BG_COLOR)
    self.charts_frame_2.pack(fill=tk.BOTH, expand=True, pady=5, padx=20)
    
    self.fig_hourly = Figure(figsize=(6, 3.2), dpi=100)
    self.ax_hourly = self.fig_hourly.add_subplot(111)
    self.lbl_chart_hourly = tk.Label(self.charts_frame_2, bg=BG_COLOR)
    self.lbl_chart_hourly.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
    
    self.fig_op = Figure(figsize=(4.5, 3.2), dpi=100)
    self.ax_op = self.fig_op.add_subplot(111)
    self.lbl_chart_op = tk.Label(self.charts_frame_2, bg=BG_COLOR)
    self.lbl_chart_op.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
    
    self.fig_pn = Figure(figsize=(4.5, 3.2), dpi=100)
    self.ax_pn = self.fig_pn.add_subplot(111)
    self.lbl_chart_pn = tk.Label(self.charts_frame_2, bg=BG_COLOR)
    self.lbl_chart_pn.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5)
    
    # 30-Day Trend Chart Area (Trends)
    tk.Label(self.kpi_tab_trends, text="WIP Trend Analysis (Last 30 Days)", bg=BG_COLOR, fg=ACCENT_COLOR, font=HMI_FONT_L).pack(pady=(20, 5))
    self.charts_frame_3 = tk.Frame(self.kpi_tab_trends, bg=BG_COLOR)
    self.charts_frame_3.pack(fill=tk.BOTH, expand=True, pady=10, padx=20)
    
    self.fig_trend = Figure(figsize=(14, 6), dpi=100)
    self.ax_trend = self.fig_trend.add_subplot(111)
    self.lbl_chart_trend = tk.Label(self.charts_frame_3, bg=BG_COLOR)
    self.lbl_chart_trend.pack(fill=tk.BOTH, expand=True, padx=10)
    
    # Style all figures
    for fig, ax, title in [(self.fig_hourly, self.ax_hourly, "Hourly Production"),
                (self.fig_shift, self.ax_shift, "Shift Output"),
                (self.fig_op, self.ax_op, "Operator Mix"),
                (self.fig_pn, self.ax_pn, "Product Mix"),
                (self.fig_trend, self.ax_trend, "30-Day WIP Accumulation by Part Number")]:
      fig.patch.set_facecolor(BG_COLOR)
      self.style_ax(ax, title)
      
    self.render_chart_to_label(self.fig_hourly, self.lbl_chart_hourly)
    self.render_chart_to_label(self.fig_shift, self.lbl_chart_shift)
    self.render_chart_to_label(self.fig_op, self.lbl_chart_op)
    self.render_chart_to_label(self.fig_pn, self.lbl_chart_pn)
    self.render_chart_to_label(self.fig_trend, self.lbl_chart_trend)
      
  def refresh_kpis(self):
    if not MATPLOTLIB_AVAILABLE:
      return
    try:
      conn = get_db_connection()
      c = conn.cursor()
      
      if hasattr(self, 'kpi_date_entry'):
        selected_date = self.kpi_date_entry.get_date()
      else:
        selected_date = datetime.datetime.now().date()
        
      start_dt = selected_date.strftime("%Y-%m-%d 00:00:00")
      end_dt = selected_date.strftime("%Y-%m-%d 23:59:59")
      
      shift_filter = getattr(self, 'cb_kpi_shift', None)
      selected_shift = shift_filter.get() if shift_filter else "All"
      
      if selected_shift == "All":
        shift_cond = ""
        shift_params = []
      else:
        shift_cond = " AND shift_sp = ?"
        shift_params = [selected_shift]
      
      # 1. Total Daily Production
      c.execute(f"SELECT SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ?{shift_cond}", [start_dt, end_dt] + shift_params)
      total_daily = c.fetchone()[0] or 0
      self.lbl_kpi_total.config(text=f"Daily Total: {total_daily}")
      
      c.execute(f"SELECT SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ? AND status='Direct to Line'{shift_cond}", [start_dt, end_dt] + shift_params)
      dtl_daily = c.fetchone()[0] or 0
      dtl_pct = (dtl_daily / total_daily * 100) if total_daily > 0 else 0
      self.lbl_kpi_dtl.config(text=f"Direct to Line: {dtl_daily} ({dtl_pct:.1f}%)")
      
      # 2. Top Operator
      c.execute(f"SELECT op_id, SUM(quantity) as q FROM records WHERE dt_sp >= ? AND dt_sp <= ?{shift_cond} GROUP BY op_id ORDER BY q DESC LIMIT 1", [start_dt, end_dt] + shift_params)
      top_op = c.fetchone()
      top_op_txt = f"{top_op[0]} ({top_op[1]})" if top_op else "-"
      self.lbl_kpi_top_op.config(text=f"Top Operator: {top_op_txt}")
      
      # OEE Calculations
      c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='quality_defects'")
      has_qual = c.fetchone()
      defective_qty = 0
      if has_qual:
        c.execute(f"SELECT SUM(CASE WHEN status='Closed' AND action_type IN ('Rework', 'Sorting', 'Use As-Is') THEN 0 ELSE qty_defective END) FROM quality_defects WHERE reported_at >= ? AND reported_at <= ?{shift_cond}", [start_dt, end_dt] + shift_params)
        defective_qty = c.fetchone()[0] or 0
        
      qual_rate = ((total_daily - defective_qty) / total_daily * 100) if total_daily > 0 else 0
      
      # Fetch Targets
      c.execute("SELECT product_pn, target_qty FROM shift_targets ORDER BY id ASC")
      pn_targets_dict = {}
      for row in c.fetchall():
        pn, tqty = row
        pn_targets_dict[pn] = tqty
        
      total_target = sum(pn_targets_dict.values())
          
      perf_rate = (total_daily / total_target * 100) if total_target > 0 else 0
      self.lbl_kpi_oee.config(text=f"Quality: {qual_rate:.1f}% | Perf: {perf_rate:.1f}%")
      
      # 3. Hourly Production
      self.ax_hourly.clear()
      self.style_ax(self.ax_hourly, "Hourly Production")
      
      c.execute(f"SELECT substr(dt_sp, 12, 2) as hr, SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ?{shift_cond} GROUP BY hr", [start_dt, end_dt] + shift_params)
      hourly_data = {row[0]: (row[1] or 0) for row in c.fetchall() if row[0] is not None}
      hrs = [f"{h:02d}" for h in range(24)]
      qts = [hourly_data.get(h, 0) for h in hrs]
      
      self.ax_hourly.plot(hrs, qts, color=ACCENT_COLOR, marker='o')
      max_qt = max(qts) if qts else 0
      self.ax_hourly.set_ylim(bottom=0, top=max(max_qt * 1.15, 1))
      for i, v in enumerate(qts):
        if v > 0:
          self.ax_hourly.text(hrs[i], v + (max_qt*0.02 if max_qt>0 else 1), str(v), color=TEXT_COLOR, ha='center', va='bottom', fontsize=8)
      
      self.ax_hourly.grid(True, color=BORDER_COLOR, alpha=0.3)
      self.ax_hourly.set_xticks(hrs)
      self.fig_hourly.subplots_adjust(bottom=0.15, top=0.85, right=0.95, left=0.12)
      self.render_chart_to_label(self.fig_hourly, self.lbl_chart_hourly)
      
      # 4. PN Comparison (Target vs Actual)
      self.ax_shift.clear()
      self.style_ax(self.ax_shift, "Actual vs Target per PN")
      
      c.execute(f"SELECT pn_sf, SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ?{shift_cond} GROUP BY pn_sf", [start_dt, end_dt] + shift_params)
      pn_actuals_dict = {row[0]: (row[1] or 0) for row in c.fetchall() if row[0] is not None}
      
      pns = list(set(pn_actuals_dict.keys()).union(set(pn_targets_dict.keys())))
      pns = sorted([p for p in pns if pn_actuals_dict.get(p, 0) > 0])
      
      sqts = [pn_actuals_dict.get(p, 0) for p in pns]
      tqts = [pn_targets_dict.get(p, 0) for p in pns]
      
      # Use part name from SF_DATA if available, else short part number
      pn_labels = []
      for p in pns:
        if p in SF_DATA and SF_DATA[p][0]:
          pn_labels.append(SF_DATA[p][0])
        else:
          pn_labels.append(p.split('-')[-1] if '-' in p else p[:8])
      
      x = range(len(pns))
      width = 0.2
      
      if len(pns) > 0:
        bars1 = self.ax_shift.bar([i - width/2 for i in x], sqts, width, color=SUCCESS_COLOR, label='Actual')
        bars2 = self.ax_shift.bar([i + width/2 for i in x], tqts, width, color=WARNING_COLOR, label='Target')
        
        max_sqt = max(max(sqts) if sqts else 0, max(tqts) if tqts else 0)
        self.ax_shift.set_ylim(bottom=0, top=max(max_sqt * 1.15, 1))
        
        for bars in [bars1, bars2]:
          for bar in bars:
            yval = bar.get_height()
            if yval > 0:
              self.ax_shift.text(bar.get_x() + bar.get_width()/2.0, yval + (max_sqt*0.02 if max_sqt>0 else 1), int(yval), va='bottom', ha='center', color=TEXT_COLOR, fontsize=8)
              
        self.ax_shift.set_xticks(x)
        import textwrap
        pn_labels_wrapped = [textwrap.fill(lbl, width=15) for lbl in pn_labels]
        self.ax_shift.set_xticklabels(pn_labels_wrapped, rotation=20, ha='right', fontsize=8)
        self.ax_shift.legend(loc='upper right', ncol=2, fontsize=8, facecolor=BG_COLOR, edgecolor=BORDER_COLOR, labelcolor=TEXT_COLOR)
      else:
        self.ax_shift.text(0.5, 0.5, "No Data", ha='center', va='center', color=TEXT_MUTED, transform=self.ax_shift.transAxes)
      self.fig_shift.subplots_adjust(bottom=0.40, top=0.90, right=0.98, left=0.05)
      self.render_chart_to_label(self.fig_shift, self.lbl_chart_shift)
      
      # 5. Production by Operator Pie
      self.ax_op.clear()
      self.style_ax(self.ax_op, "Operator Mix")
      
      c.execute(f"SELECT op_id, SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ?{shift_cond} GROUP BY op_id", [start_dt, end_dt] + shift_params)
      op_data = c.fetchall()
      if op_data:
        ops = [f"Op: {row[0]}" for row in op_data]
        o_qts = [row[1] for row in op_data]
        wedges, texts, autotexts = self.ax_op.pie(o_qts, autopct='%1.1f%%', textprops={'color': 'white', 'fontsize': 8, 'weight': 'bold'}, radius=0.9, startangle=140)
        self.ax_op.legend(wedges, ops, loc="center left", bbox_to_anchor=(0.85, 0.5), fontsize=8, facecolor=BG_COLOR, edgecolor=BORDER_COLOR, labelcolor=TEXT_COLOR)
      self.fig_op.tight_layout(rect=(0, 0, 1, 0.95))
      self.render_chart_to_label(self.fig_op, self.lbl_chart_op)
      
      # 6. Top PNs Pie
      self.ax_pn.clear()
      self.style_ax(self.ax_pn, "Product Mix")
      
      c.execute(f"SELECT pn_sf, SUM(quantity) FROM records WHERE dt_sp >= ? AND dt_sp <= ?{shift_cond} GROUP BY pn_sf ORDER BY SUM(quantity) DESC LIMIT 5", [start_dt, end_dt] + shift_params)
      pn_data = c.fetchall()
      if pn_data:
        pns_labels = []
        import textwrap
        for row in pn_data:
          p = row[0]
          if p in SF_DATA and SF_DATA[p][0]:
            pns_labels.append(textwrap.fill(SF_DATA[p][0], width=15))
          else:
            pns_labels.append(p.split('-')[-1] if '-' in p else p[:8])
        p_qts = [row[1] for row in pn_data]
        wedges, texts, autotexts = self.ax_pn.pie(p_qts, autopct='%1.1f%%', textprops={'color': 'white', 'fontsize': 8, 'weight': 'bold'}, radius=0.9, startangle=140)
        self.ax_pn.legend(wedges, pns_labels, loc="center left", bbox_to_anchor=(0.85, 0.5), fontsize=8, facecolor=BG_COLOR, edgecolor=BORDER_COLOR, labelcolor=TEXT_COLOR)
      self.fig_pn.tight_layout(rect=(0, 0, 1, 0.95))
      self.render_chart_to_label(self.fig_pn, self.lbl_chart_pn)

      # 7. 30-Day WIP Trend Chart
      self.ax_trend.clear()
      self.style_ax(self.ax_trend, "30-Day WIP Accumulation by Part Number")
      thirty_days_ago = (datetime.datetime.now() - datetime.timedelta(days=30)).strftime("%Y-%m-%d")
      c.execute('''SELECT snapshot_date, pn_sf, total_qty_in_rack 
             FROM inventory_snapshots 
             WHERE snapshot_date >= ?
             ORDER BY snapshot_date ASC''', (thirty_days_ago,))
      trend_data = c.fetchall()
      
      if trend_data:
        pn_series = {}
        dates_set = set()
        for r in trend_data:
          dt, pn, qty = r
          if pn not in pn_series: pn_series[pn] = {}
          pn_series[pn][dt] = qty
          dates_set.add(dt)
        
        sorted_dates = sorted(list(dates_set))
        x_labels = [d[5:] for d in sorted_dates]
        
        for pn, data_dict in pn_series.items():
          y_vals = [data_dict.get(d, 0) for d in sorted_dates]
          self.ax_trend.plot(x_labels, y_vals, marker='o', markersize=4, label=pn)
          
        self.ax_trend.legend(loc='center left', bbox_to_anchor=(1.02, 0.5), ncol=1, fontsize=8, facecolor=BG_COLOR, edgecolor=BORDER_COLOR, labelcolor=TEXT_COLOR)
        self.ax_trend.tick_params(axis='x', rotation=45, labelsize=8)
      else:
        self.ax_trend.text(0.5, 0.5, "No Snapshot Data Yet\n(Snapshots are taken nightly at 23:50)", ha='center', va='center', color=TEXT_MUTED, transform=self.ax_trend.transAxes)
        
      self.fig_trend.subplots_adjust(bottom=0.3, top=0.85, right=0.75)
      self.render_chart_to_label(self.fig_trend, self.lbl_chart_trend)

      
      # --- 8. Downtime Analytics ---
      if selected_shift == "All":
        dt_shift_cond = ""
        dt_params = [start_dt, end_dt]
      else:
        dt_shift_cond = " AND d.shift = ?"
        dt_params = [start_dt, end_dt, selected_shift]
        
      for fig in [self.fig_dt_reasons, self.fig_dt_station, self.fig_dt_pn, self.fig_dt_op]:
        fig.patch.set_facecolor(BG_COLOR)
        
      self.ax_dt_reasons.clear()
      self.style_ax(self.ax_dt_reasons, "Downtime by Reason")
      self.ax_dt_station.clear()
      self.style_ax(self.ax_dt_station, "Downtime by Station (Mins)")
      self.ax_dt_pn.clear()
      self.style_ax(self.ax_dt_pn, "Downtime by Product (Mins)")
      self.ax_dt_op.clear()
      self.style_ax(self.ax_dt_op, "Downtime by Operator (Mins)")
      
      # 8A. Top Reasons
      c.execute(f"SELECT CASE WHEN d.reason LIKE 'Other:%' THEN 'Other' ELSE d.reason END as short_reason, SUM(d.duration_min) FROM downtime_logs d WHERE d.created_at >= ? AND d.created_at <= ?{dt_shift_cond} GROUP BY CASE WHEN d.reason LIKE 'Other:%' THEN 'Other' ELSE d.reason END", dt_params)
      reason_data = c.fetchall()
      if reason_data:
        r_labels = [row[0][:25] for row in reason_data]
        r_vals = [row[1] for row in reason_data]
        wedges, texts, autotexts = self.ax_dt_reasons.pie(r_vals, autopct='%1.0f%%', textprops={'color': 'white', 'fontsize': 8, 'weight': 'bold'}, radius=0.9)
        self.ax_dt_reasons.legend(wedges, r_labels, loc="upper center", bbox_to_anchor=(0.5, 0.0), ncol=2, fontsize=8, facecolor=BG_COLOR, edgecolor=BORDER_COLOR, labelcolor=TEXT_COLOR)
      else:
        self.ax_dt_reasons.text(0.5, 0.5, "No Delays Recorded", ha='center', va='center', color=TEXT_MUTED, transform=self.ax_dt_reasons.transAxes)
      self.fig_dt_reasons.subplots_adjust(bottom=0.25, top=0.85, right=0.95, left=0.15)
      self.render_chart_to_label(self.fig_dt_reasons, self.lbl_chart_dt_reasons)
      
      # 8B. By Station
      c.execute(f"SELECT d.station, SUM(d.duration_min) FROM downtime_logs d WHERE d.created_at >= ? AND d.created_at <= ?{dt_shift_cond} GROUP BY d.station ORDER BY SUM(d.duration_min) DESC LIMIT 5", dt_params)
      st_data = c.fetchall()
      if st_data:
        x_lbls = [row[0][:10] for row in st_data]
        y_vals = [row[1] for row in st_data]
        self.ax_dt_station.bar(x_lbls, y_vals, color=WARNING_COLOR, width=0.4)
        self.ax_dt_station.set_xlim(-0.5, max(1.5, len(x_lbls)-0.5))
        self.ax_dt_station.tick_params(axis='x', rotation=20, labelsize=8)
      else:
        self.ax_dt_station.text(0.5, 0.5, "No Data", ha='center', va='center', color=TEXT_MUTED, transform=self.ax_dt_station.transAxes)
      self.fig_dt_station.subplots_adjust(bottom=0.25, top=0.85, right=0.95, left=0.15)
      self.render_chart_to_label(self.fig_dt_station, self.lbl_chart_dt_station)
      
      # 8C. By PN
      c.execute(f"SELECT r.pn_sf, SUM(d.duration_min) FROM downtime_logs d JOIN records r ON d.sub_batch_id = r.sub_batch_id WHERE d.created_at >= ? AND d.created_at <= ?{dt_shift_cond} GROUP BY r.pn_sf ORDER BY SUM(d.duration_min) DESC LIMIT 5", dt_params)
      pn_dt_data = c.fetchall()
      if pn_dt_data:
        x_lbls = [row[0][:20] for row in pn_dt_data]
        y_vals = [row[1] for row in pn_dt_data]
        self.ax_dt_pn.bar(x_lbls, y_vals, color=ERROR_COLOR, width=0.4)
        self.ax_dt_pn.set_xlim(-0.5, max(1.5, len(x_lbls)-0.5))
        self.ax_dt_pn.tick_params(axis='x', rotation=20, labelsize=8)
      else:
        self.ax_dt_pn.text(0.5, 0.5, "No Data", ha='center', va='center', color=TEXT_MUTED, transform=self.ax_dt_pn.transAxes)
      self.fig_dt_pn.subplots_adjust(bottom=0.25, top=0.85, right=0.95, left=0.15)
      self.render_chart_to_label(self.fig_dt_pn, self.lbl_chart_dt_pn)
      
      # 8D. By Operator
      c.execute(f"SELECT r.op_id, SUM(d.duration_min) FROM downtime_logs d JOIN records r ON d.sub_batch_id = r.sub_batch_id WHERE d.created_at >= ? AND d.created_at <= ?{dt_shift_cond} GROUP BY r.op_id ORDER BY SUM(d.duration_min) DESC LIMIT 5", dt_params)
      op_dt_data = c.fetchall()
      if op_dt_data:
        x_lbls = [row[0] for row in op_dt_data]
        y_vals = [row[1] for row in op_dt_data]
        self.ax_dt_op.bar(x_lbls, y_vals, color=ACCENT_COLOR, width=0.4)
        self.ax_dt_op.set_xlim(-0.5, max(1.5, len(x_lbls)-0.5))
        self.ax_dt_op.tick_params(axis='x', rotation=20, labelsize=8)
      else:
        self.ax_dt_op.text(0.5, 0.5, "No Data", ha='center', va='center', color=TEXT_MUTED, transform=self.ax_dt_op.transAxes)
      self.fig_dt_op.subplots_adjust(bottom=0.25, top=0.85, right=0.95, left=0.15)
      self.render_chart_to_label(self.fig_dt_op, self.lbl_chart_dt_op)

      conn.close()
    except Exception as e:
      print("KPI Refresh Error:", e)

  def export_kpis_to_excel(self):
    try:
      conn = get_db_connection()
      c = conn.cursor()
      
      # Production day logic (Calendar Day)
      prod_date_sql = "substr(dt_sp, 1, 10)"
      
      c.execute(f"SELECT DISTINCT {prod_date_sql} as prod_date FROM records ORDER BY prod_date DESC")
      dates = [r[0] for r in c.fetchall() if r[0]]
      
      c.execute("SELECT SUM(quantity) FROM records")
      grand_total = c.fetchone()[0] or 0

      if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Advanced KPI Reports"
      else:
        wb = load_workbook(EXCEL_FILE)
        if "Advanced KPI Reports" in wb.sheetnames:
          del wb["Advanced KPI Reports"]
        ws = wb.create_sheet("Advanced KPI Reports", 0)
          
        # Clean up old sheet if exists
        if "KPI Reports" in wb.sheetnames:
          del wb["KPI Reports"]
      
      # Define Styles
      header_font = Font(name='Segoe UI', size=16, bold=True, color="FFFFFF")
      header_fill = PatternFill(start_color="1B232C", end_color="1B232C", fill_type="solid")
      
      date_font = Font(name='Segoe UI', size=14, bold=True, color="FFFFFF")
      date_fill = PatternFill(start_color="0078D7", end_color="0078D7", fill_type="solid")
      
      sub_header_font = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
      sub_header_fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
      
      data_font = Font(name='Segoe UI', size=11, color="000000")
      bold_data_font = Font(name='Segoe UI', size=11, bold=True, color="000000")
      
      align_center = Alignment(horizontal="center", vertical="center")
      thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
      
      # Main Header
      ws.merge_cells('A1:N2')
      top_cell = ws.cell(row=1, column=1, value=f"KPI REPORT | GENERATED: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} | TOTAL ALL-TIME PRODUCTION: {grand_total}")
      top_cell.font = header_font
      top_cell.fill = header_fill
      top_cell.alignment = align_center
      
      current_row = 4
      
      for d in dates:
        c.execute(f"SELECT SUM(quantity) FROM records WHERE {prod_date_sql} = ?", (d,))
        d_total = c.fetchone()[0] or 0
        
        c.execute(f"SELECT shift_sp, SUM(quantity) FROM records WHERE {prod_date_sql} = ? GROUP BY shift_sp ORDER BY shift_sp", (d,))
        shifts = c.fetchall()
        
        c.execute(f"SELECT op_id, SUM(quantity) as q FROM records WHERE {prod_date_sql} = ? GROUP BY op_id ORDER BY q DESC LIMIT 5", (d,))
        ops = c.fetchall()
        
        c.execute(f"SELECT pn_sf, SUM(quantity) as q FROM records WHERE {prod_date_sql} = ? GROUP BY pn_sf ORDER BY q DESC LIMIT 5", (d,))
        pns = c.fetchall()
        
        c.execute(f"SELECT station, SUM(quantity) as q FROM records WHERE {prod_date_sql} = ? GROUP BY station ORDER BY q DESC LIMIT 5", (d,))
        stations = c.fetchall()
        
        c.execute(f"SELECT op_id, GROUP_CONCAT(DISTINCT reason), SUM(duration_min) as m FROM downtime_logs WHERE substr(created_at, 1, 10) = ? GROUP BY op_id ORDER BY m DESC LIMIT 5", (d,))
        downtimes = c.fetchall()
        
        # Date Header
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=14)
        d_cell = ws.cell(row=current_row, column=1, value=f"PRODUCTION DATE: {d}  |  TOTAL DAILY OUTPUT: {d_total}")
        d_cell.font = date_font
        d_cell.fill = date_fill
        d_cell.alignment = align_center
        current_row += 1
        
        # Grid Headers
        headers = [
          (1, "SHIFT", "OUTPUT"),
          (4, "TOP 5 OPERATORS", "OUTPUT"),
          (7, "TOP 5 PRODUCTS", "OUTPUT"),
          (10, "STATIONS", "OUTPUT"),
          (13, "DOWNTIME (OP)", "MINS")
        ]
        
        for col_start, t1, t2 in headers:
          ws.cell(row=current_row, column=col_start, value=t1).font = sub_header_font
          ws.cell(row=current_row, column=col_start).fill = sub_header_fill
          ws.cell(row=current_row, column=col_start).alignment = align_center
          ws.cell(row=current_row, column=col_start).border = thin_border
          
          ws.cell(row=current_row, column=col_start+1, value=t2).font = sub_header_font
          ws.cell(row=current_row, column=col_start+1).fill = sub_header_fill
          ws.cell(row=current_row, column=col_start+1).alignment = align_center
          ws.cell(row=current_row, column=col_start+1).border = thin_border
        
        current_row += 1
        
        # Grid Data (5 rows)
        start_data_row = current_row
        for i in range(5):
          # Shift
          if i < len(shifts):
            ws.cell(row=start_data_row+i, column=1, value=f"Shift {shifts[i][0]}").font = data_font
            ws.cell(row=start_data_row+i, column=2, value=shifts[i][1]).font = bold_data_font
          else:
            ws.cell(row=start_data_row+i, column=1, value="").font = data_font
            ws.cell(row=start_data_row+i, column=2, value="").font = data_font
            
          # Operator
          if i < len(ops):
            ws.cell(row=start_data_row+i, column=4, value=ops[i][0]).font = data_font
            ws.cell(row=start_data_row+i, column=5, value=ops[i][1]).font = bold_data_font
          else:
            ws.cell(row=start_data_row+i, column=4, value="")
            ws.cell(row=start_data_row+i, column=5, value="")
            
          # Product
          if i < len(pns):
            ws.cell(row=start_data_row+i, column=7, value=pns[i][0]).font = data_font
            ws.cell(row=start_data_row+i, column=8, value=pns[i][1]).font = bold_data_font
          else:
            ws.cell(row=start_data_row+i, column=7, value="")
            ws.cell(row=start_data_row+i, column=8, value="")
            
          # Station
          if i < len(stations):
            ws.cell(row=start_data_row+i, column=10, value=stations[i][0]).font = data_font
            ws.cell(row=start_data_row+i, column=11, value=stations[i][1]).font = bold_data_font
          else:
            ws.cell(row=start_data_row+i, column=10, value="")
            ws.cell(row=start_data_row+i, column=11, value="")
            
          # Downtime
          if i < len(downtimes):
            op_id_val = downtimes[i][0] or '-'
            reasons_val = downtimes[i][1][:20] if downtimes[i][1] else '-'
            reason_txt = f"{op_id_val} ({reasons_val})"
            ws.cell(row=start_data_row+i, column=13, value=reason_txt).font = data_font
            ws.cell(row=start_data_row+i, column=14, value=downtimes[i][2]).font = bold_data_font
          else:
            ws.cell(row=start_data_row+i, column=13, value="")
            ws.cell(row=start_data_row+i, column=14, value="")
            
          # Apply Borders
          for c_idx in [1,2,4,5,7,8,10,11,13,14]:
            ws.cell(row=start_data_row+i, column=c_idx).border = thin_border
            ws.cell(row=start_data_row+i, column=c_idx).alignment = align_center

        current_row += 6

      ws.column_dimensions['A'].width = 15
      ws.column_dimensions['B'].width = 15
      ws.column_dimensions['C'].width = 3
      ws.column_dimensions['D'].width = 25
      ws.column_dimensions['E'].width = 15
      ws.column_dimensions['F'].width = 3
      ws.column_dimensions['G'].width = 35
      ws.column_dimensions['H'].width = 15
      ws.column_dimensions['I'].width = 3
      ws.column_dimensions['J'].width = 15
      ws.column_dimensions['K'].width = 15
      ws.column_dimensions['L'].width = 3
      ws.column_dimensions['M'].width = 30
      ws.column_dimensions['N'].width = 15
      
      ws.sheet_view.showGridLines = False

      conn.close()
      wb.save(EXCEL_FILE)
    except Exception as e:
      print(f"Background KPI Export Error: {e}")

  def open_excel(self):
    if os.path.exists(EXCEL_FILE):
      os.startfile(EXCEL_FILE)
    else:
      messagebox.showinfo("Info", "Excel file not created yet. Save a record first.")

  def take_snapshot(self, date_str):
    conn = get_db_connection()
    c = conn.cursor()
    
    c.execute('''SELECT pn_sf, SUM(quantity), COUNT(id)
           FROM records WHERE status="In Rack" GROUP BY pn_sf''')
    agg_rows = c.fetchall()
    
    c.execute('''SELECT pn_sf, dt_sp FROM records WHERE status="In Rack"''')
    det_rows = c.fetchall()
    
    now = datetime.datetime.now()
    oldest_dt_per_pn = {}
    for row in det_rows:
      pn, dt_str = row
      try:
        try:
          dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
          dt_obj = datetime.datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
        if pn not in oldest_dt_per_pn or dt_obj < oldest_dt_per_pn[pn]:
          oldest_dt_per_pn[pn] = dt_obj
      except: pass
      
    if not agg_rows:
      c.execute('''INSERT OR REPLACE INTO inventory_snapshots 
             (snapshot_date, pn_sf, boxes_in_rack, total_qty_in_rack, oldest_box_age_hours)
             VALUES (?, ?, ?, ?, ?)''', 
             (date_str, "NONE", 0, 0, 0.0))
    else:
      for row in agg_rows:
        pn, total_qty, box_count = row
        oldest_age_hours = 0.0
        if pn in oldest_dt_per_pn:
          oldest_age_hours = (now - oldest_dt_per_pn[pn]).total_seconds() / 3600.0
          
        c.execute('''INSERT OR REPLACE INTO inventory_snapshots 
               (snapshot_date, pn_sf, boxes_in_rack, total_qty_in_rack, oldest_box_age_hours)
               VALUES (?, ?, ?, ?, ?)''', 
               (date_str, pn, box_count, total_qty, oldest_age_hours))
    conn.commit()
    conn.close()
    print(f"[{datetime.datetime.now()}] Automated Inventory Snapshot Taken for {date_str}.")
    
  def schedule_daily_snapshot(self):
    now = datetime.datetime.now()
    yesterday = now - datetime.timedelta(days=1)
    yesterday_str = yesterday.strftime("%Y-%m-%d")
    
    try:
      conn = get_db_connection()
      c = conn.cursor()
      c.execute("SELECT COUNT(*) FROM inventory_snapshots WHERE snapshot_date=?", (yesterday_str,))
      count = c.fetchone()[0]
      conn.close()
      
      if count == 0:
        self.take_snapshot(yesterday_str)
    except Exception as e:
      print("Error checking/taking daily snapshot:", e)

    self.after(60000, self.schedule_daily_snapshot)

  def on_closing(self):
    msg = "Are you sure you want to close the dashboard?\n\n(Automated Reports & Snapshots will continue running in the background)"
    if messagebox.askyesno("Confirm Exit", msg):
      self.withdraw_to_tray()
  def withdraw_to_tray(self):
    self.withdraw()
    try:
      import pystray
      from PIL import Image
      import threading

      def show_window(icon, item):
        icon.stop()
        self.after(0, self.deiconify)

      def quit_window(icon, item):
        icon.stop()
        self.after(0, self.destroy)

      menu = pystray.Menu(
        pystray.MenuItem('Restore Dashboard', show_window, default=True),
        pystray.MenuItem('Exit Completely', quit_window)
      )

      icon_path = resource_path(os.path.join("assets", "taskbar_logo.png"))
      if os.path.exists(icon_path):
        image = Image.open(icon_path)
      else:
        image = Image.new('RGB', (64, 64), color='red')

      icon = pystray.Icon("traceability", image, "Sub-Process Traceability", menu)
      threading.Thread(target=icon.run, daemon=True).start()

    except ImportError:
      self.destroy()

if __name__ == "__main__":
  init_db()
  app = TraceabilityApp()
  app.mainloop()
