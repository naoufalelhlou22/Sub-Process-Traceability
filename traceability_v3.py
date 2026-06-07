import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import sqlite3
import datetime
import os
import webbrowser
import tempfile
import json
import base64
import io
import qrcode
import cv2
import numpy as np
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
import sys
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

try:
    import win32print
    WIN32_PRINT_AVAILABLE = True
except ImportError:
    WIN32_PRINT_AVAILABLE = False

CONFIG_FILE = "traceability_config.json"
# Premium Industrial HMI Theme

BG_COLOR = "#0F1419"       # Main Background
SURFACE_COLOR = "#1B232C"  # Panels & Cards

ACCENT_COLOR = "#00B4D8"   # Primary Action
SUCCESS_COLOR = "#2DC653"  # Running / Completed
WARNING_COLOR = "#F59E0B"  # Warning
ERROR_COLOR = "#DC2626"    # Error / Stop

BORDER_COLOR = "#334155"   # Borders

TEXT_COLOR = "#F8FAFC"     # Main Text
TEXT_MUTED = "#94A3B8"     # Secondary Text

# Status Colors
STATUS_RUNNING = "#22C55E"
STATUS_IDLE = "#64748B"
STATUS_WARNING = "#F59E0B"
STATUS_STOPPED = "#EF4444"

# Modern HMI Fonts
HMI_FONT_XL = ("Segoe UI", 20, "bold")
HMI_FONT_L = ("Segoe UI", 16, "bold")
HMI_FONT_M = ("Segoe UI", 12, "bold")
HMI_FONT_S = ("Segoe UI", 10)
HMI_FONT_XS = ("Segoe UI", 9)

# Hardcoded Data
SF_DATA = {
    "AT-V07-11313-A-SUB": ("SUB BODY N R", [("AT-V07-11313-P-WIE", "BODY R"), ("AT-V07-11323-E-WIE", "BUMPER RUBBER R")]),
    "AT-V07-11213-A-SUB": ("SUB BODY EMG R", [("AT-V07-11213-E-WIE", "BODY(EMERGENCY) R"), ("AT-V07-11323-E-WIE", "BUMPER RUBBER R")]),
    "AT-V07-11314-A-SUB": ("SUB BODY N L", [("AT-V07-11314-M-WIE", "BODY L"), ("AT-V07-11324-E-WIE", "BUMPER RUBBER L")]),
    "AT-V07-11214-A-SUB": ("SUB BODY EMG L", [("AT-V07-11214-E-WIE", "BODY(EMERGENCY) L"), ("AT-V07-11324-E-WIE", "BUMPER RUBBER L")]),
    "AT-V80-11213-A-SUB": ("SUB BODY EMG R XDD", [("AT-V07-11213-E-WIE", "BODY(EMERGENCY) R"), ("AT-V07-11223-E-WIE", "CATCH BUMPER R"), ("AT-V07-11323-E-WIE", "BUMPER RUBBER R")]),
    "AT-V80-11314-A-SUB": ("SUB BODY N L XDD", [("AT-V07-11314-M-WIE", "BODY L"), ("AT-V07-11224-E-WIE", "CATCH BUMPER L"), ("AT-V07-11324-E-WIE", "BUMPER RUBBER L")]),
    "AT-V07-11311-A-SUB": ("Back plate R sub", [("AT-V07-11311-J-WIE", "BACK PLATE R"), ("AT-V07-11359-B-WIE", "STOPPER RUBBER")]),
    "AT-V07-11312-A-SUB": ("Back plate L sub", [("AT-V07-11312-J-WIE", "BACK PLATE L"), ("AT-V07-11359-B-WIE", "STOPPER RUBBER")]),
    "AT-V80-11311-A-SUB": ("Back plate R sub XDD", [("AT-V80-11311-C-WIE", "BACK PLATE (ZnNi) R"), ("AT-V07-11359-B-WIE", "STOPPER RUBBER")]),
    "AT-V80-11312-A-SUB": ("Back plate L sub XDD", [("AT-V80-11312-C-WIE", "BACK PLATE (ZnNi) L"), ("AT-V07-11359-B-WIE", "STOPPER RUBBER")]),
    "AT-V07-14311-A-SUB": ("Lever Child R sub (short type)", [("AT-V07-14311-F-WIE", "LEVER CHILD R"), ("AT-V07-14315-C-WIE", "SPRING CHILD R"), ("AT-V07-14319-B-WIE", "PIN CHILD")]),
    "AT-V07-14341-A-SUB": ("Lever Child R sub (long type)", [("AT-V07-14341-A-WIE", "LEVER CHILD LONG TYPE R"), ("AT-V07-14315-C-WIE", "SPRING CHILD R"), ("AT-V07-14319-B-WIE", "PIN CHILD")]),
    "AT-V07-14312-A-SUB": ("Lever Child L sub (short type)", [("AT-V07-14312-F-WIE", "LEVER CHILD L"), ("AT-V07-14316-C-WIE", "SPRING CHILD L"), ("AT-V07-14319-B-WIE", "PIN CHILD")]),
    "AT-V07-14342-A-SUB": ("Lever Child L sub (long type)", [("AT-V07-14342-A-WIE", "LEVER CHILD LONG TYPE L"), ("AT-V07-14316-C-WIE", "SPRING CHILD L"), ("AT-V07-14319-B-WIE", "PIN CHILD")]),
    "AT-T61-11313-A-SUB": ("Access Key R sub", [("AT-T61-11313-C-TIE", "ACCESS KEY R"), ("AT-D93-11349-A-TIE", "X-RING")]),
    "AT-V07-11315-A-SUB": ("COVER PLATE SUB R", [("AT-V07-11315-E-WIE", "COVER PLATE R"), ("AT-V07-11319-B-WIE", "SHAFT LATCH")]),
    "AT-V07-11316-A-SUB": ("COVER PLATE SUB L", [("AT-V07-11316-E-WIE", "COVER PLATE L"), ("AT-V07-11319-B-WIE", "SHAFT LATCH")]),
    "AT-V80-11315-A-SUB": ("COVER PLATE SUB R XDD", [("AT-V80-11315-B-WIE", "COVER PLATE (ZnNi) R"), ("AT-V80-11319-B-WIE", "SHAFT LATCH (ZnNi)"), ("AT-V07-11319-B-WIE", "SHAFT LATCH")]),
    "AT-V80-11319-A-SUB": ("COVER PLATE SUB L XDD", [("AT-V80-11319-B-WIE", "SHAFT LATCH (ZnNi)"), ("AT-V07-11319-B-WIE", "SHAFT LATCH")]),
    "AT-V07-31351-A-SUB": ("Motor Terminal NL R", [("AT-W22-31311-B-WIE", "MOTOR"), ("AT-V07-31351-C-WIE", "MOTOR TERMINAL R"), ("AT-V07-31311-A-WIE", "WORM")]),
    "AT-V07-31352-A-SUB": ("Motor Terminal NL L", [("AT-W22-31311-B-WIE", "MOTOR"), ("AT-V07-31352-C-WIE", "MOTOR TERMINAL L"), ("AT-V07-31311-A-WIE", "WORM")]),
    "AT-V90-12323-A-SUB": ("Sub Assy Motor SL R", [("AT-W22-31311-B-WIE", "MOTOR"), ("AT-W22-31312-B-WIE", "MOTOR"), ("AT-V90-12323-B-WIE", "MOTOR TERMINAL SL 3P R"), ("AT-V07-31311-A-WIE", "WORM")]),
    "AT-V90-12328-A-SUB": ("Sub Assy Motor SL L", [("AT-W22-31311-B-WIE", "MOTOR"), ("AT-W22-31312-B-WIE", "MOTOR"), ("AT-V90-12328-A-WIE", "MOTOR TERMINAL SL-A L"), ("AT-V90-12332-A-WIE", "MOTOR TERMINAL SL-B L"), ("AT-V07-31311-A-WIE", "WORM")]),
    "AT-E71-12331-A-SUB": ("HOLDER-ROD RH", [("AT-E71-12331-A-WIE", "HOLDER-ROD RH"), ("AT-E71-31355-C-WIE", "KEY LEVER B")]),
    "AT-E71-12332-A-SUB": ("HOLDER-ROD LH", [("AT-E71-12332-A-WIE", "HOLDER-ROD LH"), ("AT-E71-31355-C-WIE", "KEY LEVER B")]),
    "AT-YA3-11332-E-SUB": ("LEVER IH SUB L", [("AT-YA3-11332-E-TIE", "LEVER IH OR L"), ("AT-000-00265-F-TIE", "SNAP L")]),
    "AT-YA3-11331-E-SUB": ("LEVER IH SUB R", [("AT-YA3-11331-E-TIE", "LEVER IH OR R"), ("AT-000-00264-F-TIE", "SNAP R")]),
    "AT-YA3-11314-K-SUB": ("BODY SUB L", [("AT-YA3-11376-A-WIE", "CATCH BUMPER L"), ("AT-YA3-11314-H-TIE", "BODY L"), ("AT-V07-11324-E-WIE", "BUMPER RUBBER L"), ("AT-L88-11319-E-TIE", "STOPPER RUBBER")]),
    "AT-YA3-11316-K-SUB": ("BODY SUB (EMERGENCY) L", [("AT-YA3-11376-A-WIE", "CATCH BUMPER L"), ("AT-YA3-11316-K-TIE", "BODY(EMERGENCY) L"), ("AT-V07-11324-E-WIE", "BUMPER RUBBER L"), ("AT-L88-11319-E-TIE", "STOPPER RUBBER")]),
    "AT-YA3-11315-K-SUB": ("BODY SUB (EMERGENCY) R", [("AT-YA3-11375-A-WIE", "CATCH BUMPER R"), ("AT-YA3-11315-K-TIE", "BODY(EMERGENCY) R"), ("AT-V07-11323-E-WIE", "BUMPER RUBBER R"), ("AT-L88-11319-E-TIE", "STOPPER RUBBER")])
}

DB_FILE = "traceability.db"
EXCEL_FILE = "production_data.xlsx"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sub_batch_id TEXT NOT NULL,
            pn_sf TEXT,
            part_sf TEXT,
            rm1_pn TEXT,
            rm1_name TEXT,
            rm2_pn TEXT,
            rm2_name TEXT,
            rm3_pn TEXT,
            rm3_name TEXT,
            rm4_pn TEXT,
            rm4_name TEXT,
            batch1 TEXT,
            batch2 TEXT,
            batch3 TEXT,
            quantity INTEGER,
            shift_sp TEXT,
            op_id TEXT,
            station TEXT,
            dt_sp TEXT,
            dt_line TEXT,
            shift_line TEXT,
            remarks TEXT,
            status TEXT DEFAULT 'pending',
            created_at TEXT NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

def db_query(query, args=(), fetch=True):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(query, args)
    result = None
    if fetch:
        result = c.fetchall()
    else:
        conn.commit()
    conn.close()
    return result

def get_next_sequence():
    today_str = datetime.datetime.now().strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM records WHERE date(created_at) = ?", (today_str,))
    count = c.fetchone()[0]
    conn.close()
    return count + 1

def generate_sub_batch_id(shift, station):
    today_str = datetime.datetime.now().strftime("%Y%m%d")
    seq = get_next_sequence()
    st_num = station.replace("S", "") if station else "00"
    if not shift:
        shift = "X"
    return f"SB-{today_str}-{shift}-{st_num}-{seq:03d}"

def save_to_excel(data):
    headers = [
        "FULL PN° Semi fini", "PART NAME (SF)", 
        "RM PN", "RM Name",
        "Batch No. 1", "Batch No. 2", "Batch No. 3", "Quantity",
        "Work in Sub-Process by Shift", "Op ID", "Station",
        "Sub-Process Work Date/Time", "Production Line Entry Date/Time",
        "Work in PROD line by Shift", "Remarks"
    ]
    
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sub-process fill by TL"
        ws.append(headers)
        
        header_font = Font(name='Calibri', size=10, bold=True, color="000000")
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for col in range(1, 16):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 15
        
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 34
    else:
        wb = load_workbook(EXCEL_FILE)
        if "Sub-process fill by TL" in wb.sheetnames:
            ws = wb["Sub-process fill by TL"]
        else:
            ws = wb.create_sheet("Sub-process fill by TL")
            ws.append(headers)
    
    sf_pn = data[0] if data[0] else "-"
    part_sf = data[1] if data[1] else "-"
    rms = []
    for i in range(4):
        rm_pn = data[2 + i*2]
        rm_name = data[3 + i*2]
        if rm_pn:
            rms.append((rm_pn if rm_pn else "-", rm_name if rm_name else "-"))
            
    if not rms:
        rms = [("-", "-")]
        
    num_rms = len(rms)
    rest_data = ["-" if (x == "" or x is None) else x for x in data[10:]]
    
    start_row = ws.max_row + 1
    end_row = start_row + num_rms - 1
    
    for i in range(num_rms):
        row_data = []
        if i == 0:
            row_data.extend([sf_pn, part_sf])
            row_data.extend([rms[i][0], rms[i][1]])
            row_data.extend(rest_data)
        else:
            row_data.extend([None, None])
            row_data.extend([rms[i][0], rms[i][1]])
            row_data.extend([None] * len(rest_data))
        ws.append(row_data)
        
    if num_rms > 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        for col in range(5, 16):
            ws.merge_cells(start_row=start_row, start_column=col, end_row=end_row, end_column=col)
    
    sf_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    rm_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    rest_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    row_font = Font(name='Calibri', size=10, color="000000")
    align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for r in range(start_row, end_row + 1):
        for c in range(1, 16):
            cell = ws.cell(row=r, column=c)
            cell.font = row_font
            cell.alignment = align
            cell.border = thin_border
            if c in (1, 2):
                cell.fill = sf_fill
            elif c in (3, 4):
                cell.fill = rm_fill
            else:
                cell.fill = rest_fill
            
    wb.save(EXCEL_FILE)

def print_html_slip(record_data):
    config = {}
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r") as f:
                config = json.load(f)
        except: pass
            
    printer_name = config.get("zebra_printer", "")
    qr_payload = {
        "full_pn_sf": record_data.get('pn_sf', ''),
        "part_name_sf": record_data.get('part_sf', ''),
        "rm1_pn": record_data.get('rm1_pn', ''),
        "rm1_name": record_data.get('rm1_name', ''),
        "rm2_pn": record_data.get('rm2_pn', ''),
        "rm2_name": record_data.get('rm2_name', ''),
        "rm3_pn": record_data.get('rm3_pn', ''),
        "rm3_name": record_data.get('rm3_name', ''),
        "rm4_pn": record_data.get('rm4_pn', ''),
        "rm4_name": record_data.get('rm4_name', ''),
        "batch_no_1": record_data.get('batch1', ''),
        "batch_no_2": record_data.get('batch2', ''),
        "batch_no_3": record_data.get('batch3', ''),
        "quantity": record_data.get('quantity', 0),
        "sub_process_shift": record_data.get('shift_sp', ''),
        "op_id": record_data.get('op_id', ''),
        "station": record_data.get('station', ''),
        "sub_process_datetime": record_data.get('dt_sp', ''),
        "production_line_entry_datetime": record_data.get('dt_line', ''),
        "production_line_shift": record_data.get('shift_line', ''),
        "remarks": record_data.get('remarks', '')
    }
    json_str = json.dumps(qr_payload)
    
    def format_val(val, default="-"):
        return str(val) if val else default

    zpl = []
    zpl.append("^XA")
    zpl.append("^PW480")
    zpl.append("^CI28")
    
    zpl.append("^FO80,20^A0N,24,24^FDSUB-PROCESS RECORD^FS")
    zpl.append("^FO10,55^GB460,3,3^FS")
    
    zpl.append("^FO10,70^GB460,40,40^FS")
    zpl.append(f"^FO10,80^A0N,24,24^FR^FB460,1,0,C^FD{format_val(record_data.get('sub_batch_id'))}\\&^FS")
    
    y = 120
    def add_row(label, value, font_size=20, y_inc=25):
        nonlocal y
        zpl.append(f"^FO10,{y}^A0N,{font_size},{font_size}^FD{label}^FS")
        zpl.append(f"^FO10,{y}^A0N,{font_size},{font_size}^FB460,1,0,R^FD{value}\\&^FS")
        y += y_inc

    add_row("PN Semi fini", format_val(record_data.get('pn_sf')))
    add_row("Part Name (SF)", format_val(record_data.get('part_sf')))
    
    for i in range(1, 5):
        pn_key = f"rm{i}_pn"
        name_key = f"rm{i}_name"
        if record_data.get(pn_key):
            lbl_suffix = "" if i == 1 else str(i)
            add_row(f"PN RM{lbl_suffix} Ref", format_val(record_data.get(pn_key)))
            add_row(f"Part Name (RM{lbl_suffix})", format_val(record_data.get(name_key)))
            
    y += 5
    zpl.append(f"^FO10,{y}^GB460,1,1^FS")
    y += 12
    
    add_row("Batch No. 1", format_val(record_data.get('batch1')))
    add_row("Batch No. 2", format_val(record_data.get('batch2')))
    add_row("Batch No. 3", format_val(record_data.get('batch3')))
    add_row("Quantity", f"{format_val(record_data.get('quantity'))} pcs", font_size=22)
    
    y += 5
    zpl.append(f"^FO10,{y}^GB460,1,1^FS")
    y += 12
    
    add_row("Shift SP", format_val(record_data.get('shift_sp')))
    add_row("Op ID", format_val(record_data.get('op_id')))
    add_row("Station", format_val(record_data.get('station')))
    dt_sp = format_val(record_data.get('dt_sp'))[:16] if record_data.get('dt_sp') else '-'
    add_row("SP Date/Time", dt_sp)
    dt_line = format_val(record_data.get('dt_line'))[:16] if record_data.get('dt_line') else '-'
    add_row("Line Entry", dt_line)
    
    y += 5
    zpl.append(f"^FO10,{y}^GB460,2,2^FS")
    y += 12
    
    status = format_val(record_data.get('status', '-'))
    printed = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    zpl.append(f"^FO10,{y}^A0N,18,18^FB460,1,0,C^FDStatus: {status} | Printed: {printed}\\&^FS")
    y += 20
    
    zpl.append(f"^FO130,{y}^BQN,2,2^FDQA,{json_str}^FS")
    zpl.append("^PQ2")
    zpl.append("^XZ")
    zpl_string = "\n".join(zpl).encode("utf-8")
    
    # Save the ZPL code to a file for preview/debugging
    try:
        os.makedirs("qr", exist_ok=True)
        sb_id_safe = format_val(record_data.get('sub_batch_id', 'unknown')).replace(":", "").replace("-", "")
        file_path = os.path.join("qr", f"{sb_id_safe}.zpl")
        with open(file_path, "wb") as f:
            f.write(zpl_string)
    except Exception as e:
        print("Could not save ZPL file:", e)
        
    if not printer_name:
        messagebox.showwarning("Setup Required", "ZPL saved to 'qr' folder. Please configure Zebra Printer in Settings to physically print.")
        return
        
    if not WIN32_PRINT_AVAILABLE:
        messagebox.showerror("Error", "pywin32 is not installed. ZPL saved to 'qr' folder, but cannot print. Please run 'pip install pywin32'.")
        return
    
    try:
        hPrinter = win32print.OpenPrinter(printer_name)
        try:
            job = win32print.StartDocPrinter(hPrinter, 1, ("Sub-Process Label", None, "RAW"))
            win32print.StartPagePrinter(hPrinter)
            win32print.WritePrinter(hPrinter, zpl_string)
            win32print.EndPagePrinter(hPrinter)
            win32print.EndDocPrinter(hPrinter)
            messagebox.showinfo("Success", f"Label sent to printer '{printer_name}' successfully!")
        finally:
            win32print.ClosePrinter(hPrinter)
    except Exception as e:
        messagebox.showerror("Printer Error", f"Failed to send to printer '{printer_name}':\\n{e}")

class TraceabilityApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("HI-LEX ACT - Sub-Process Traceability")
        self.geometry("1280x800")
        try:
            taskbar_logo = resource_path(os.path.join("assets", "taskbar_logo.png"))
            if os.path.exists(taskbar_logo):
                self.iconphoto(True, tk.PhotoImage(file=taskbar_logo))
        except: pass
        try:
            self.state('zoomed')
        except:
            pass
        self.configure(bg=BG_COLOR)
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # Modern Windows 11 Title Bar Color (White background, Blue text)
        try:
            import ctypes
            self.update_idletasks()
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
            # DWMWA_CAPTION_COLOR = 35
            ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 35, ctypes.byref(ctypes.c_int(0x00FFFFFF)), 4)
            # DWMWA_TEXT_COLOR = 36 (HI-LEX Blue #005A8C -> 0x008C5A00)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(hwnd, 36, ctypes.byref(ctypes.c_int(0x008C5A00)), 4)
        except Exception:
            pass
        
        self.option_add('*insertBackground', 'white')
        self.option_add('*Listbox.background', SURFACE_COLOR)
        self.option_add('*Listbox.foreground', TEXT_COLOR)
        self.option_add('*Listbox.selectBackground', ACCENT_COLOR)
        self.option_add('*Listbox.selectForeground', 'black')
        self.option_add('*TCombobox*Listbox.background', SURFACE_COLOR)
        self.option_add('*TCombobox*Listbox.foreground', TEXT_COLOR)
        self.option_add('*TCombobox*Listbox.selectBackground', ACCENT_COLOR)
        self.option_add('*TCombobox*Listbox.selectForeground', 'black')
        
        # Configure Styles
        self.style = ttk.Style(self)
        self.style.theme_use('clam')
        self.style.configure('.', background=BG_COLOR, foreground=TEXT_COLOR, fieldbackground=SURFACE_COLOR)
        self.style.configure('TNotebook', background=BG_COLOR)
        self.style.configure('TNotebook.Tab', background=SURFACE_COLOR, foreground=TEXT_COLOR, padding=[20, 8], font=("Segoe UI", 11, "bold"))
        self.style.map('TNotebook.Tab', background=[('selected', ACCENT_COLOR)])
        self.style.configure('TFrame', background=BG_COLOR)
        self.style.configure('TLabel', background=SURFACE_COLOR, foreground=TEXT_COLOR)
        self.style.configure('TCheckbutton', background=SURFACE_COLOR, foreground=TEXT_COLOR)
        
        self.style.configure('TCombobox', fieldbackground=SURFACE_COLOR, background=SURFACE_COLOR, foreground=TEXT_COLOR)
        self.style.map('TCombobox', 
                       fieldbackground=[('readonly', SURFACE_COLOR)], 
                       selectbackground=[('readonly', ACCENT_COLOR)],
                       selectforeground=[('readonly', '#000000')])
        
        self.style.configure('TEntry', fieldbackground=SURFACE_COLOR, foreground=TEXT_COLOR, insertcolor='white')
        self.style.configure('TButton', background=SURFACE_COLOR, foreground=TEXT_COLOR, font=HMI_FONT_M, padding=5)
        self.style.map('TButton', background=[('active', BORDER_COLOR)])
        
        self.style.configure('Success.TButton', background=SUCCESS_COLOR, foreground="#000000", font=HMI_FONT_M, padding=5)
        self.style.map('Success.TButton', background=[('active', STATUS_RUNNING)])
        
        self.style.configure('Primary.TButton', background=ACCENT_COLOR, foreground="#000000", font=HMI_FONT_M, padding=5)
        self.style.map('Primary.TButton', background=[('active', "#0096C7")])
        
        self.style.configure('Secondary.TButton', background=STATUS_IDLE, foreground="#FFFFFF", font=HMI_FONT_M, padding=5)
        self.style.map('Secondary.TButton', background=[('active', BORDER_COLOR)])
        
        self.style.configure('Header.TButton', background="#005A8C", foreground="#FFFFFF", font=HMI_FONT_M, padding=5)
        self.style.map('Header.TButton', background=[('active', "#00456B")])
        
        self.style.configure('Warning.TButton', background=WARNING_COLOR, foreground="#000000", font=HMI_FONT_M, padding=5)
        self.style.map('Warning.TButton', background=[('active', "#D97706")])
        
        self.style.configure('Danger.TButton', background=ERROR_COLOR, foreground="#FFFFFF", font=HMI_FONT_M, padding=5)
        self.style.map('Danger.TButton', background=[('active', "#B91C1C")])
        
        self.style.configure('Treeview', background=SURFACE_COLOR, foreground=TEXT_COLOR, fieldbackground=SURFACE_COLOR, rowheight=30, font=HMI_FONT_S)
        self.style.configure('Treeview.Heading', background=BORDER_COLOR, foreground=TEXT_COLOR, font=HMI_FONT_M)
        self.style.map('Treeview.Heading',
                       background=[('active', ACCENT_COLOR)],
                       foreground=[('active', '#000000')])
        
        self.build_ui()
        self.update_clock()
        self.update_stats()
        self.populate_sf_combobox()
        self.refresh_recent_treeview()
        self.refresh_records_treeview()
        self.update_sub_batch_preview()

    def create_card(self, parent, title, fg_color=ACCENT_COLOR):
        card = tk.Frame(parent, bg=SURFACE_COLOR)
        card.pack(fill=tk.X, pady=5, padx=5)
        
        title_lbl = tk.Label(card, text=title, bg=SURFACE_COLOR, fg=fg_color, font=HMI_FONT_M, anchor="w", padx=10, pady=5)
        title_lbl.pack(fill=tk.X)
        
        tk.Frame(card, bg=BORDER_COLOR, height=1).pack(fill=tk.X)
        
        content = tk.Frame(card, bg=SURFACE_COLOR, padx=10, pady=10)
        content.pack(fill=tk.BOTH, expand=True)
        return card, content

    def open_settings(self):
        top = tk.Toplevel(self)
        top.title("Settings")
        top.geometry("450x200")
        top.configure(bg=BG_COLOR)
        top.transient(self)
        top.grab_set()
        
        tk.Label(top, text="⛭ Printer Settings", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L).pack(pady=10)
        
        frame = tk.Frame(top, bg=BG_COLOR)
        frame.pack(pady=10, fill=tk.X, padx=20)
        
        tk.Label(frame, text="Zebra Printer:", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(side=tk.LEFT)
        
        printers = []
        if WIN32_PRINT_AVAILABLE:
            try:
                # PRINTER_ENUM_LOCAL = 2, PRINTER_ENUM_CONNECTIONS = 4
                flags = 2 | 4
                printers = [p[2] for p in win32print.EnumPrinters(flags)]
            except Exception as e:
                print("Error listing printers:", e)
        
        cb = ttk.Combobox(frame, values=printers, state="readonly", width=35)
        cb.pack(side=tk.RIGHT)
        
        config = {}
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r") as f:
                    config = json.load(f)
            except: pass
            
        if config.get("zebra_printer") in printers:
            cb.set(config.get("zebra_printer"))
        elif printers:
            cb.set(printers[0])
            
        def save():
            config["zebra_printer"] = cb.get()
            with open(CONFIG_FILE, "w") as f:
                json.dump(config, f)
            top.destroy()
            messagebox.showinfo("Saved", "Settings saved successfully.")
            
        ttk.Button(top, text="✔ Save Settings", style="Success.TButton", command=save).pack(pady=20)

    def build_ui(self):
        HEADER_BG = "#FFFFFF"
        HEADER_FG = "#005A8C"
        
        self.header_frame = tk.Frame(self, bg=HEADER_BG, height=70)
        self.header_frame.pack(side=tk.TOP, fill=tk.X)
        self.header_frame.pack_propagate(False)
        
        try:
            logo_path = resource_path(os.path.join("assets", "logo_en.png"))
            if os.path.exists(logo_path):
                try:
                    from PIL import Image, ImageTk
                    img = Image.open(logo_path)
                    
                    ratio = 50 / img.height
                    new_size = (int(img.width * ratio), 50)
                    if hasattr(Image, "Resampling"):
                        res_filter = Image.Resampling.LANCZOS
                    else:
                        res_filter = Image.ANTIALIAS
                    img = img.resize(new_size, res_filter)
                    self.logo_img = ImageTk.PhotoImage(img)
                    tk.Label(self.header_frame, image=self.logo_img, bg=HEADER_BG).pack(side=tk.LEFT, padx=(20, 10))
                except Exception as pil_e:
                    self.logo_img = tk.PhotoImage(file=logo_path)
                    tk.Label(self.header_frame, image=self.logo_img, bg=HEADER_BG).pack(side=tk.LEFT, padx=(20, 10))
            
            tk.Label(self.header_frame, text="SUB-PROCESS TRACEABILITY", bg=HEADER_BG, fg=HEADER_FG, font=HMI_FONT_L).pack(side=tk.LEFT)
        except Exception as e:
            tk.Label(self.header_frame, text="⛭ HI-LEX ACT - SUB-PROCESS TRACEABILITY", bg=HEADER_BG, fg=HEADER_FG, font=HMI_FONT_L).pack(side=tk.LEFT, padx=20)
        
        self.settings_btn = ttk.Button(self.header_frame, text="⛭ Settings", style="Header.TButton", command=self.open_settings)
        self.settings_btn.pack(side=tk.RIGHT, padx=20)
        
        try:
            settings_path = resource_path(os.path.join("assets", "settings_icon.png"))
            if os.path.exists(settings_path):
                from PIL import Image, ImageTk
                s_img = Image.open(settings_path).convert("RGBA")
                data = s_img.getdata()
                new_data = []
                for item in data:
                    if item[3] > 10:
                        new_data.append((255, 255, 255, item[3]))
                    else:
                        new_data.append((255, 255, 255, 0))
                s_img.putdata(new_data)
                
                if hasattr(Image, "Resampling"):
                    res_filter = Image.Resampling.LANCZOS
                else:
                    res_filter = Image.ANTIALIAS
                s_img = s_img.resize((20, 20), res_filter)
                self.settings_icon = ImageTk.PhotoImage(s_img)
                self.settings_btn.config(image=self.settings_icon, compound=tk.LEFT, text=" Settings")
        except Exception as e:
            print("Could not load settings icon:", e)
        
        # PanedWindow
        self.paned = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        self.paned.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Sidebar
        self.sidebar = tk.Frame(self.paned, bg=SURFACE_COLOR, width=180)
        self.sidebar.pack_propagate(False)
        self.paned.add(self.sidebar, weight=0)
        
        self.lbl_clock = tk.Label(self.sidebar, text="00:00", bg=SURFACE_COLOR, fg=ACCENT_COLOR, font=("Consolas", 28, "bold"))
        self.lbl_clock.pack(pady=(20,0))
        self.lbl_date = tk.Label(self.sidebar, text="YYYY-MM-DD", bg=SURFACE_COLOR, fg=TEXT_MUTED, font=HMI_FONT_S)
        self.lbl_date.pack(pady=(0, 20))

        
        tk.Label(self.sidebar, text="SHIFT OVERVIEW", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack()
        
        def create_side_kpi(parent, title):
            f = tk.Frame(parent, bg=BG_COLOR, highlightbackground=BORDER_COLOR, highlightthickness=1)
            f.pack(fill=tk.X, padx=15, pady=5)
            lbl_title = tk.Label(f, text=title, bg=BG_COLOR, fg=TEXT_MUTED, font=HMI_FONT_S)
            lbl_title.pack(pady=(5,0))
            lbl_val = tk.Label(f, text="0", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L)
            lbl_val.pack(pady=(0,5))
            return lbl_title, lbl_val
            
        self.lbl_title_recs, self.side_stat_recs = create_side_kpi(self.sidebar, "Records (Shift -)")
        self.lbl_title_qty, self.side_stat_qty = create_side_kpi(self.sidebar, "Qty (Shift -)")
        _, self.side_stat_today = create_side_kpi(self.sidebar, "Total Today")
        
        tk.Label(self.sidebar, text="RECENT PNs", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=HMI_FONT_M).pack(pady=(10,0))
        self.recent_pns_listbox = tk.Listbox(self.sidebar, bg=BG_COLOR, fg=TEXT_COLOR, bd=0, relief="flat", height=8, font=HMI_FONT_S)
        self.recent_pns_listbox.pack(fill=tk.X, padx=15, pady=10)
        
        ttk.Button(self.sidebar, text="🗁 Open Excel", style="Secondary.TButton", command=self.open_excel).pack(fill=tk.X, padx=10, pady=2)
        ttk.Button(self.sidebar, text="⎙ Print Last Slip", style="Warning.TButton", command=self.print_last_slip).pack(fill=tk.X, padx=10, pady=2)
        
        # Notebook
        self.notebook = ttk.Notebook(self.paned)
        self.paned.add(self.notebook, weight=1)
        
        self.tab1 = ttk.Frame(self.notebook)
        self.tab4 = ttk.Frame(self.notebook)
        self.tab2 = ttk.Frame(self.notebook)
        self.tab3 = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab1, text="✚ New Entry")
        self.notebook.add(self.tab4, text="⎙ Print Label")
        self.notebook.add(self.tab2, text="🔍︎ Records")
        self.notebook.add(self.tab3, text="📈 KPIs")
        
        self.build_tab1()
        self.build_tab4()
        self.build_tab2()
        self.build_tab3()
        
    def build_tab1(self):
        # Container Split
        top_panel = tk.Frame(self.tab1, bg=BG_COLOR)
        top_panel.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        bottom_panel = tk.Frame(self.tab1, bg=BG_COLOR)
        bottom_panel.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=5)

        # 2-Column Layout
        self.t1_left = tk.Frame(top_panel, bg=BG_COLOR, width=400)
        self.t1_left.pack(side=tk.LEFT, fill=tk.Y, padx=10, pady=5)
        self.t1_left.pack_propagate(False) # Keep fixed width for left panel
        
        self.t1_right = tk.Frame(top_panel, bg=BG_COLOR)
        self.t1_right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=5)

        # ---------------- LEFT PANEL ----------------
        # QR Scan Frame
        _, scan_frame = self.create_card(self.t1_left, "⛶ Scan Label (QR Code)", fg_color=WARNING_COLOR)
        
        ttk.Label(scan_frame, text="Scan Label Here:").pack(side=tk.TOP, anchor="w", pady=2)
        self.var_scan_input = tk.StringVar()
        scan_entry = ttk.Entry(scan_frame, textvariable=self.var_scan_input, width=35)
        scan_entry.pack(side=tk.TOP, fill=tk.X, pady=2)
        scan_entry.bind("<Return>", self.on_scan_enter)
        
        ttk.Button(scan_frame, text="🗁 Upload QR Image", command=self.upload_qr_image).pack(side=tk.TOP, fill=tk.X, pady=5)

        # Quick Search
        _, search_frame = self.create_card(self.t1_left, "⌕ Quick Search")
        self.sv_search = tk.StringVar()
        self.sv_search.trace_add("write", self.filter_sf_combobox)
        search_entry = ttk.Entry(search_frame, textvariable=self.sv_search, width=30)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        self.lbl_search_count = tk.Label(search_frame, text="30", fg=ACCENT_COLOR, bg=BG_COLOR)
        self.lbl_search_count.pack(side=tk.LEFT, padx=5)
        ttk.Button(search_frame, text="✕", width=3, command=lambda: self.sv_search.set("")).pack(side=tk.LEFT)
        
        # Status Label
        self.lbl_status = tk.Label(self.t1_left, text="", bg=BG_COLOR, fg=SUCCESS_COLOR, font=HMI_FONT_M)
        self.lbl_status.pack(pady=10)
        
        btn_frame = tk.Frame(self.t1_left, bg=BG_COLOR)
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10, padx=10)
        ttk.Button(btn_frame, text="✔ Save Record", style="Success.TButton", command=self.save_record).pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        ttk.Button(btn_frame, text="⟳ Same as Last", style="Primary.TButton", command=self.same_as_last).pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        ttk.Button(btn_frame, text="↺ Clear Form", style="Secondary.TButton", command=self.clear_form).pack(side=tk.TOP, fill=tk.X, padx=10, pady=5)
        
        # ---------------- BOTTOM PANEL ----------------
        # Recent Records Tree
        _, tree_frame = self.create_card(bottom_panel, "📋 Last 5 Records (Double-click to print)")
        
        cols = ("SB_ID", "SF_PN", "Qty", "Shift", "Station", "DateTime")
        self.tree_recent = ttk.Treeview(tree_frame, columns=cols, show="headings", height=5)
        for c in cols:
            self.tree_recent.heading(c, text=c)
            self.tree_recent.column(c, width=100)
        self.tree_recent.pack(fill=tk.X)
        self.tree_recent.bind("<Double-1>", self.on_recent_double_click)
        
        self.tree_recent.tag_configure("even", background="#141A20")
        self.tree_recent.tag_configure("odd", background="#1B232C")
        

        # ---------------- RIGHT PANEL (FORM) ----------------
        self.canvas_t1 = tk.Canvas(self.t1_right, bg=BG_COLOR, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.t1_right, orient="vertical", command=self.canvas_t1.yview)
        self.form_frame = tk.Frame(self.canvas_t1, bg=BG_COLOR)
        
        # Make the form frame fill the canvas width
        self.t1_window = self.canvas_t1.create_window((0, 0), window=self.form_frame, anchor="nw")
        
        def resize_form(event):
            canvas_width = event.width
            self.canvas_t1.itemconfig(self.t1_window, width=canvas_width)
            self.canvas_t1.configure(scrollregion=self.canvas_t1.bbox("all"))
            
        self.canvas_t1.bind("<Configure>", resize_form)
        self.form_frame.bind("<Configure>", lambda e: self.canvas_t1.configure(scrollregion=self.canvas_t1.bbox("all")))
        
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas_t1.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        def _on_mousewheel(event):
            try:
                widget = self.winfo_containing(event.x_root, event.y_root)
                if widget and widget.winfo_class() in ("Listbox", "Treeview"):
                    return
                if self.notebook.index("current") == 0:
                    self.canvas_t1.yview_scroll(int(-1*(event.delta/120)), "units")
            except Exception:
                pass
        self.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Section 1: Part Reference
        _, lf1 = self.create_card(self.form_frame, "Part Reference")
        
        ttk.Label(lf1, text="FULL PN° Semi fini ★", width=32).grid(row=0, column=0, sticky="w", pady=2)
        self.cb_sf_pn = ttk.Combobox(lf1, state="readonly", width=30)
        self.cb_sf_pn.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        self.cb_sf_pn.bind("<<ComboboxSelected>>", self.on_sf_selected)
        
        ttk.Label(lf1, text="PART NAME (SF)").grid(row=0, column=2, sticky="w", pady=2, padx=10)
        self.var_part_sf = tk.StringVar()
        ttk.Entry(lf1, textvariable=self.var_part_sf, state="disabled", width=30).grid(row=0, column=3, sticky="w", pady=2)
        
        # Dynamic RM Container T1
        self.rm_container_t1 = tk.Frame(lf1, bg=SURFACE_COLOR)
        self.rm_container_t1.grid(row=1, column=0, columnspan=4, sticky="w", pady=5)
        self.rm_vars_t1 = []
        
        # Section 2: Batch & Qty
        _, lf2 = self.create_card(self.form_frame, "Batch Numbers & Quantity")
        
        ttk.Label(lf2, text="Batch No. 1 / 2 / 3", width=32).grid(row=0, column=0, sticky="w", pady=2)
        b_frame = tk.Frame(lf2, bg=SURFACE_COLOR)
        b_frame.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        self.var_b1 = tk.StringVar()
        self.var_b2 = tk.StringVar()
        self.var_b3 = tk.StringVar()
        self.var_b1.trace_add("write", lambda *args: self.var_b1.set(self.var_b1.get().upper()))
        self.var_b2.trace_add("write", lambda *args: self.var_b2.set(self.var_b2.get().upper()))
        self.var_b3.trace_add("write", lambda *args: self.var_b3.set(self.var_b3.get().upper()))
        ttk.Entry(b_frame, textvariable=self.var_b1, width=15).pack(side=tk.LEFT, padx=2)
        ttk.Entry(b_frame, textvariable=self.var_b2, width=15).pack(side=tk.LEFT, padx=2)
        ttk.Entry(b_frame, textvariable=self.var_b3, width=15).pack(side=tk.LEFT, padx=2)
        
        ttk.Label(lf2, text="Quantity ★", width=32).grid(row=1, column=0, sticky="w", pady=2)
        q_frame = tk.Frame(lf2, bg=SURFACE_COLOR)
        q_frame.grid(row=1, column=1, sticky="w", pady=2, padx=5)
        self.var_qty = tk.StringVar()
        qty_entry = ttk.Entry(q_frame, textvariable=self.var_qty, width=10, font=HMI_FONT_M)
        qty_entry.pack(side=tk.LEFT, padx=2)
        ttk.Label(q_frame, text="pcs").pack(side=tk.LEFT, padx=5)
        
        # Section 3: Operation Details
        _, lf3 = self.create_card(self.form_frame, "Operation Details")
        
        ttk.Label(lf3, text="Work in Sub-Process by Shift ★", width=32).grid(row=0, column=0, sticky="w", pady=2)
        self.cb_shift_sp = ttk.Combobox(lf3, values=["A", "B", "C"], state="readonly", width=5)
        self.cb_shift_sp.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        self.cb_shift_sp.set("")
        self.cb_shift_sp.bind("<<ComboboxSelected>>", lambda e: [self.update_sub_batch_preview(), self.update_stats()])
        
        ttk.Label(lf3, text="Op ID ★").grid(row=0, column=2, sticky="w", pady=2, padx=10)
        self.var_op_id = tk.StringVar()
        ttk.Entry(lf3, textvariable=self.var_op_id, width=15).grid(row=0, column=3, sticky="w", pady=2)
        
        ttk.Label(lf3, text="Station ★").grid(row=0, column=4, sticky="w", pady=2, padx=10)
        self.cb_station = ttk.Combobox(lf3, values=["S06", "S07", "S10", "S11"], state="readonly", width=5)
        self.cb_station.grid(row=0, column=5, sticky="w", pady=2)
        self.cb_station.bind("<<ComboboxSelected>>", lambda e: self.update_sub_batch_preview())
        
        # Section 4: Date & Time
        _, lf4 = self.create_card(self.form_frame, "Date & Time")
        
        def create_dt_picker(parent, label_text):
            frame = tk.Frame(parent, bg=SURFACE_COLOR)
            ttk.Label(frame, text=label_text, width=32).pack(side=tk.LEFT, padx=5)
            de = DateEntry(frame, width=12, background=ACCENT_COLOR, foreground='white', borderwidth=2)
            de.pack(side=tk.LEFT, padx=2)
            h_spin = tk.Spinbox(frame, from_=0, to=23, wrap=True, width=3, format="%02.0f", bg=SURFACE_COLOR, fg=TEXT_COLOR)
            h_spin.pack(side=tk.LEFT, padx=2)
            ttk.Label(frame, text=":").pack(side=tk.LEFT)
            m_spin = tk.Spinbox(frame, from_=0, to=59, wrap=True, width=3, format="%02.0f", bg=SURFACE_COLOR, fg=TEXT_COLOR)
            m_spin.pack(side=tk.LEFT, padx=2)
            
            def set_now():
                now = datetime.datetime.now()
                de.set_date(now.date())
                h_spin.delete(0, "end")
                h_spin.insert(0, f"{now.hour:02d}")
                m_spin.delete(0, "end")
                m_spin.insert(0, f"{now.minute:02d}")
                
            live_var = tk.BooleanVar(value=True)
            cb_live = ttk.Checkbutton(frame, text="Live", variable=live_var)
            cb_live.pack(side=tk.LEFT, padx=5)
            
            def auto_update():
                if live_var.get():
                    set_now()
                parent.after(1000, auto_update)
                
            def stop_live(*args): live_var.set(False)
            de.bind("<<DateEntrySelected>>", stop_live)
            h_spin.bind("<Button-1>", stop_live)
            h_spin.bind("<Key>", stop_live)
            m_spin.bind("<Button-1>", stop_live)
            m_spin.bind("<Key>", stop_live)
            
            auto_update()
            return frame, de, h_spin, m_spin

        self.f_dt_sp, self.de_sp, self.h_sp, self.m_sp = create_dt_picker(lf4, "Sub-Process Work Date/Time")
        self.f_dt_sp.grid(row=0, column=0, sticky="w", pady=5)
        
        self.f_dt_line, self.de_line, self.h_line, self.m_line = create_dt_picker(lf4, "Production Line Entry Date/Time")
        self.f_dt_line.grid(row=1, column=0, sticky="w", pady=5)
        
        # Section 5: Additional Info
        _, lf5 = self.create_card(self.form_frame, "Additional Info")
        
        ttk.Label(lf5, text="Work in PROD line by Shift", width=32).grid(row=0, column=0, sticky="w", pady=2)
        self.cb_shift_line = ttk.Combobox(lf5, values=["A", "B", "C"], state="readonly", width=5)
        self.cb_shift_line.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        
        ttk.Label(lf5, text="Remarks", width=32).grid(row=1, column=0, sticky="nw", pady=2)
        self.txt_remarks = tk.Text(lf5, height=3, width=50, bg=SURFACE_COLOR, fg=TEXT_COLOR)
        self.txt_remarks.grid(row=1, column=1, sticky="w", pady=2, padx=5)
        
    def build_tab4(self):
        # 3-Column Layout for True Centering
        self.t4_left = tk.Frame(self.tab4, bg=BG_COLOR)
        self.t4_left.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.t4_center = tk.Frame(self.tab4, bg=BG_COLOR, width=950)
        self.t4_center.pack(side=tk.LEFT, fill=tk.Y, pady=10)
        self.t4_center.pack_propagate(False)
        
        self.t4_right = tk.Frame(self.tab4, bg=BG_COLOR)
        self.t4_right.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Dedicated Print Label Form (No DB Saving)
        self.pl_frame = tk.Frame(self.t4_center, bg=BG_COLOR)
        self.pl_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        
        # Part Reference
        _, lf1 = self.create_card(self.pl_frame, "Part Reference")
        
        ttk.Label(lf1, text="FULL PN° Semi fini ★", width=32).grid(row=0, column=0, sticky="w", pady=2)
        self.pl_cb_sf_pn = ttk.Combobox(lf1, values=list(SF_DATA.keys()), state="readonly", width=30)
        self.pl_cb_sf_pn.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        self.pl_cb_sf_pn.bind("<<ComboboxSelected>>", self.pl_on_sf_selected)
        
        ttk.Label(lf1, text="PART NAME (SF)").grid(row=0, column=2, sticky="w", pady=2, padx=10)
        self.pl_var_part_sf = tk.StringVar()
        ttk.Entry(lf1, textvariable=self.pl_var_part_sf, state="disabled", width=30).grid(row=0, column=3, sticky="w", pady=2)
        
        # Dynamic RM Container T4
        self.rm_container_t4 = tk.Frame(lf1, bg=SURFACE_COLOR)
        self.rm_container_t4.grid(row=1, column=0, columnspan=4, sticky="w", pady=5)
        self.rm_vars_t4 = []
        
        # Batch & Qty
        _, lf2 = self.create_card(self.pl_frame, "Batch Numbers & Quantity")
        
        ttk.Label(lf2, text="Batch No. 1 / 2 / 3", width=32).grid(row=0, column=0, sticky="w", pady=2)
        b_frame = tk.Frame(lf2, bg=SURFACE_COLOR)
        b_frame.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        self.pl_var_b1 = tk.StringVar()
        self.pl_var_b2 = tk.StringVar()
        self.pl_var_b3 = tk.StringVar()
        self.pl_var_b1.trace_add("write", lambda *args: self.pl_var_b1.set(self.pl_var_b1.get().upper()))
        self.pl_var_b2.trace_add("write", lambda *args: self.pl_var_b2.set(self.pl_var_b2.get().upper()))
        self.pl_var_b3.trace_add("write", lambda *args: self.pl_var_b3.set(self.pl_var_b3.get().upper()))
        ttk.Entry(b_frame, textvariable=self.pl_var_b1, width=15).pack(side=tk.LEFT, padx=2)
        ttk.Entry(b_frame, textvariable=self.pl_var_b2, width=15).pack(side=tk.LEFT, padx=2)
        ttk.Entry(b_frame, textvariable=self.pl_var_b3, width=15).pack(side=tk.LEFT, padx=2)
        
        ttk.Label(lf2, text="Quantity ★", width=32).grid(row=1, column=0, sticky="w", pady=2)
        q_frame = tk.Frame(lf2, bg=SURFACE_COLOR)
        q_frame.grid(row=1, column=1, sticky="w", pady=2, padx=5)
        self.pl_var_qty = tk.StringVar()
        ttk.Entry(q_frame, textvariable=self.pl_var_qty, width=10, font=HMI_FONT_M).pack(side=tk.LEFT, padx=2)
        ttk.Label(q_frame, text="pcs").pack(side=tk.LEFT, padx=5)
        
        # Operations
        _, lf3 = self.create_card(self.pl_frame, "Operation Details")
        
        ttk.Label(lf3, text="Work in Sub-Process by Shift ★", width=32).grid(row=0, column=0, sticky="w", pady=2)
        self.pl_cb_shift_sp = ttk.Combobox(lf3, values=["A", "B", "C"], state="readonly", width=5)
        self.pl_cb_shift_sp.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        self.pl_cb_shift_sp.set("")
        
        ttk.Label(lf3, text="Op ID ★").grid(row=0, column=2, sticky="w", pady=2, padx=10)
        self.pl_var_op_id = tk.StringVar()
        ttk.Entry(lf3, textvariable=self.pl_var_op_id, width=15).grid(row=0, column=3, sticky="w", pady=2)
        
        ttk.Label(lf3, text="Station ★").grid(row=0, column=4, sticky="w", pady=2, padx=10)
        self.pl_cb_station = ttk.Combobox(lf3, values=["S06", "S07", "S10", "S11"], state="readonly", width=5)
        self.pl_cb_station.grid(row=0, column=5, sticky="w", pady=2)
        
        # Dates
        _, lf4 = self.create_card(self.pl_frame, "Date & Time")
        
        def create_pl_dt(parent, label_text):
            frame = tk.Frame(parent, bg=SURFACE_COLOR)
            ttk.Label(frame, text=label_text, width=32).pack(side=tk.LEFT, padx=5)
            de = DateEntry(frame, width=12, background=ACCENT_COLOR, foreground='white', borderwidth=2)
            de.pack(side=tk.LEFT, padx=2)
            h_spin = tk.Spinbox(frame, from_=0, to=23, wrap=True, width=3, format="%02.0f", bg=SURFACE_COLOR, fg=TEXT_COLOR)
            h_spin.pack(side=tk.LEFT, padx=2)
            ttk.Label(frame, text=":").pack(side=tk.LEFT)
            m_spin = tk.Spinbox(frame, from_=0, to=59, wrap=True, width=3, format="%02.0f", bg=SURFACE_COLOR, fg=TEXT_COLOR)
            m_spin.pack(side=tk.LEFT, padx=2)
            def set_now():
                now = datetime.datetime.now()
                de.set_date(now.date())
                h_spin.delete(0, "end"); h_spin.insert(0, f"{now.hour:02d}")
                m_spin.delete(0, "end"); m_spin.insert(0, f"{now.minute:02d}")
            
            live_var = tk.BooleanVar(value=True)
            cb_live = ttk.Checkbutton(frame, text="Live", variable=live_var)
            cb_live.pack(side=tk.LEFT, padx=5)
            
            def auto_update():
                if live_var.get(): set_now()
                parent.after(1000, auto_update)
                
            def stop_live(*args): live_var.set(False)
            de.bind("<<DateEntrySelected>>", stop_live)
            h_spin.bind("<Button-1>", stop_live); h_spin.bind("<Key>", stop_live)
            m_spin.bind("<Button-1>", stop_live); m_spin.bind("<Key>", stop_live)
            
            auto_update()
            return frame, de, h_spin, m_spin
            
        self.pl_f_sp, self.pl_de_sp, self.pl_h_sp, self.pl_m_sp = create_pl_dt(lf4, "Sub-Process Work Date/Time")
        self.pl_f_sp.grid(row=0, column=0, sticky="w", pady=5)
        self.pl_f_line, self.pl_de_line, self.pl_h_line, self.pl_m_line = create_pl_dt(lf4, "Production Line Entry Date/Time")
        self.pl_f_line.grid(row=1, column=0, sticky="w", pady=5)
        
        # Add info
        _, lf5 = self.create_card(self.pl_frame, "Additional Info")
        ttk.Label(lf5, text="Work in PROD line by Shift", width=32).grid(row=0, column=0, sticky="w", pady=2)
        self.pl_cb_shift_line = ttk.Combobox(lf5, values=["A", "B", "C"], state="readonly", width=5)
        self.pl_cb_shift_line.grid(row=0, column=1, sticky="w", pady=2, padx=5)
        ttk.Label(lf5, text="Remarks", width=32).grid(row=1, column=0, sticky="nw", pady=2)
        self.pl_txt_remarks = tk.Text(lf5, height=3, width=50, bg=SURFACE_COLOR, fg=TEXT_COLOR)
        self.pl_txt_remarks.grid(row=1, column=1, sticky="w", pady=2, padx=5)
        
        btn_frame = tk.Frame(self.t4_center, bg=BG_COLOR)
        btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)
        ttk.Button(btn_frame, text="🖨 Generate & Print Label", style="Warning.TButton", command=self.print_standalone_label).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="↺ Clear", style="Secondary.TButton", command=self.clear_pl_form).pack(side=tk.LEFT, padx=5)

    def build_tab2(self):
        filter_frame = tk.Frame(self.tab2, bg=SURFACE_COLOR)
        filter_frame.pack(fill=tk.X, padx=10, pady=10)
        
        ttk.Label(filter_frame, text="Search:").pack(side=tk.LEFT, padx=2)
        self.var_rec_search = tk.StringVar()
        ttk.Entry(filter_frame, textvariable=self.var_rec_search, width=20).pack(side=tk.LEFT, padx=2)
        
        ttk.Label(filter_frame, text="Shift:").pack(side=tk.LEFT, padx=(10,2))
        self.cb_rec_shift = ttk.Combobox(filter_frame, values=["All", "A", "B", "C"], state="readonly", width=5)
        self.cb_rec_shift.set("All")
        self.cb_rec_shift.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(filter_frame, text="Station:").pack(side=tk.LEFT, padx=(10,2))
        self.cb_rec_station = ttk.Combobox(filter_frame, values=["All", "S06", "S07", "S10", "S11"], state="readonly", width=5)
        self.cb_rec_station.set("All")
        self.cb_rec_station.pack(side=tk.LEFT, padx=2)
        
        ttk.Button(filter_frame, text="Search", style="Primary.TButton", command=self.refresh_records_treeview).pack(side=tk.LEFT, padx=10)
        ttk.Button(filter_frame, text="⟳ All", style="Secondary.TButton", command=self.reset_records_filter).pack(side=tk.LEFT, padx=2)
        ttk.Button(filter_frame, text="🖨 Print Selected", style="Warning.TButton", command=self.print_selected_record).pack(side=tk.RIGHT, padx=5)
        
        self.lbl_rec_count = tk.Label(filter_frame, text="0 records", bg=SURFACE_COLOR, fg=TEXT_COLOR)
        self.lbl_rec_count.pack(side=tk.RIGHT, padx=10)
        
        tree_frame = tk.Frame(self.tab2, bg=BG_COLOR)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        
        cols = ("SB_ID", "SF_PN", "SF_Name", "Qty", "Shift", "Station", "DateTime", "Status")
        self.tree_records = ttk.Treeview(tree_frame, columns=cols, show="headings")
        for c in cols:
            self.tree_records.heading(c, text=c, command=lambda _c=c: self.sort_treeview(self.tree_records, _c, False))
            self.tree_records.column(c, width=100)
            
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_records.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree_records.xview)
        self.tree_records.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree_records.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)
        
        self.tree_records.tag_configure("pending", background=WARNING_COLOR, foreground="black")
        self.tree_records.tag_configure("even", background="#141A20")
        self.tree_records.tag_configure("odd", background="#1B232C")
        self.tree_records.bind("<Double-1>", lambda e: self.print_selected_record())

    def get_dashboard_data(self):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        
        now = datetime.datetime.now()
        today_date = now.date()
        if now.hour < 6: today_date -= datetime.timedelta(days=1)
        
        start_today_str = f"{today_date.strftime('%Y-%m-%d')} 06:00:00"
        
        start_week_date = today_date - datetime.timedelta(days=today_date.weekday())
        start_week_str = f"{start_week_date.strftime('%Y-%m-%d')} 06:00:00"
        
        start_month_date = today_date.replace(day=1)
        start_month_str = f"{start_month_date.strftime('%Y-%m-%d')} 06:00:00"
        
        def get_stats(start_dt):
            c.execute("SELECT shift_sp, SUM(quantity) FROM records WHERE created_at >= ? GROUP BY shift_sp", (start_dt,))
            rows = c.fetchall()
            stats = {'A': 0, 'B': 0, 'C': 0}
            for r in rows:
                if r[0] in stats:
                    stats[r[0]] = r[1] if r[1] else 0
            total = sum(stats.values())
            return total, stats['A'], stats['B'], stats['C']
            
        today_stats = get_stats(start_today_str)
        week_stats = get_stats(start_week_str)
        month_stats = get_stats(start_month_str)
        
        conn.close()
        return today_stats, week_stats, month_stats

    def build_tab3(self):
        for widget in self.tab3.winfo_children():
            widget.destroy()
            
        header = tk.Frame(self.tab3, bg=BG_COLOR)
        header.pack(fill=tk.X, pady=10)
        tk.Label(header, text="📊 Production KPIs", bg=BG_COLOR, fg=TEXT_COLOR, font=HMI_FONT_L).pack(side=tk.LEFT, padx=20)
        ttk.Button(header, text="⟳ Refresh Data", command=self.build_tab3).pack(side=tk.RIGHT, padx=20)
        
        content = tk.Frame(self.tab3, bg=BG_COLOR)
        content.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        try:
            today_stats, week_stats, month_stats = self.get_dashboard_data()
        except Exception as e:
            tk.Label(content, text=f"Error loading data: {e}", fg="red", bg=BG_COLOR).pack()
            return

        def create_kpi_card(parent, title, stats):
            total, a, b, c = stats
            card = tk.Frame(parent, bg=SURFACE_COLOR, bd=2, relief="flat", highlightbackground=BORDER_COLOR, highlightthickness=1)
            card.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)
            
            tk.Label(card, text=title, bg=SURFACE_COLOR, fg=WARNING_COLOR, font=HMI_FONT_L).pack(pady=10)
            tk.Label(card, text=f"{total:,} pcs", bg=SURFACE_COLOR, fg=TEXT_COLOR, font=("Helvetica", 36, "bold")).pack(pady=5)
            
            c_width = 300
            c_height = 200
            canvas = tk.Canvas(card, width=c_width, height=c_height, bg=SURFACE_COLOR, highlightthickness=0)
            canvas.pack(pady=20)
            
            max_val = max(1, a, b, c)
            shifts = [("Shift A", a, STATUS_RUNNING), ("Shift B", b, SUCCESS_COLOR), ("Shift C", c, ERROR_COLOR)]
            
            bar_w = 50
            spacing = 30
            start_x = (c_width - (3*bar_w + 2*spacing)) / 2
            
            for i, (name, val, color) in enumerate(shifts):
                x0 = start_x + i*(bar_w + spacing)
                x1 = x0 + bar_w
                h = (val / max_val) * (c_height - 50)
                y0 = c_height - 30 - h
                y1 = c_height - 30
                
                canvas.create_rectangle(x0, y0, x1, y1, fill=color, outline="")
                canvas.create_text(x0 + bar_w/2, y0 - 15, text=str(val), fill=TEXT_COLOR, font=HMI_FONT_S)
                canvas.create_text(x0 + bar_w/2, c_height - 10, text=name, fill=TEXT_MUTED, font=("Helvetica", 10))
                
        cards_frame = tk.Frame(content, bg=BG_COLOR)
        cards_frame.pack(fill=tk.X, pady=20)
        
        create_kpi_card(cards_frame, "Today", today_stats)
        create_kpi_card(cards_frame, "This Week", week_stats)
        create_kpi_card(cards_frame, "This Month", month_stats)

    def update_clock(self):
        now = datetime.datetime.now()
        self.lbl_clock.config(text=now.strftime("%H:%M:%S"))
        self.lbl_date.config(text=now.strftime("%Y-%m-%d"))
        self.after(1000, self.update_clock)

    def update_stats(self):
        now = datetime.datetime.now()
        # Production day starts at 06:00 and ends at 05:59 the next day
        if now.hour < 6:
            start_dt = (now - datetime.timedelta(days=1)).strftime("%Y-%m-%d 06:00:00")
            end_dt = now.strftime("%Y-%m-%d 05:59:59")
        else:
            start_dt = now.strftime("%Y-%m-%d 06:00:00")
            end_dt = (now + datetime.timedelta(days=1)).strftime("%Y-%m-%d 05:59:59")
            
        cur_shift = getattr(self, 'cb_shift_sp', None)
        if cur_shift and cur_shift.get():
            shift_val = cur_shift.get()
        else:
            shift_val = "-"
        
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        
        # Shift stats
        c.execute("SELECT COUNT(*), SUM(quantity) FROM records WHERE created_at >= ? AND created_at <= ? AND shift_sp = ?", (start_dt, end_dt, shift_val))
        res_shift = c.fetchone()
        shift_count = res_shift[0] if res_shift[0] else 0
        shift_qty = res_shift[1] if res_shift[1] else 0
        
        # Today stats
        c.execute("SELECT COUNT(*) FROM records WHERE created_at >= ? AND created_at <= ?", (start_dt, end_dt))
        today_count = c.fetchone()[0]
        
        conn.close()
        

        
        self.lbl_title_recs.config(text=f"Records (Shift {shift_val})")
        self.side_stat_recs.config(text=str(shift_count))
        self.lbl_title_qty.config(text=f"Qty (Shift {shift_val})")
        self.side_stat_qty.config(text=str(shift_qty))
        self.side_stat_today.config(text=str(today_count))

    def populate_sf_combobox(self):
        sf_list = list(SF_DATA.keys())
        self.cb_sf_pn['values'] = sf_list
        self.lbl_search_count.config(text=f"{len(sf_list)} matches", fg=SUCCESS_COLOR)

    def filter_sf_combobox(self, *args):
        search_term = self.sv_search.get().lower()
        filtered = [pn for pn in SF_DATA.keys() if search_term in pn.lower()]
        self.cb_sf_pn['values'] = filtered
        if filtered:
            self.lbl_search_count.config(text=f"{len(filtered)} matches", fg=SUCCESS_COLOR)
            self.cb_sf_pn.set(filtered[0])
            self.on_sf_selected(None)
        else:
            self.lbl_search_count.config(text="0 matches", fg="red")
            self.cb_sf_pn.set("")

    def on_sf_selected(self, event):
        sf_pn = self.cb_sf_pn.get()
        for widget in getattr(self, 'rm_container_t1', tk.Frame()).winfo_children():
            widget.destroy()
        self.rm_vars_t1 = []

        if sf_pn in SF_DATA:
            sf_name, rm_list = SF_DATA[sf_pn]
            self.var_part_sf.set(sf_name)
            
            for idx, (rm_id, rm_name) in enumerate(rm_list):
                ttk.Label(self.rm_container_t1, text=f"RM {idx+1} Ref ★", font=HMI_FONT_S).grid(row=idx, column=0, sticky="w", padx=5, pady=2)
                cb_var = tk.StringVar(value=rm_id)
                cb = ttk.Entry(self.rm_container_t1, textvariable=cb_var, state="readonly", width=25)
                cb.grid(row=idx, column=1, sticky="w", padx=5, pady=2)
                
                ttk.Label(self.rm_container_t1, text=f"RM {idx+1} Name", font=HMI_FONT_S).grid(row=idx, column=2, sticky="w", padx=15, pady=2)
                name_var = tk.StringVar(value=rm_name)
                ttk.Entry(self.rm_container_t1, textvariable=name_var, state="readonly", width=25).grid(row=idx, column=3, sticky="w", padx=5, pady=2)
                
                self.rm_vars_t1.append((cb_var, name_var))

    def update_sub_batch_preview(self):
        pass

    def get_dt_string(self, de, h, m):
        d = de.get_date()
        return f"{d.strftime('%Y-%m-%d')} {h.get()}:{m.get()}"

    def clear_form(self):
        self.cb_sf_pn.set("")
        self.var_part_sf.set("")
        for widget in getattr(self, "rm_container_t1", tk.Frame()).winfo_children():
            widget.destroy()
        self.rm_vars_t1 = []
        self.var_b1.set("")
        self.var_b2.set("")
        self.var_b3.set("")
        self.var_qty.set("")
        self.cb_shift_sp.set("")
        self.var_op_id.set("")
        self.cb_station.set("")
        
        now = datetime.datetime.now()
        self.de_sp.set_date(now.date())
        self.h_sp.delete(0, "end"); self.h_sp.insert(0, f"{now.hour:02d}")
        self.m_sp.delete(0, "end"); self.m_sp.insert(0, f"{now.minute:02d}")
        
        self.de_line.set_date(now.date())
        self.h_line.delete(0, "end"); self.h_line.insert(0, f"{now.hour:02d}")
        self.m_line.delete(0, "end"); self.m_line.insert(0, f"{now.minute:02d}")
        
        self.cb_shift_line.set("")
        self.txt_remarks.delete("1.0", tk.END)
        self.update_sub_batch_preview()

    def same_as_last(self):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT * FROM records ORDER BY id DESC LIMIT 1")
        rec = c.fetchone()
        conn.close()
        if rec:
            self.cb_sf_pn.set(rec[2])
            self.on_sf_selected(None)
            self.var_qty.set(rec[9])
            self.cb_shift_sp.set(rec[10])
            self.cb_station.set(rec[12])
            self.update_sub_batch_preview()
            self.lbl_status.config(text="Pre-filled with last record data.", fg=SUCCESS_COLOR)
            self.after(3000, lambda: self.lbl_status.config(text=""))

    def save_record(self):
        sf_pn = self.cb_sf_pn.get()
        qty_str = self.var_qty.get()
        op_id = self.var_op_id.get()
        station = self.cb_station.get()
        shift_sp = self.cb_shift_sp.get()
        
        if not all([sf_pn, qty_str, op_id, station, shift_sp]) or not self.rm_vars_t1:
            messagebox.showerror("Error", "Please fill all required fields (★).")
            return
            
        if not any([self.var_b1.get().strip(), self.var_b2.get().strip(), self.var_b3.get().strip()]):
            messagebox.showerror("Error", "Please provide at least one Batch Number.")
            return
            
        rm_pns = [""] * 4
        rm_names = [""] * 4
        for idx, (cb_var, name_var) in enumerate(self.rm_vars_t1):
            if not cb_var.get():
                messagebox.showerror("Error", "Please fill all RM Reference fields.")
                return
            if idx < 4:
                rm_pns[idx] = cb_var.get()
                rm_names[idx] = name_var.get()

        try:
            qty = int(qty_str)
        except ValueError:
            messagebox.showerror("Error", "Quantity must be a valid integer.")
            return

        shift_sp = self.cb_shift_sp.get()
        dt_sp = self.get_dt_string(self.de_sp, self.h_sp, self.m_sp)
        dt_line = f"{self.de_line.get_date()} {self.h_line.get()}:{self.m_line.get()}"
        
        sb_id = f"SB{dt_sp.replace('-', '').replace(':', '').replace(' ', '')}{station}{shift_sp}"

        created_at = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        data = (
            sb_id, sf_pn, self.var_part_sf.get(), 
            rm_pns[0], rm_names[0], rm_pns[1], rm_names[1],
            rm_pns[2], rm_names[2], rm_pns[3], rm_names[3],
            self.var_b1.get(), self.var_b2.get(), self.var_b3.get(), qty,
            shift_sp, op_id, station, dt_sp, dt_line,
            self.cb_shift_line.get(), self.txt_remarks.get("1.0", tk.END).strip(),
            'pending', created_at
        )

        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute('''
            INSERT INTO records (
                sub_batch_id, pn_sf, part_sf, 
                rm1_pn, rm1_name, rm2_pn, rm2_name,
                rm3_pn, rm3_name, rm4_pn, rm4_name,
                batch1, batch2, batch3, quantity,
                shift_sp, op_id, station, dt_sp, dt_line,
                shift_line, remarks, status, created_at
            ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', data)
        conn.commit()
        conn.close()
        
        excel_data = [
            sf_pn, self.var_part_sf.get(), 
            rm_pns[0], rm_names[0], rm_pns[1], rm_names[1],
            rm_pns[2], rm_names[2], rm_pns[3], rm_names[3],
            self.var_b1.get(), self.var_b2.get(), self.var_b3.get(), qty,
            shift_sp, op_id, station, dt_sp, dt_line,
            self.cb_shift_line.get(), self.txt_remarks.get("1.0", tk.END).strip()
        ]
        
        try:
            save_to_excel(excel_data)
        except Exception as e:
            messagebox.showwarning("Excel Error", f"Failed to write to Excel:\n{e}")

        self.lbl_status.config(text=f"Saved successfully: {sb_id}", fg=SUCCESS_COLOR)
        self.after(7000, lambda: self.lbl_status.config(text=""))
        
        self.update_stats()
        self.refresh_recent_treeview()
        self.refresh_records_treeview()
        self.update_sub_batch_preview()
        
        # update recent PNs sidebar
        pns = self.recent_pns_listbox.get(0, tk.END)
        if sf_pn not in pns:
            self.recent_pns_listbox.insert(0, sf_pn)
            if self.recent_pns_listbox.size() > 8:
                self.recent_pns_listbox.delete(tk.END)
                
        self.clear_form()

    def refresh_recent_treeview(self):
        for item in self.tree_recent.get_children():
            self.tree_recent.delete(item)
            
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT sub_batch_id, pn_sf, quantity, shift_sp, station, dt_sp FROM records ORDER BY id DESC LIMIT 5")
        rows = c.fetchall()
        conn.close()
        
        for i, row in enumerate(rows):
            tag = "even" if i % 2 == 0 else "odd"
            self.tree_recent.insert("", "end", values=row, tags=(tag,))

    def refresh_records_treeview(self):
        for item in self.tree_records.get_children():
            self.tree_records.delete(item)
            
        query = "SELECT sub_batch_id, pn_sf, part_sf, quantity, shift_sp, station, dt_sp, status FROM records WHERE 1=1"
        params = []
        
        search = self.var_rec_search.get()
        if search:
            query += " AND (sub_batch_id LIKE ? OR pn_sf LIKE ? OR part_sf LIKE ?)"
            params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])
            
        shift = self.cb_rec_shift.get()
        if shift != "All":
            query += " AND shift_sp = ?"
            params.append(shift)
            
        station = self.cb_rec_station.get()
        if station != "All":
            query += " AND station = ?"
            params.append(station)
            
        query += " ORDER BY id DESC"
        
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute(query, params)
        rows = c.fetchall()
        conn.close()
        
        is_filtered = bool(search) or shift != "All" or station != "All"
        total_count = len(rows)
        
        if not is_filtered:
            limit = total_count % 10
            if limit == 0 and total_count > 0:
                limit = 10
            display_rows = rows[:limit] if total_count > 0 else []
        else:
            display_rows = rows
        
        for i, row in enumerate(display_rows):
            tags = ()
            if row[7] == 'pending':
                tags = ("pending",)
            else:
                tags = ("even" if i % 2 == 0 else "odd",)
            self.tree_records.insert("", "end", values=row, tags=tags)
            
        self.lbl_rec_count.config(text=f"{len(display_rows)} records displayed (Total: {total_count})")

    def reset_records_filter(self):
        self.var_rec_search.set("")
        self.cb_rec_shift.set("All")
        self.cb_rec_station.set("All")
        self.refresh_records_treeview()

    def sort_treeview(self, tree, col, reverse):
        l = [(tree.set(k, col), k) for k in tree.get_children('')]
        try:
            l.sort(key=lambda t: int(t[0]), reverse=reverse)
        except ValueError:
            l.sort(reverse=reverse)
            
        for index, (val, k) in enumerate(l):
            tree.move(k, '', index)
        tree.heading(col, command=lambda: self.sort_treeview(tree, col, not reverse))

    def on_recent_double_click(self, event):
        item = self.tree_recent.selection()
        if not item: return
        sb_id = self.tree_recent.item(item[0], "values")[0]
        self.do_print(sb_id)

    def print_selected_record(self):
        item = self.tree_records.selection()
        if not item: return
        sb_id = self.tree_records.item(item[0], "values")[0]
        self.do_print(sb_id)

    def print_last_slip(self):
        conn = sqlite3.connect(DB_FILE)
        c = conn.cursor()
        c.execute("SELECT sub_batch_id FROM records ORDER BY id DESC LIMIT 1")
        row = c.fetchone()
        conn.close()
        if row:
            self.do_print(row[0])
        else:
            messagebox.showinfo("Info", "No records found.")

    def do_print(self, sb_id):
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM records WHERE sub_batch_id=?", (sb_id,))
        row = c.fetchone()
        conn.close()
        
        if row:
            print_html_slip(dict(row))
        else:
            messagebox.showerror("Error", f"Record not found: {sb_id}")

    def on_scan_enter(self, event):
        qr_data = self.var_scan_input.get()
        if qr_data:
            self.process_qr_data(qr_data)
            self.var_scan_input.set("")

    def upload_qr_image(self):
        file_path = filedialog.askopenfilename(title="Select QR Image", filetypes=[("Image Files", "*.png *.jpg *.jpeg *.bmp")])
        if file_path:
            img = cv2.imread(file_path)
            detector = cv2.QRCodeDetector()
            data, bbox, _ = detector.detectAndDecode(img)
            if data:
                self.process_qr_data(data)
            else:
                messagebox.showerror("Error", "Could not decode any QR code from the selected image.")

    def process_qr_data(self, json_str):
        try:
            data = json.loads(json_str)
        except json.JSONDecodeError:
            messagebox.showerror("Error", "Invalid QR code format. Must be valid JSON.")
            return

        rm_pn = data.get("rm1_pn", "")
        if not rm_pn:
            messagebox.showerror("Error", "QR data does not contain 'rm1_pn'.")
            return
            
        rm_pn_list = [
            data.get("rm1_pn", ""), data.get("rm2_pn", ""),
            data.get("rm3_pn", ""), data.get("rm4_pn", "")
        ]
        rm_pn_list = [r.strip() for r in rm_pn_list if r and r.strip()]

        # Filter SF dropdown based on the RM PN
        compatible_sfs = []
        for sf, info in SF_DATA.items():
            rm_list = info[1]
            if any(rm[0] in rm_pn_list for rm in rm_list):
                compatible_sfs.append(sf)

        target_sf = data.get("full_pn_sf", "")
        
        if not compatible_sfs:
            messagebox.showwarning("Warning", f"RM PN '{rm_pn}' is not found in any known SF parts.")
            self.cb_sf_pn.set("")
            self.var_part_sf.set("")
            self.cb_sf_pn['values'] = list(SF_DATA.keys())
        else:
            self.cb_sf_pn['values'] = compatible_sfs
            if target_sf and target_sf in compatible_sfs:
                self.cb_sf_pn.set(target_sf)
            else:
                self.cb_sf_pn.set(compatible_sfs[0])
            self.on_sf_selected(None)

        self.var_b1.set(data.get("batch_no_1", ""))
        self.var_b2.set(data.get("batch_no_2", ""))
        self.var_b3.set(data.get("batch_no_3", ""))
        
        self.var_qty.set(str(data.get("quantity", "")))
        
        shift_sp = data.get("sub_process_shift", "")
        if shift_sp in ["A", "B", "C"]:
            self.cb_shift_sp.set(shift_sp)
            
        self.var_op_id.set(data.get("op_id", ""))
        
        station = data.get("station", "")
        if station in ["S06", "S07", "S10", "S11"]:
            self.cb_station.set(station)
            
        dt_sp_str = data.get("sub_process_datetime", "")
        if dt_sp_str:
            try:
                dt_sp = datetime.datetime.strptime(dt_sp_str, "%Y-%m-%d %H:%M:%S")
                self.de_sp.set_date(dt_sp.date())
                self.h_sp.delete(0, "end"); self.h_sp.insert(0, f"{dt_sp.hour:02d}")
                self.m_sp.delete(0, "end"); self.m_sp.insert(0, f"{dt_sp.minute:02d}")
            except ValueError:
                pass
                
        dt_line_str = data.get("production_line_entry_datetime", "")
        if dt_line_str:
            try:
                dt_line = datetime.datetime.strptime(dt_line_str, "%Y-%m-%d %H:%M:%S")
                self.de_line.set_date(dt_line.date())
                self.h_line.delete(0, "end"); self.h_line.insert(0, f"{dt_line.hour:02d}")
                self.m_line.delete(0, "end"); self.m_line.insert(0, f"{dt_line.minute:02d}")
            except ValueError:
                pass
                
        shift_line = data.get("production_line_shift", "")
        if shift_line in ["A", "B", "C"]:
            self.cb_shift_line.set(shift_line)
            
        self.txt_remarks.delete("1.0", tk.END)
        self.txt_remarks.insert(tk.END, data.get("remarks", ""))

        self.update_sub_batch_preview()
        
        # Highlight success
        self.lbl_status.config(text="QR Data Loaded Successfully!", fg=SUCCESS_COLOR)
        self.after(3000, lambda: self.lbl_status.config(text=""))

    def pl_on_sf_selected(self, event):
        sf_pn = self.pl_cb_sf_pn.get()
        for widget in getattr(self, 'rm_container_t4', tk.Frame()).winfo_children():
            widget.destroy()
        self.rm_vars_t4 = []

        if sf_pn in SF_DATA:
            sf_name, rm_list = SF_DATA[sf_pn]
            self.pl_var_part_sf.set(sf_name)
            
            for idx, (rm_id, rm_name) in enumerate(rm_list):
                ttk.Label(self.rm_container_t4, text=f"RM {idx+1} Ref ★", font=HMI_FONT_S).grid(row=idx, column=0, sticky="w", padx=5, pady=2)
                cb_var = tk.StringVar(value=rm_id)
                cb = ttk.Entry(self.rm_container_t4, textvariable=cb_var, state="readonly", width=25)
                cb.grid(row=idx, column=1, sticky="w", padx=5, pady=2)
                
                ttk.Label(self.rm_container_t4, text=f"RM {idx+1} Name", font=HMI_FONT_S).grid(row=idx, column=2, sticky="w", padx=15, pady=2)
                name_var = tk.StringVar(value=rm_name)
                ttk.Entry(self.rm_container_t4, textvariable=name_var, state="readonly", width=25).grid(row=idx, column=3, sticky="w", padx=5, pady=2)
                
                self.rm_vars_t4.append((cb_var, name_var))

    def clear_pl_form(self):
        self.pl_cb_sf_pn.set("")
        self.pl_var_part_sf.set("")
        for widget in getattr(self, "rm_container_t4", tk.Frame()).winfo_children():
            widget.destroy()
        self.rm_vars_t4 = []
        self.pl_var_b1.set("")
        self.pl_var_b2.set("")
        self.pl_var_b3.set("")
        self.pl_var_qty.set("")
        self.pl_var_op_id.set("")
        self.pl_cb_station.set("")
        self.pl_cb_shift_sp.set("")
        now = datetime.datetime.now()
        self.pl_de_sp.set_date(now.date())
        self.pl_h_sp.delete(0, "end"); self.pl_h_sp.insert(0, f"{now.hour:02d}")
        self.pl_m_sp.delete(0, "end"); self.pl_m_sp.insert(0, f"{now.minute:02d}")
        self.pl_de_line.set_date(now.date())
        self.pl_h_line.delete(0, "end"); self.pl_h_line.insert(0, f"{now.hour:02d}")
        self.pl_m_line.delete(0, "end"); self.pl_m_line.insert(0, f"{now.minute:02d}")
        self.pl_cb_shift_line.set("")
        self.pl_txt_remarks.delete("1.0", tk.END)

    def print_standalone_label(self):
        sf_pn = self.pl_cb_sf_pn.get()
        qty_str = self.pl_var_qty.get()
        op_id = self.pl_var_op_id.get()
        station = self.pl_cb_station.get()
        shift_sp = self.pl_cb_shift_sp.get()
        
        if not all([sf_pn, qty_str, op_id, station, shift_sp]) or not self.rm_vars_t4:
            messagebox.showerror("Error", "Please fill all required fields (★).")
            return
            
        if not any([self.pl_var_b1.get().strip(), self.pl_var_b2.get().strip(), self.pl_var_b3.get().strip()]):
            messagebox.showerror("Error", "Please provide at least one Batch Number.")
            return
            
        rm_pns = [""] * 4
        rm_names = [""] * 4
        for idx, (cb_var, name_var) in enumerate(self.rm_vars_t4):
            if not cb_var.get():
                messagebox.showerror("Error", "Please fill all RM Reference fields.")
                return
            if idx < 4:
                rm_pns[idx] = cb_var.get()
                rm_names[idx] = name_var.get()

        dt_sp = f"{self.pl_de_sp.get_date()} {self.pl_h_sp.get()}:{self.pl_m_sp.get()}:00"
        dt_line = f"{self.pl_de_line.get_date()} {self.pl_h_line.get()}:{self.pl_m_line.get()}:00"
        
        try:
            qty = int(qty_str)
        except ValueError:
            messagebox.showerror("Error", "Quantity must be a valid integer.")
            return

        sb_id = f"PL{dt_sp.replace('-', '').replace(':', '').replace(' ', '')}{self.pl_cb_station.get()}{self.pl_cb_shift_sp.get()}"

        record_data = {
            'sub_batch_id': sb_id,
            'pn_sf': sf_pn,
            'part_sf': self.pl_var_part_sf.get(),
            'rm1_pn': rm_pns[0],
            'rm1_name': rm_names[0],
            'rm2_pn': rm_pns[1],
            'rm2_name': rm_names[1],
            'rm3_pn': rm_pns[2],
            'rm3_name': rm_names[2],
            'rm4_pn': rm_pns[3],
            'rm4_name': rm_names[3],
            'batch1': self.pl_var_b1.get(),
            'batch2': self.pl_var_b2.get(),
            'batch3': self.pl_var_b3.get(),
            'quantity': qty,
            'shift_sp': self.pl_cb_shift_sp.get(),
            'op_id': self.pl_var_op_id.get(),
            'station': self.pl_cb_station.get(),
            'dt_sp': dt_sp,
            'dt_line': dt_line,
            'shift_line': self.pl_cb_shift_line.get(),
            'remarks': self.pl_txt_remarks.get("1.0", "end-1c"),
            'status': 'Standalone Print'
        }
        
        print_html_slip(record_data)

    def open_excel(self):
        if os.path.exists(EXCEL_FILE):
            os.startfile(EXCEL_FILE)
        else:
            messagebox.showinfo("Info", "Excel file not created yet. Save a record first.")

    def on_closing(self):
        if messagebox.askyesno("Confirm Exit", "Are you sure you want to close the application?"):
            self.destroy()

if __name__ == "__main__":
    init_db()
    app = TraceabilityApp()
    app.mainloop()
